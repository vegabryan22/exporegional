import os
import re
import secrets
import uuid
import json
import subprocess
import base64
import hmac
import shutil
from html import escape
from io import BytesIO
from datetime import datetime
from pathlib import Path

from functools import wraps

from flask import abort, current_app, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required
from sqlalchemy import or_, text
from sqlalchemy.engine import make_url
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models.assignment import Assignment
from app.models.campaign import Campaign
from app.models.project_document_revision import ProjectDocumentRevision
from app.models.category import Category
from app.models.evaluation import Evaluation
from app.models.evaluation_score import EvaluationScore
from app.models.evaluation_type import EvaluationType
from app.models.judge import Judge
from app.models.level import Level
from app.models.project import Project
from app.models.project_member_change import ProjectMemberChange
from app.models.project_member import ProjectMember
from app.models.project_member_edit_request import ProjectMemberEditRequest
from app.models.project_type import ProjectType
from app.models.rubric_criterion import RubricCriterion
from app.models.section import Section
from app.models.specialty import Specialty
from app.models.system_audit_log import SystemAuditLog
from app.models.system_setting import SystemSetting
from app.models.thematic_axis import ThematicAxis
from app.models.workshop import Workshop
from app.services.audit_service import log_event
from app.services.evaluation_service import (
    ENGLISH_EVAL_TYPE_CODE,
    assignment_allows_evaluation_type,
    build_admin_evaluation_overview,
    get_project_available_evaluation_types,
    infer_evaluation_type_kind,
)
from app.services.mail_service import send_email, smtp_is_configured

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}
ALLOWED_PROJECT_DOCUMENT_EXTENSIONS = {"pdf"}
LOGISTICS_STATUSES = [
    ("pendiente_revision", "Revision"),
    ("completo", "Completo"),
    ("incompleto", "Incompleto"),
]
USER_DEPARTMENTS = [
    ("logistica", "Logistica"),
    ("datos", "Datos"),
    ("diseno", "Diseno"),
    ("qa", "QA"),
]
USER_ROLES = [
    (Judge.ROLE_JUDGE, "Juez"),
    (Judge.ROLE_ADMIN, "Administrador"),
    (Judge.ROLE_SUPERADMIN, "Superadministrador"),
]

ADMIN_MENU_ITEMS = [
    ("overview", "admin.overview", "Resumen"),
    ("assignments", "admin.assignments_page", "Asignaciones"),
    ("judge_pool", "admin.judge_pool_page", "Jueces"),
    ("judges", "admin.judges_page", "Usuarios"),
    ("permissions", "admin.permissions_page", "Permisos"),
    ("campaigns", "admin.campaigns_page", "Campañas"),
    ("categories", "admin.categories_page", "Categorías"),
    ("academic", "admin.academic_page", "Académico"),
    ("rubrics", "admin.rubrics_page", "Rúbricas"),
    ("projects", "admin.projects_page", "Proyectos"),
    ("evaluations", "admin.evaluations_page", "Evaluaciones"),
    ("documents", "admin.documents_page", "Actas y certificados"),
    ("smtp", "admin.smtp_page", "SMTP"),
    ("institution", "admin.institution_page", "Institución"),
    ("maintenance", "admin.maintenance_page", "Mantenimiento"),
    ("database", "admin.database_page", "Base de datos"),
    ("gitops", "admin.gitops_page", "Mantenimiento Git"),
    ("logs", "admin.logs_page", "Bitácora"),
]

ADMIN_MENU_GROUPS = [
    ("General", ["overview"]),
    ("Operación", ["assignments", "judge_pool", "projects", "evaluations"]),
    ("Catálogos", ["campaigns", "categories", "academic", "rubrics"]),
    ("Sistema", ["judges", "permissions", "smtp", "institution", "maintenance", "database", "gitops", "logs"]),
]

ADMIN_MENU_ICONS = {
    "overview": "settings",
    "assignments": "users",
    "judge_pool": "users",
    "judges": "users",
    "permissions": "settings",
    "campaigns": "doc",
    "categories": "filter",
    "academic": "doc",
    "rubrics": "chart",
    "projects": "box",
    "evaluations": "chart",
    "documents": "doc",
    "smtp": "send",
    "institution": "box",
    "maintenance": "settings",
    "database": "box",
    "gitops": "settings",
    "dependencies": "box",
    "logs": "doc",
    "students_stats": "chart",
}

# Override mojibake labels with clean UTF-8 text.
ADMIN_MENU_ITEMS = [
    ("overview", "admin.overview", "Resumen"),
    ("assignments", "admin.assignments_page", "Asignaciones"),
    ("judge_pool", "admin.judge_pool_page", "Jueces"),
    ("judges", "admin.judges_page", "Usuarios"),
    ("permissions", "admin.permissions_page", "Permisos"),
    ("campaigns", "admin.campaigns_page", "Campañas"),
    ("categories", "admin.categories_page", "Categorías"),
    ("academic", "admin.academic_page", "Académico"),
    ("rubrics", "admin.rubrics_page", "Rúbricas"),
    ("projects", "admin.projects_page", "Proyectos"),
    ("evaluations", "admin.evaluations_page", "Evaluaciones"),
    ("documents", "admin.documents_page", "Actas y certificados"),
    ("smtp", "admin.smtp_page", "SMTP"),
    ("institution", "admin.institution_page", "Institución"),
    ("maintenance", "admin.maintenance_page", "Mantenimiento"),
    ("database", "admin.database_page", "Base de datos"),
    ("gitops", "admin.gitops_page", "Mantenimiento Git"),
    ("dependencies", "admin.dependencies_page", "Dependencias"),
    ("logs", "admin.logs_page", "Bitácora"),
    ("students_stats", "admin.students_stats_page", "Estadísticas de estudiantes"),
]

ADMIN_MENU_GROUPS = [
    ("General", ["overview"]),
    ("Documentos", ["documents"]),
    ("Operación", ["assignments", "judge_pool", "projects", "evaluations", "students_stats"]),
    ("Catálogos", ["campaigns", "categories", "academic", "rubrics"]),
    ("Sistema", ["judges", "permissions", "smtp", "institution", "maintenance", "database", "gitops", "dependencies", "logs"]),
]

ADMIN_DEPARTMENT_MODULE_ACCESS = {
    "logistica": {"overview", "assignments", "judge_pool", "projects", "documents"},
    "datos": {"overview", "evaluations", "documents"},
    "diseno": {"overview", "campaigns", "categories", "academic", "rubrics", "institution"},
    "qa": {"overview", "logs", "maintenance", "database", "gitops"},
}
PERMISSIONS_SETTING_KEY = "permissions_department_modules"
PERMISSION_MANAGEABLE_MODULES = [
    key for key, _, _ in ADMIN_MENU_ITEMS if key not in {"overview", "permissions"}
]

ACTION_MODULE_MAP = {
    "create_campaign": "campaigns",
    "update_campaign": "campaigns",
    "delete_campaign": "campaigns",
    "activate_campaign": "campaigns",
    "deactivate_campaign": "campaigns",
    "create_assignment": "assignments",
    "replace_assignment": "assignments",
    "quick_create_assignment_judge": "assignments",
    "delete_assignment": "assignments",
    "auto_assign": "assignments",
    "confirm_draft_assignments": "assignments",
    "discard_draft_assignments": "assignments",
    "create_judge": "judges",
    "update_judge": "judges",
    "reset_judge_password": "judges",
    "set_judge_password": "judges",
    "toggle_judge_active": "judges",
    "toggle_judge_admin": "judges",
    "delete_judge": "judges",
    "save_judge_form_settings": "judges",
    "rotate_judge_form_secret": "judges",
    "update_advisor": "projects",
    "update_project": "projects",
    "update_project_logistics": "projects",
    "replace_project_document": "projects",
    "approve_document_revision": "projects",
    "reject_document_revision": "projects",
    "approve_member_edit": "projects",
    "reject_member_edit": "projects",
    "upload_project_logo": "projects",
    "delete_project": "projects",
    "upload_member_photo": "projects",
    "create_project_member": "projects",
    "update_project_member": "projects",
    "delete_project_member": "projects",
    "create_category": "categories",
    "update_category": "categories",
    "delete_category": "categories",
    "create_level": "academic",
    "update_level": "academic",
    "create_section": "academic",
    "update_section": "academic",
    "delete_section": "academic",
    "create_specialty": "academic",
    "update_specialty": "academic",
    "delete_specialty": "academic",
    "create_workshop": "academic",
    "update_workshop": "academic",
    "delete_workshop": "academic",
    "create_thematic_axis": "academic",
    "update_thematic_axis": "academic",
    "delete_thematic_axis": "academic",
    "create_project_type": "academic",
    "update_project_type": "academic",
    "delete_project_type": "academic",
    "create_evaluation_type": "rubrics",
    "update_evaluation_type": "rubrics",
    "delete_evaluation_type": "rubrics",
    "create_rubric": "rubrics",
    "update_rubric": "rubrics",
    "delete_rubric": "rubrics",
    "delete_evaluation": "evaluations",
    "save_smtp": "smtp",
    "test_smtp": "smtp",
    "save_institution": "institution",
    "save_maintenance_settings": "maintenance",
    "cleanup_expotecnica": "database",
    "backup_database": "database",
    "restore_database": "database",
    "delete_database_backup": "database",
    "database_service_reload": "database",
    "gitops_fetch": "gitops",
    "gitops_pull_ff": "gitops",
    "gitops_pull_apply": "gitops",
    "gitops_revert_commit": "gitops",
    "gitops_refresh": "gitops",
    "gitops_service_reload": "gitops",
    "gitops_service_restart": "gitops",
    "gitops_service_check": "gitops",
    "save_gitops_remote": "gitops",
    "gitops_test_remote": "gitops",
    "save_permissions_matrix": "permissions",
    "send_logistics_reminder": "projects",
    "install_package": "dependencies",
}


def admin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapped(*args, **kwargs):
        if not current_user.has_admin_access:
            flash("Acceso denegado.", "error")
            return redirect(url_for("judge.dashboard"))
        return view_func(*args, **kwargs)

    wrapped.__name__ = view_func.__name__
    return wrapped


def _current_department():
    return (getattr(current_user, "department", "") or "").strip().lower()


def _current_role():
    return (getattr(current_user, "effective_role", "") or "").strip().lower()


def _valid_role(value: str):
    role = (value or "").strip().lower()
    valid_roles = {code for code, _ in USER_ROLES}
    return role if role in valid_roles else Judge.ROLE_JUDGE


def _assignment_scope_from_form(prefix: str = "assignment_scope"):
    scopes = set(request.form.getlist(prefix))
    if not scopes:
        if request.form.get(f"{prefix}_present") == "1":
            return False, False
        return True, True
    can_documentation = "documentacion" in scopes
    can_exposition = "exposicion" in scopes
    return can_documentation, can_exposition


def _assignment_scope_valid(can_documentation: bool, can_exposition: bool) -> bool:
    return bool(can_documentation or can_exposition)


def _apply_assignment_scope(assignment: Assignment, can_documentation: bool, can_exposition: bool):
    assignment.can_evaluate_documentation = bool(can_documentation)
    assignment.can_evaluate_exposition = bool(can_exposition)


def _project_requires_english(project: Project) -> bool:
    if not project:
        return False
    return bool(getattr(project, "requires_english_evaluation", False))


def _assignment_compatibility_error(
    judge: Judge,
    project: Project,
    can_documentation: bool,
    can_exposition: bool,
) -> str:
    if not judge:
        return "Debes seleccionar un juez válido."
    if not project:
        return "Debes seleccionar un proyecto válido."
    if not judge.is_active_user:
        return f"{judge.full_name} está inactivo y no puede recibir asignaciones."
    if can_documentation and not judge.can_evaluate_documentation:
        return f"{judge.full_name} no indicó disponibilidad para evaluar documento escrito."
    if can_exposition and not judge.can_evaluate_exposition:
        return f"{judge.full_name} no indicó disponibilidad para evaluar exposición oral."
    if _project_requires_english(project) and not judge.can_evaluate_english:
        return f"{project.title} requiere evaluación en inglés y {judge.full_name} indicó que no evalúa inglés."
    if not judge.can_evaluate_category(project.category):
        return f"{judge.full_name} no está clasificado para la categoría {project.category}."
    return ""


def _auto_assign_judges(max_per_project: int, replace_drafts: bool) -> tuple[int, int]:
    """Assign up to max_per_project judges per project ensuring both doc and expo
    are covered collectively. Single-dimension judges are paired with complementary
    ones. 'Ambos' judges are preferred. Soft cap of 3 projects per judge.
    Returns (created_count, skipped_count)."""
    SOFT_CAP = 3

    if replace_drafts:
        Assignment.query.filter_by(status=Assignment.STATUS_DRAFT).delete()
        db.session.flush()

    active_projects = Project.query.filter(Project.is_active == True).all()  # noqa: E712
    eligible_judges = Judge.query.filter(
        Judge.is_active_user == True,  # noqa: E712
        Judge.role == Judge.ROLE_JUDGE,
    ).all()
    if not eligible_judges or not active_projects:
        return 0, 0

    judge_load: dict[int, int] = {
        j.id: Assignment.query.filter_by(judge_id=j.id).count()
        for j in eligible_judges
    }

    created = 0
    skipped = 0

    for project in active_projects:
        existing = Assignment.query.filter_by(project_id=project.id).all()
        confirmed = [a for a in existing if a.status == Assignment.STATUS_CONFIRMED]
        existing_drafts = [a for a in existing if a.status == Assignment.STATUS_DRAFT]
        slots = max_per_project - len(confirmed) - len(existing_drafts)
        if slots <= 0:
            skipped += 1
            continue

        # Running coverage counts from confirmed + existing drafts
        doc_count = sum(1 for a in confirmed + existing_drafts if a.can_evaluate_documentation)
        expo_count = sum(1 for a in confirmed + existing_drafts if a.can_evaluate_exposition)
        already_assigned_ids = {a.judge_id for a in existing}
        needs_english = _project_requires_english(project)
        new_assignments: list[Assignment] = []

        def compatible(j) -> bool:
            if j.id in already_assigned_ids:
                return False
            if not j.can_evaluate_category(project.category):
                return False
            return bool(j.can_evaluate_documentation or j.can_evaluate_exposition)

        def base_sort(j):
            over_cap = 1 if judge_load.get(j.id, 0) >= SOFT_CAP else 0
            not_both = 0 if (j.can_evaluate_documentation and j.can_evaluate_exposition) else 1
            no_english = 1 if (needs_english and not j.can_evaluate_english) else 0
            return (over_cap, not_both, no_english, judge_load.get(j.id, 0))

        def pools():
            cands = sorted([j for j in eligible_judges if compatible(j)], key=base_sort)
            both = [j for j in cands if j.can_evaluate_documentation and j.can_evaluate_exposition]
            doc  = [j for j in cands if j.can_evaluate_documentation and not j.can_evaluate_exposition]
            expo = [j for j in cands if not j.can_evaluate_documentation and j.can_evaluate_exposition]
            return both, doc, expo

        def assign(judge):
            nonlocal doc_count, expo_count, slots
            a = Assignment(
                judge_id=judge.id,
                project_id=project.id,
                can_evaluate_documentation=judge.can_evaluate_documentation,
                can_evaluate_exposition=judge.can_evaluate_exposition,
                status=Assignment.STATUS_DRAFT,
            )
            db.session.add(a)
            judge_load[judge.id] = judge_load.get(judge.id, 0) + 1
            already_assigned_ids.add(judge.id)
            new_assignments.append(a)
            slots -= 1
            if judge.can_evaluate_documentation:
                doc_count += 1
            if judge.can_evaluate_exposition:
                expo_count += 1

        # Phase 1: guarantee at least one judge covers doc
        if doc_count == 0 and slots > 0:
            both, doc, _ = pools()
            pick = (both or doc or [None])[0]
            if pick:
                assign(pick)

        # Phase 2: guarantee at least one judge covers expo
        if expo_count == 0 and slots > 0:
            both, _, expo = pools()
            pick = (both or expo or [None])[0]
            if pick:
                assign(pick)

        # Phase 3: fill remaining slots with balanced doc/expo distribution
        while slots > 0:
            both, doc, expo = pools()
            if not both and not doc and not expo:
                break
            # Decide priority based on which dimension has less coverage
            if doc_count <= expo_count:
                # Need more doc coverage: prefer ambos, then doc-only, then expo-only
                pick = (both or doc or expo or [None])[0]
            else:
                # Need more expo coverage: prefer ambos, then expo-only, then doc-only
                pick = (both or expo or doc or [None])[0]
            if pick is None:
                break
            assign(pick)

        created += len(new_assignments)

    db.session.commit()
    return created, skipped


def _build_default_department_access():
    return {dept: sorted(modules) for dept, modules in ADMIN_DEPARTMENT_MODULE_ACCESS.items()}


def _load_department_module_access():
    raw = SystemSetting.get_value(PERMISSIONS_SETTING_KEY, "")
    defaults = _build_default_department_access()
    if not raw:
        return defaults
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return defaults

    sanitized = {}
    valid_departments = {code for code, _ in USER_DEPARTMENTS}
    valid_modules = {key for key, _, _ in ADMIN_MENU_ITEMS if key != "permissions"}

    for dept_code in valid_departments:
        modules = parsed.get(dept_code, defaults.get(dept_code, ["overview"]))
        if not isinstance(modules, list):
            modules = defaults.get(dept_code, ["overview"])
        clean_modules = sorted({module for module in modules if module in valid_modules})
        if "overview" not in clean_modules:
            clean_modules.insert(0, "overview")
        sanitized[dept_code] = clean_modules
    return sanitized


def _save_department_module_access(access_map):
    SystemSetting.set_value(PERMISSIONS_SETTING_KEY, json.dumps(access_map, ensure_ascii=True))


def _allowed_modules_for_current_user():
    if not current_user.is_authenticated or not current_user.has_admin_access:
        return set()
    if current_user.is_superadmin:
        return {module for module, _, _ in ADMIN_MENU_ITEMS}
    if _current_role() == Judge.ROLE_ADMIN:
        dynamic_map = _load_department_module_access()
        return set(dynamic_map.get(_current_department(), {"overview"}))
    return set()


def _can_access_module(module_key: str):
    if module_key == "permissions":
        return current_user.is_superadmin
    return module_key in _allowed_modules_for_current_user()


def _admin_fallback_redirect():
    allowed = _allowed_modules_for_current_user()
    for module_key, endpoint, _ in ADMIN_MENU_ITEMS:
        if module_key in allowed:
            return redirect(url_for(endpoint))
    return redirect(url_for("judge.dashboard"))


def admin_module_required(module_key: str):
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def wrapped(*args, **kwargs):
            if not current_user.has_admin_access:
                flash("Acceso denegado.", "error")
                return redirect(url_for("judge.dashboard"))
            if not _can_access_module(module_key):
                flash("No tienes permisos para este modulo.", "error")
                return _admin_fallback_redirect()
            return view_func(*args, **kwargs)

        wrapped.__name__ = view_func.__name__
        return wrapped

    return decorator


def _can_perform_action(action: str):
    required_module = ACTION_MODULE_MAP.get(action)
    if not required_module:
        return False
    return _can_access_module(required_module)


def _normalize_code(raw_value: str):
    raw_value = (raw_value or "").strip().lower()
    raw_value = re.sub(r"\s+", "_", raw_value)
    raw_value = re.sub(r"[^a-z0-9_]", "", raw_value)
    return raw_value


def _str_to_bool(value: str):
    return str(value).strip().lower() in {"1", "true", "on", "yes"}


def _valid_department(value: str):
    department = (value or "").strip().lower()
    valid_departments = {code for code, _ in USER_DEPARTMENTS}
    return department if department in valid_departments else ""


def _role_requires_department(role: str) -> bool:
    return role in Judge.ADMIN_ROLES


def _normalize_department_for_role(role: str, department: str) -> str:
    return department if _role_requires_department(role) else ""


def _parse_date(raw_value):
    try:
        return datetime.strptime((raw_value or "").strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _validate_category_evaluation_types(exposition_eval_type, documentation_eval_type):
    if not exposition_eval_type or not documentation_eval_type:
        return "Debes asignar una rubrica de Exposicion y una de Documentacion."
    if exposition_eval_type.id == documentation_eval_type.id:
        return "Exposicion y Documentacion deben usar rubricas distintas."
    if infer_evaluation_type_kind(exposition_eval_type) != "exposicion":
        return "La rubrica de Exposicion no corresponde a una evaluacion de exposicion."
    if infer_evaluation_type_kind(documentation_eval_type) != "documentacion":
        return "La rubrica de Documentacion no corresponde a una evaluacion documental."
    return None


def _redirect_next():
    next_url = request.form.get("next", "").strip()
    if next_url and next_url.startswith("/admin/"):
        return redirect(next_url)
    return redirect(url_for("admin.overview"))


def _git_repo_path() -> Path:
    return Path(current_app.root_path).resolve().parent


def _run_git_command(args: list[str], timeout: int = 120) -> dict:
    repo_path = _git_repo_path()
    if not (repo_path / ".git").exists():
        return {"ok": False, "code": -1, "out": "", "err": f"No es un repositorio git: {repo_path}"}
    try:
        proc = subprocess.run(
            args,
            cwd=str(repo_path),
            capture_output=True,
            text=True,
            timeout=timeout,
            shell=False,
        )
        return {"ok": proc.returncode == 0, "code": proc.returncode, "out": (proc.stdout or "").strip(), "err": (proc.stderr or "").strip()}
    except subprocess.TimeoutExpired:
        return {"ok": False, "code": -2, "out": "", "err": f"Timeout ejecutando: {' '.join(args)}"}
    except Exception as ex:
        return {"ok": False, "code": -3, "out": "", "err": str(ex)}


def _git_remote_auth_args() -> list[str]:
    remote_url = (SystemSetting.get_value("gitops_remote_url", "") or "").strip()
    token = (SystemSetting.get_value("gitops_private_token", "") or "").strip()
    username = (SystemSetting.get_value("gitops_username", "x-access-token") or "x-access-token").strip()
    args = []
    if remote_url:
        args.extend(["-c", f"remote.origin.url={remote_url}"])
    if token:
        raw = f"{username}:{token}".encode("utf-8")
        b64 = base64.b64encode(raw).decode("ascii")
        args.extend(["-c", f"http.extraHeader=Authorization: Basic {b64}"])
    return args


def _run_git_remote_command(base_args: list[str], timeout: int = 120) -> dict:
    args = ["git"] + _git_remote_auth_args() + base_args[1:]
    return _run_git_command(args, timeout=timeout)


def _git_status_snapshot() -> dict:
    branch = _run_git_command(["git", "rev-parse", "--abbrev-ref", "HEAD"], timeout=20)
    head = _run_git_command(["git", "rev-parse", "--short", "HEAD"], timeout=20)
    remote = _run_git_command(["git", "config", "--get", "remote.origin.url"], timeout=20)
    status = _run_git_command(["git", "status", "--porcelain"], timeout=20)
    last = _run_git_command(["git", "log", "--pretty=format:%h%x09%s%x09%cr", "-8"], timeout=20)
    ahead = 0
    behind = 0
    upstream = _run_git_command(["git", "rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{u}"], timeout=20)
    if upstream["ok"]:
        counts = _run_git_command(["git", "rev-list", "--left-right", "--count", "@{u}...HEAD"], timeout=20)
        if counts["ok"] and counts["out"]:
            parts = counts["out"].split()
            if len(parts) == 2:
                behind = int(parts[0])
                ahead = int(parts[1])
    changed_files = []
    for raw_line in (status["out"] or "").splitlines():
        line = raw_line.rstrip()
        if not line.strip():
            continue
        code = line[:2]
        path = line[3:].strip() if len(line) > 3 else ""
        if code == "??":
            label = "Nuevo"
        elif "D" in code:
            label = "Eliminado"
        elif "M" in code:
            label = "Modificado"
        elif "R" in code:
            label = "Renombrado"
        else:
            label = "Cambio"
        changed_files.append({"code": code.strip(), "label": label, "path": path})

    revision_rows = []
    for line in (last["out"] or "").splitlines():
        parts = line.split("\t")
        if len(parts) >= 3:
            revision_rows.append({"hash": parts[0], "subject": parts[1], "relative": parts[2]})
        elif line.strip():
            raw_parts = line.split(" ", 1)
            revision_rows.append({"hash": raw_parts[0], "subject": raw_parts[1] if len(raw_parts) > 1 else "", "relative": ""})

    return {
        "repo_path": str(_git_repo_path()),
        "branch": branch["out"] if branch["ok"] else "N/D",
        "head": head["out"] if head["ok"] else "N/D",
        "remote": remote["out"] if remote["ok"] else "N/D",
        "dirty_count": len(changed_files),
        "changed_files": changed_files,
        "ahead": ahead,
        "behind": behind,
        "last_commits": [line for line in (last["out"] or "").splitlines() if line.strip()],
        "revisions": revision_rows,
    }


def _read_gunicorn_config_text() -> str:
    config_path = _git_repo_path() / "gunicorn_conf.py"
    if not config_path.exists():
        return ""
    try:
        return config_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""


def _extract_gunicorn_value(config_text: str, key: str, default: str = "") -> str:
    match = re.search(rf"^\s*{re.escape(key)}\s*=\s*['\"]([^'\"]+)['\"]", config_text, re.MULTILINE)
    return match.group(1).strip() if match else default


def _gitops_service_config() -> dict:
    repo_path = _git_repo_path()
    config_text = _read_gunicorn_config_text()
    bind = _extract_gunicorn_value(config_text, "bind", "127.0.0.1:5055")
    pidfile = _extract_gunicorn_value(config_text, "pidfile", str(repo_path / "logs" / "expotecnica.pid"))
    if "chdir +" in pidfile:
        pidfile = str(repo_path / "logs" / "expotecnica.pid")
    return {
        "repo_path": str(repo_path),
        "config_path": str(repo_path / "gunicorn_conf.py"),
        "bind": bind,
        "pidfile": pidfile,
        "health_path": SystemSetting.get_value("gitops_health_path", "/registro-jueces") or "/registro-jueces",
    }


def _pid_is_running(pid: int) -> bool:
    if not pid or pid < 1:
        return False
    try:
        os.kill(pid, 0)
        return True
    except PermissionError:
        return True
    except OSError:
        return False


def _gitops_service_status() -> dict:
    config = _gitops_service_config()
    pid = None
    pidfile_exists = os.path.exists(config["pidfile"])
    if pidfile_exists:
        try:
            with open(config["pidfile"], "r", encoding="utf-8") as handle:
                pid = int((handle.read() or "").strip())
        except (OSError, ValueError):
            pid = None

    running = _pid_is_running(pid or 0)
    bind = config["bind"]
    host = "127.0.0.1"
    port = ""
    if ":" in bind:
        host_part, port = bind.rsplit(":", 1)
        if host_part and host_part != "0.0.0.0":
            host = host_part

    health_url = f"http://{host}:{port}{config['health_path']}" if port else ""
    http_code = "N/D"
    health_ok = False
    if health_url and running:
        curl = subprocess.run(
            ["curl", "-sS", "-o", os.devnull, "-w", "%{http_code}", "--max-time", "8", health_url],
            capture_output=True,
            text=True,
            timeout=12,
        )
        http_code = (curl.stdout or curl.stderr or "").strip() or "N/D"
        health_ok = http_code.startswith("2") or http_code.startswith("3")

    return {
        **config,
        "pid": pid,
        "pidfile_exists": pidfile_exists,
        "running": running,
        "health_url": health_url,
        "http_code": http_code,
        "health_ok": health_ok,
        "status_label": "Activo" if running and health_ok else ("Proceso activo con alerta" if running else "Detenido"),
    }


def _gitops_reload_service() -> dict:
    status = _gitops_service_status()
    pid = status.get("pid")
    if not pid or not status.get("running"):
        return {"ok": False, "code": -1, "out": "", "err": "No hay PID activo para recargar Gunicorn."}
    try:
        os.kill(pid, 1)
    except OSError as ex:
        return {"ok": False, "code": -2, "out": "", "err": str(ex)}
    recheck = _gitops_service_status()
    out = f"HUP enviado a PID {pid}. Estado: {recheck['status_label']} HTTP {recheck['http_code']}"
    return {"ok": recheck["running"], "code": 0 if recheck["running"] else -3, "out": out, "err": ""}


def _gitops_restart_service() -> dict:
    status = _gitops_service_status()
    pid = status.get("pid")
    pidfile = status.get("pidfile")
    repo_path = _git_repo_path()
    venv_bins = sorted(repo_path.glob("*_venv/bin/gunicorn"))
    gunicorn_bin = str(venv_bins[0]) if venv_bins else "gunicorn"
    config_path = repo_path / "gunicorn_conf.py"
    if not config_path.exists():
        return {"ok": False, "code": -1, "out": "", "err": f"No existe {config_path}"}
    if gunicorn_bin != "gunicorn" and not os.path.exists(gunicorn_bin):
        return {"ok": False, "code": -2, "out": "", "err": f"No existe el binario Gunicorn: {gunicorn_bin}"}

    if pid and status.get("running"):
        try:
            os.kill(pid, 15)
        except OSError:
            pass

        import time as _time

        for _ in range(20):
            if not _pid_is_running(pid):
                break
            _time.sleep(0.5)

    if pidfile:
        try:
            stale_pid = None
            if os.path.exists(pidfile):
                with open(pidfile, "r", encoding="utf-8") as handle:
                    raw_pid = (handle.read() or "").strip()
                stale_pid = int(raw_pid) if raw_pid.isdigit() else None
            if not stale_pid or not _pid_is_running(stale_pid):
                os.remove(pidfile)
        except OSError:
            pass

    try:
        subprocess.Popen(
            [gunicorn_bin, "-c", str(config_path), "run:app"],
            cwd=str(repo_path),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
    except Exception as ex:
        return {"ok": False, "code": -3, "out": "", "err": str(ex)}

    import time as _time

    recheck = {}
    for _ in range(20):
        _time.sleep(0.5)
        recheck = _gitops_service_status()
        if recheck.get("running") and recheck.get("health_ok"):
            break

    out = (
        f"Reinicio solicitado con {gunicorn_bin}. "
        f"Estado: {recheck.get('status_label', 'N/D')} HTTP {recheck.get('http_code', 'N/D')} "
        f"PID {recheck.get('pid') or 'N/D'}"
    )
    ok = bool(recheck.get("running") and recheck.get("health_ok"))
    return {"ok": ok, "code": 0 if ok else -4, "out": out, "err": "" if ok else "El servicio no paso la verificacion despues del reinicio."}


def _save_gitops_result(action: str, result: dict):
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    output = result.get("out") or result.get("err") or "(sin salida)"
    output = output[:8000]
    SystemSetting.set_value("gitops_last_action", action)
    SystemSetting.set_value("gitops_last_status", "ok" if result.get("ok") else "error")
    SystemSetting.set_value("gitops_last_output", output)
    SystemSetting.set_value("gitops_last_ran_at", stamp)


def _project_logistics_missing_items(project):
    missing = []
    if not project.project_document_path:
        missing.append("documento digital adjunto")
    if not project.logistics_document_ok:
        missing.append("proyecto escrito completo")
    if not project.has_real_logo or not project.logistics_logo_ok:
        missing.append("logo validado")
    missing_member_photos = len([member for member in project.members if not member.photo_url])
    if missing_member_photos > 0 or not project.logistics_photos_ok:
        missing.append("fotos de integrantes")
    if not project.logistics_registration_form_signed_ok:
        missing.append("formulario fisico firmado")
    if not project.logistics_student_consents_signed_ok:
        missing.append("consentimientos fisicos firmados")
    if not project.logistics_requirements_reviewed_ok:
        missing.append("revision de requisitos")
    return missing


def _project_logistics_group_missing(project):
    missing = []
    if not project.has_real_logo or not project.logistics_logo_ok:
        missing.append("Logo del proyecto")
    if not project.logistics_document_ok:
        missing.append("Documento escrito")
    if not project.logistics_registration_form_signed_ok:
        missing.append("Formulario físico de inscripción firmado")
    if not project.logistics_student_consents_signed_ok:
        missing.append("Consentimientos físicos estudiantiles firmados")
    if not project.logistics_requirements_reviewed_ok:
        missing.append("Revisión de requisitos completada")
    return missing


def _build_logistics_reminder_data(projects, campaigns):
    from datetime import timedelta

    active_campaign = next((c for c in campaigns if c.is_active), None)
    deadline = None
    if active_campaign and active_campaign.end_date:
        deadline = active_campaign.end_date - timedelta(days=1)
    institution_name = SystemSetting.get_value("school_name", "ExpoTécnica")

    reminder_rows = []
    for project in projects:
        if not project.is_active:
            continue
        missing_group = _project_logistics_group_missing(project)
        member_rows = []
        for member in sorted(project.members, key=lambda m: m.student_number or 0):
            missing_individual = []
            if not member.photo_url:
                missing_individual.append("Foto de perfil")
            if missing_group or missing_individual:
                member_rows.append({
                    "member": member,
                    "missing_individual": missing_individual,
                    "has_email": bool(member.email and member.email.strip()),
                })
        if missing_group or any(r["missing_individual"] for r in member_rows):
            recipients = [r for r in member_rows if r["has_email"]]
            reminder_rows.append({
                "project": project,
                "missing_group": missing_group,
                "member_rows": member_rows,
                "recipients": recipients,
                "recipient_count": len(recipients),
            })

    return {
        "reminder_rows": reminder_rows,
        "active_campaign": active_campaign,
        "deadline": deadline,
        "total_recipients": sum(r["recipient_count"] for r in reminder_rows),
        "total_projects": len(reminder_rows),
        "institution_name": institution_name,
    }


def _render_logistics_reminder_email(member, project, missing_group, missing_individual, deadline, institution_name):
    return render_template(
        "admin/email_logistics_reminder.html",
        member=member,
        project=project,
        missing_group=missing_group,
        missing_individual=missing_individual,
        deadline=deadline,
        institution_name=institution_name,
    )


def _build_overview_metrics(projects, assignments, logistics_page=1, logistics_per_page=5, pending_revisions=None):
    active_projects = [project for project in projects if project.is_active]
    active_project_ids = {project.id for project in active_projects}
    active_assignments = [assignment for assignment in assignments if assignment.project_id in active_project_ids]
    active_members = [member for project in active_projects for member in project.members]

    projects_without_judges = []
    projects_with_pending_evaluations = []
    projects_pending_logistics = []
    total_expected_evaluations = 0
    total_completed_evaluations = 0

    for project in active_projects:
        available_types = get_project_available_evaluation_types(project)
        assigned_count = len(project.assignments)
        expected_evaluations = sum(
            len([eval_type for eval_type in available_types if assignment_allows_evaluation_type(assignment, eval_type)])
            for assignment in project.assignments
        )
        completed_evaluations = len(project.evaluations)

        total_expected_evaluations += expected_evaluations
        total_completed_evaluations += completed_evaluations

        if assigned_count == 0:
            projects_without_judges.append(project)

        if expected_evaluations > 0 and completed_evaluations < expected_evaluations:
            projects_with_pending_evaluations.append(
                {
                    "project": project,
                    "completed": completed_evaluations,
                    "expected": expected_evaluations,
                }
            )

        members_without_photos = [member for member in project.members if not member.photo_url]
        missing_member_photos = len(members_without_photos)
        missing_logistics_items = _project_logistics_missing_items(project)
        if project.logistics_status != "completo" or missing_logistics_items:
            projects_pending_logistics.append(
                {
                    "project": project,
                    "missing_member_photos": missing_member_photos,
                    "members_without_photos": members_without_photos,
                    "missing_logistics_items": missing_logistics_items,
                }
            )

    pending_logistics_rows = sorted(
        projects_pending_logistics,
        key=lambda item: item["project"].created_at,
        reverse=True,
    )
    pending_logistics_total = len(pending_logistics_rows)
    logistics_per_page = max(1, logistics_per_page)
    pending_logistics_pages = max(1, (pending_logistics_total + logistics_per_page - 1) // logistics_per_page)
    logistics_page = min(max(1, logistics_page), pending_logistics_pages)
    logistics_start = 0
    logistics_end = min(logistics_per_page, pending_logistics_total)

    return {
        "active_projects": len(active_projects),
        "active_assignments": len(active_assignments),
        "members_without_photo": len([member for member in active_members if not member.photo_url]),
        "projects_without_logo": len([project for project in active_projects if not project.has_real_logo]),
        "projects_without_document": len([project for project in active_projects if not project.project_document_path]),
        "projects_without_judges": len(projects_without_judges),
        "projects_pending_evaluations": len(projects_with_pending_evaluations),
        "projects_pending_review": len([project for project in active_projects if project.logistics_status == "pendiente_revision"]),
        "projects_incomplete_logistics": len([project for project in active_projects if project.logistics_status == "incompleto" or _project_logistics_missing_items(project)]),
        "completed_evaluations": total_completed_evaluations,
        "expected_evaluations": total_expected_evaluations,
        "urgent_projects": sorted(projects_without_judges, key=lambda item: item.created_at, reverse=True)[:5],
        "pending_evaluation_rows": sorted(
            projects_with_pending_evaluations,
            key=lambda item: (item["expected"] - item["completed"], item["project"].created_at),
            reverse=True,
        )[:5],
        "pending_logistics_total": pending_logistics_total,
        "pending_logistics_displayed": logistics_end,
        "pending_logistics_projects": pending_logistics_rows,
        "pending_logistics_page": logistics_page,
        "pending_logistics_pages": pending_logistics_pages,
        "pending_logistics_has_prev": logistics_page > 1,
        "pending_logistics_has_next": logistics_page < pending_logistics_pages,
        "pending_logistics_prev_page": logistics_page - 1,
        "pending_logistics_next_page": logistics_page + 1,
        "pending_logistics_start": 1 if pending_logistics_total else 0,
        "pending_logistics_end": logistics_end,
        "pending_logistics_per_page": logistics_per_page,
        "pending_document_revisions_count": len(pending_revisions) if pending_revisions is not None else 0,
    }


def _collect_project_acta_data(project: Project, evaluation_types_by_code: dict[str, EvaluationType]):
    category = Category.query.filter_by(code=(project.category or "").strip().lower()).first()
    assigned_judges = sorted(
        [assignment.judge for assignment in project.assignments if assignment.judge],
        key=lambda judge: (judge.full_name or "").lower(),
    )

    evaluations = sorted(
        project.evaluations,
        key=lambda item: (
            (evaluation_types_by_code.get(item.evaluation_type).name if evaluation_types_by_code.get(item.evaluation_type) else item.evaluation_type).lower(),
            (item.judge.full_name if item.judge else "").lower(),
            item.id,
        ),
    )

    evaluation_rows = []
    total_percentage = 0.0
    counted_percentages = 0
    evaluations_by_judge = {}
    for evaluation in evaluations:
        evaluation_type = evaluation_types_by_code.get(evaluation.evaluation_type)
        judge_name = evaluation.judge.full_name if evaluation.judge else "N/D"
        if evaluation.judge:
            evaluations_by_judge.setdefault(evaluation.judge.id, set()).add(evaluation.evaluation_type)

        criteria_rows = sorted(
            [
                {
                    "section_name": (score.criterion.section_name if score.criterion else "") or "",
                    "criterion_name": score.criterion.name if score.criterion else "Criterio",
                    "score": score.score,
                    "max_score": score.criterion.max_score if score.criterion else None,
                    "observation": score.observation or "",
                    "sort_order": score.criterion.sort_order if score.criterion else 0,
                }
                for score in evaluation.scores
            ],
            key=lambda row: (row["sort_order"], row["criterion_name"].lower()),
        )

        if evaluation.percentage is not None:
            total_percentage += evaluation.percentage
            counted_percentages += 1

        evaluation_rows.append(
            {
                "id": evaluation.id,
                "evaluation_type_code": evaluation.evaluation_type,
                "evaluation_type_name": evaluation_type.name if evaluation_type else evaluation.evaluation_type,
                "judge_name": judge_name,
                "judge_email": evaluation.judge.email if evaluation.judge else "",
                "created_at": evaluation.created_at,
                "total_score": evaluation.total_score,
                "max_score": evaluation.max_score,
                "percentage": evaluation.percentage,
                "comments": (evaluation.comments or "").strip(),
                "recommendations": (evaluation.recommendations or "").strip(),
                "criteria_rows": criteria_rows,
            }
        )

    assigned_judge_rows = []
    expected_types = [item.code for item in get_project_available_evaluation_types(project)]
    for judge in assigned_judges:
        submitted_types = evaluations_by_judge.get(judge.id, set())
        assigned_judge_rows.append(
            {
                "id": judge.id,
                "name": judge.full_name,
                "email": judge.email,
                "role_label": judge.role_label,
                "submitted_count": len(submitted_types),
                "expected_count": len(expected_types),
            }
        )

    average_percentage = round(total_percentage / counted_percentages, 2) if counted_percentages else None
    return {
        "project": project,
        "category": category,
        "assigned_judges": assigned_judge_rows,
        "evaluations": evaluation_rows,
        "evaluations_count": len(evaluation_rows),
        "average_percentage": average_percentage,
    }


def _load_project_for_acta(project_id: int):
    return (
        Project.query.options(
            joinedload(Project.assignments).joinedload(Assignment.judge),
            joinedload(Project.evaluations).joinedload(Evaluation.judge),
            joinedload(Project.evaluations).joinedload(Evaluation.scores).joinedload(EvaluationScore.criterion),
        )
        .filter(Project.id == project_id)
        .first()
    )


def _build_project_acta_context(project_id: int):
    project = _load_project_for_acta(project_id)
    if not project:
        return None
    evaluation_types = EvaluationType.query.order_by(EvaluationType.sort_order.asc(), EvaluationType.name.asc()).all()
    evaluation_types_by_code = {item.code: item for item in evaluation_types}
    return _collect_project_acta_data(project, evaluation_types_by_code)


def _build_all_projects_acta_context():
    projects = (
        Project.query.options(
            joinedload(Project.assignments).joinedload(Assignment.judge),
            joinedload(Project.evaluations).joinedload(Evaluation.judge),
            joinedload(Project.evaluations).joinedload(Evaluation.scores).joinedload(EvaluationScore.criterion),
        )
        .filter(Project.is_active.is_(True))
        .order_by(Project.title.asc())
        .all()
    )
    evaluation_types = EvaluationType.query.order_by(EvaluationType.sort_order.asc(), EvaluationType.name.asc()).all()
    evaluation_types_by_code = {item.code: item for item in evaluation_types}
    project_actas = [_collect_project_acta_data(project, evaluation_types_by_code) for project in projects]
    total_evaluations = sum(item["evaluations_count"] for item in project_actas)
    valid_averages = [item["average_percentage"] for item in project_actas if item["average_percentage"] is not None]
    global_average = round(sum(valid_averages) / len(valid_averages), 2) if valid_averages else None
    return {
        "generated_at": datetime.now(),
        "project_actas": project_actas,
        "projects_count": len(project_actas),
        "total_evaluations": total_evaluations,
        "global_average": global_average,
    }


def _build_judge_acta_context(judge_id: int):
    judge = (
        Judge.query.options(
            joinedload(Judge.evaluations).joinedload(Evaluation.project).joinedload(Project.members),
            joinedload(Judge.evaluations).joinedload(Evaluation.project).joinedload(Project.workshop_ref),
            joinedload(Judge.evaluations).joinedload(Evaluation.project).joinedload(Project.specialty_ref),
            joinedload(Judge.evaluations).joinedload(Evaluation.scores).joinedload(EvaluationScore.criterion),
        )
        .filter(Judge.id == judge_id)
        .first()
    )
    if not judge:
        return None

    evaluation_types = EvaluationType.query.order_by(EvaluationType.sort_order.asc(), EvaluationType.name.asc()).all()
    evaluation_types_by_code = {item.code: item for item in evaluation_types}
    active_evaluations = [
        evaluation
        for evaluation in judge.evaluations
        if evaluation.project and evaluation.project.is_active
    ]
    active_judge_users = [judge for judge in judge_users if judge.is_active_user]
    active_evaluations.sort(
        key=lambda item: (
            (item.project.title if item.project else "").lower(),
            (evaluation_types_by_code.get(item.evaluation_type).name if evaluation_types_by_code.get(item.evaluation_type) else item.evaluation_type).lower(),
            item.id,
        )
    )

    project_map = {}
    evaluation_rows = []

    for evaluation in active_evaluations:
        project = evaluation.project
        evaluation_type = evaluation_types_by_code.get(evaluation.evaluation_type)
        criteria_rows = sorted(
            [
                {
                    "section_name": (score.criterion.section_name if score.criterion else "") or "",
                    "criterion_name": score.criterion.name if score.criterion else "Criterio",
                    "score": score.score,
                    "max_score": score.criterion.max_score if score.criterion else None,
                    "observation": score.observation or "",
                    "sort_order": score.criterion.sort_order if score.criterion else 0,
                }
                for score in evaluation.scores
            ],
            key=lambda row: (row["sort_order"], row["criterion_name"].lower()),
        )
        row = {
            "id": evaluation.id,
            "project_id": project.id,
            "project_title": project.title,
            "project_team": project.team_name,
            "evaluation_type_code": evaluation.evaluation_type,
            "evaluation_type_name": evaluation_type.name if evaluation_type else evaluation.evaluation_type,
            "created_at": evaluation.created_at,
            "total_score": evaluation.total_score,
            "max_score": evaluation.max_score,
            "percentage": evaluation.percentage,
            "comments": (evaluation.comments or "").strip(),
            "recommendations": (evaluation.recommendations or "").strip(),
            "criteria_rows": criteria_rows,
        }
        evaluation_rows.append(row)

        project_entry = project_map.setdefault(
            project.id,
            {
                "project": project,
                "evaluation_rows": [],
            },
        )
        project_entry["evaluation_rows"].append(row)

    evaluated_projects = []
    for project_entry in project_map.values():
        project_entry["evaluation_count"] = len(project_entry["evaluation_rows"])
        evaluated_projects.append(project_entry)

    evaluated_projects.sort(key=lambda item: item["project"].title.lower())
    category_sections = [
        {
            "key": "steam",
            "title": "Categoría Desafío STEAM",
            "projects": [item for item in evaluated_projects if (item["project"].category or "").strip().lower() == "steam"],
        },
        {
            "key": "emprendimiento",
            "title": "Categoría Emprendimiento e innovación",
            "projects": [item for item in evaluated_projects if (item["project"].category or "").strip().lower() == "emprendimiento"],
        },
    ]
    return {
        "judge": judge,
        "evaluated_projects": evaluated_projects,
        "category_sections": category_sections,
        "evaluations": evaluation_rows,
        "projects_count": len(evaluated_projects),
        "evaluations_count": len(evaluation_rows),
        "generated_at": datetime.now(),
    }


def _institution_name():
    return SystemSetting.get_value("school_name", "CTP Roberto Gamboa Valverde")


def _month_name_es(month_number: int) -> str:
    months = [
        "enero",
        "febrero",
        "marzo",
        "abril",
        "mayo",
        "junio",
        "julio",
        "agosto",
        "septiembre",
        "octubre",
        "noviembre",
        "diciembre",
    ]
    if 1 <= month_number <= 12:
        return months[month_number - 1]
    return ""


def _long_date_es(value: datetime) -> str:
    return f"{value.day} de {_month_name_es(value.month)} de {value.year}"


def _build_participation_certificate_context(project_id: int | None = None):
    query = Project.query.options(joinedload(Project.members)).filter(Project.is_active.is_(True))
    if project_id is not None:
        query = query.filter(Project.id == project_id)
    projects = query.order_by(Project.title.asc()).all()

    recipients = []
    for project in projects:
        members = sorted(project.members, key=lambda item: (item.student_number, (item.full_name or "").lower(), item.id))
        for member in members:
            recipients.append(
                {
                    "project": project,
                    "member": member,
                    "category_label": (project.category or "N/D").replace("_", " ").title(),
                    "grade_label": member.section_name or project.grade_level or "",
                    "focus_label": member.specialty or project.specialty or "",
                }
            )

    title = "Certificados de participacion" if project_id is None else f"Certificados del proyecto: {projects[0].title}" if projects else "Certificados"
    return {
        "generated_at": datetime.now(),
        "institution_name": _institution_name(),
        "expo_logo_path": SystemSetting.get_value("expo_logo_path", ""),
        "director_name": SystemSetting.get_value("expotec_director_name", ""),
        "technical_coordinator_name": SystemSetting.get_value("expotec_technical_coordinator_name", ""),
        "projects": projects,
        "recipients": recipients,
        "projects_count": len(projects),
        "certificates_count": len(recipients),
        "title": title,
        "single_project": projects[0] if len(projects) == 1 else None,
    }


def _draw_certificate_watermark(pdf, width, height, relative_logo_path: str):
    logo_path = (relative_logo_path or "").strip()
    if not logo_path:
        return
    if logo_path.startswith("http://") or logo_path.startswith("https://"):
        return

    absolute_path = os.path.join(current_app.static_folder, logo_path.replace("/", os.sep))
    if not os.path.exists(absolute_path):
        return

    try:
        image = ImageReader(absolute_path)
        image_width, image_height = image.getSize()
    except Exception:
        return

    if not image_width or not image_height:
        return

    scale = min((width * 0.42) / image_width, (height * 0.52) / image_height)
    draw_width = image_width * scale
    draw_height = image_height * scale
    draw_x = (width - draw_width) / 2
    draw_y = (height - draw_height) / 2 - 18

    pdf.saveState()
    try:
        if hasattr(pdf, "setFillAlpha"):
            pdf.setFillAlpha(0.08)
    except Exception:
        pass
    pdf.drawImage(image, draw_x, draw_y, width=draw_width, height=draw_height, mask="auto", preserveAspectRatio=True)
    pdf.restoreState()


def _build_documents_context():
    judges = (
        Judge.query.options(joinedload(Judge.evaluations).joinedload(Evaluation.project))
        .filter(Judge.is_active_user.is_(True))
        .order_by(Judge.full_name.asc())
        .all()
    )
    judge_rows = []
    for judge in judges:
        active_evaluations = [evaluation for evaluation in judge.evaluations if evaluation.project and evaluation.project.is_active]
        if not active_evaluations:
            continue
        active_project_ids = sorted({evaluation.project_id for evaluation in active_evaluations})
        percentages = [evaluation.percentage for evaluation in active_evaluations if evaluation.percentage is not None]
        judge_rows.append(
            {
                "judge": judge,
                "projects_count": len(active_project_ids),
                "evaluations_count": len(active_evaluations),
                "average_percentage": round(sum(percentages) / len(percentages), 2) if percentages else None,
            }
        )

    active_projects = (
        Project.query.options(joinedload(Project.members))
        .filter(Project.is_active.is_(True))
        .order_by(Project.title.asc())
        .all()
    )
    project_certificate_rows = [
        {
            "project": project,
            "members_count": len(project.members),
            "category_label": (project.category or "N/D").replace("_", " ").title(),
        }
        for project in active_projects
    ]

    return {
        "judge_rows": judge_rows,
        "project_certificate_rows": project_certificate_rows,
        "judge_reports_count": len(judge_rows),
        "active_projects_count": len(active_projects),
        "certificates_count": sum(item["members_count"] for item in project_certificate_rows),
    }


def _pdf_normalize_text(value):
    text = (value or "").strip()
    if not text:
        return ""
    # ReportLab core fonts work better with latin-1 compatible strings.
    return text.encode("latin-1", "replace").decode("latin-1")


def _pdf_wrap_text(text, max_width, font_name, font_size):
    normalized = _pdf_normalize_text(text)
    if not normalized:
        return [""]
    words = normalized.split()
    lines = []
    current = ""
    for word in words:
        trial = f"{current} {word}".strip()
        if pdfmetrics.stringWidth(trial, font_name, font_size) <= max_width:
            current = trial
            continue
        if current:
            lines.append(current)
        current = word
    if current:
        lines.append(current)
    return lines or [""]


def _pdf_draw_wrapped_line_set(pdf, text, x, y, max_width, font_name="Helvetica", font_size=10, line_gap=12):
    pdf.setFont(font_name, font_size)
    lines = _pdf_wrap_text(text, max_width, font_name, font_size)
    for line in lines:
        pdf.drawString(x, y, line)
        y -= line_gap
    return y


def _pdf_draw_centered_wrapped_line_set(pdf, text, center_x, y, max_width, font_name="Helvetica", font_size=10, line_gap=12):
    pdf.setFont(font_name, font_size)
    lines = _pdf_wrap_text(text, max_width, font_name, font_size)
    for line in lines:
        pdf.drawCentredString(center_x, y, line)
        y -= line_gap
    return y


def _certificate_script_font():
    if not REPORTLAB_AVAILABLE:
        return "Helvetica-Bold"

    static_folder = current_app.static_folder if current_app else ""
    candidates = [
        ("LucidaCalligraphy", os.path.join(static_folder, "fonts", "LucidaCalligraphy.ttf")),
        ("LucidaCalligraphy", os.path.join(static_folder, "fonts", "lucida-calligraphy.ttf")),
    ]
    for font_name, font_path in candidates:
        if not os.path.exists(font_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
            return font_name
        except Exception:
            continue

    for fallback_name in ["ZapfChancery-MediumItalic", "Times-Italic"]:
        try:
            pdfmetrics.getFont(fallback_name)
            return fallback_name
        except Exception:
            continue
    return "Helvetica-Oblique"


def _pdf_new_page_with_header(pdf, width, height, title, subtitle):
    pdf.showPage()
    return _pdf_draw_header(pdf, width, height, title, subtitle)


def _pdf_draw_header(pdf, width, height, title, subtitle):
    top = height - 45
    pdf.setFillColor(colors.HexColor("#103f78"))
    pdf.rect(28, height - 78, width - 56, 36, stroke=0, fill=1)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 17)
    pdf.drawString(38, height - 66, "EXPOTECNICA")
    pdf.setFillColor(colors.HexColor("#103f78"))
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(32, top - 48, _pdf_normalize_text(title))
    pdf.setFont("Helvetica", 10)
    pdf.setFillColor(colors.HexColor("#385c88"))
    pdf.drawString(32, top - 63, _pdf_normalize_text(subtitle))
    return top - 84


def _render_project_acta_pdf(acta_data):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = _pdf_draw_header(
        pdf,
        width,
        height,
        f"Acta de evaluacion del proyecto: {acta_data['project'].title}",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )

    project = acta_data["project"]
    category_name = acta_data["category"].name if acta_data["category"] else (project.category or "N/D")
    info_lines = [
        f"Proyecto: {project.title}",
        f"Equipo: {project.team_name}",
        f"Categoria: {category_name}",
        f"Promedio general: {acta_data['average_percentage'] if acta_data['average_percentage'] is not None else 'N/D'}",
    ]
    pdf.setFillColor(colors.black)
    for line in info_lines:
        y = _pdf_draw_wrapped_line_set(pdf, line, 32, y, width - 64, "Helvetica", 10, 13)

    y -= 4
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(32, y, "Jueces asignados")
    y -= 14
    if not acta_data["assigned_judges"]:
        pdf.setFont("Helvetica", 10)
        pdf.drawString(32, y, "Sin jueces asignados.")
        y -= 14
    else:
        for judge in acta_data["assigned_judges"]:
            judge_line = f"- {judge['name']} ({judge['email']}) {judge['submitted_count']}/{judge['expected_count']}"
            y = _pdf_draw_wrapped_line_set(pdf, judge_line, 40, y, width - 80, "Helvetica", 9, 12)
            if y < 80:
                y = _pdf_new_page_with_header(pdf, width, height, f"Acta de evaluacion: {project.title}", "Continuacion")

    y -= 2
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(32, y, "Evaluaciones")
    y -= 12

    if not acta_data["evaluations"]:
        pdf.setFont("Helvetica", 10)
        pdf.drawString(32, y, "No hay evaluaciones registradas.")
        y -= 14
    else:
        for item in acta_data["evaluations"]:
            if y < 120:
                y = _pdf_new_page_with_header(pdf, width, height, f"Acta de evaluacion: {project.title}", "Continuacion")
            pdf.setFillColor(colors.HexColor("#e7f0fb"))
            pdf.roundRect(30, y - 66, width - 60, 62, 6, stroke=0, fill=1)
            pdf.setFillColor(colors.HexColor("#0f3c73"))
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawString(38, y - 16, _pdf_normalize_text(item["evaluation_type_name"]))
            pdf.setFont("Helvetica", 9)
            pdf.drawString(38, y - 30, _pdf_normalize_text(f"Juez: {item['judge_name']}"))
            pdf.drawString(
                38,
                y - 44,
                _pdf_normalize_text(
                    f"Puntaje: {item['total_score']}/{item['max_score'] or 'N/D'} | Porcentaje: {item['percentage'] if item['percentage'] is not None else 'N/D'}"
                ),
            )
            y -= 74
            if item["comments"]:
                y = _pdf_draw_wrapped_line_set(pdf, f"Comentarios: {item['comments']}", 38, y, width - 76, "Helvetica", 9, 11)
            if item["recommendations"]:
                y = _pdf_draw_wrapped_line_set(pdf, f"Recomendaciones: {item['recommendations']}", 38, y, width - 76, "Helvetica", 9, 11)
            if item["criteria_rows"]:
                for criterion in item["criteria_rows"]:
                    criterion_label = f"- {criterion['criterion_name']}: {criterion['score']}/{criterion['max_score'] or 'N/D'}"
                    y = _pdf_draw_wrapped_line_set(pdf, criterion_label, 46, y, width - 92, "Helvetica", 8, 10)
                    if y < 80:
                        y = _pdf_new_page_with_header(pdf, width, height, f"Acta de evaluacion: {project.title}", "Continuacion")
            y -= 6

    pdf.save()
    buffer.seek(0)
    return buffer


def _render_all_projects_acta_pdf(context):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = _pdf_draw_header(
        pdf,
        width,
        height,
        "Acta general de evaluaciones",
        f"Generado: {context['generated_at'].strftime('%Y-%m-%d %H:%M')}",
    )

    summary_line = (
        f"Proyectos: {context['projects_count']} | Evaluaciones: {context['total_evaluations']} | "
        f"Promedio global: {context['global_average'] if context['global_average'] is not None else 'N/D'}"
    )
    y = _pdf_draw_wrapped_line_set(pdf, summary_line, 32, y, width - 64, "Helvetica", 10, 13)
    y -= 8

    for project_data in context["project_actas"]:
        project = project_data["project"]
        if y < 120:
            y = _pdf_new_page_with_header(pdf, width, height, "Acta general de evaluaciones", "Continuacion")

        category_name = project_data["category"].name if project_data["category"] else (project.category or "N/D")
        pdf.setFillColor(colors.HexColor("#dce9f9"))
        pdf.roundRect(30, y - 46, width - 60, 42, 6, stroke=0, fill=1)
        pdf.setFillColor(colors.HexColor("#0f3c73"))
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(38, y - 18, _pdf_normalize_text(project.title))
        pdf.setFont("Helvetica", 9)
        detail_line = (
            f"Equipo: {project.team_name} | Categoria: {category_name} | "
            f"Evaluaciones: {project_data['evaluations_count']} | "
            f"Promedio: {project_data['average_percentage'] if project_data['average_percentage'] is not None else 'N/D'}"
        )
        pdf.drawString(38, y - 33, _pdf_normalize_text(detail_line))
        y -= 54

        if not project_data["evaluations"]:
            pdf.setFont("Helvetica-Oblique", 9)
            pdf.drawString(40, y, "Sin evaluaciones registradas.")
            y -= 14
            continue

        for item in project_data["evaluations"]:
            if y < 80:
                y = _pdf_new_page_with_header(pdf, width, height, "Acta general de evaluaciones", "Continuacion")
            line = (
                f"- {item['evaluation_type_name']} | Juez: {item['judge_name']} | "
                f"Porcentaje: {item['percentage'] if item['percentage'] is not None else 'N/D'}"
            )
            y = _pdf_draw_wrapped_line_set(pdf, line, 42, y, width - 84, "Helvetica", 9, 11)
        y -= 4

    pdf.save()
    buffer.seek(0)
    return buffer


def _render_judge_acta_pdf(context):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    judge = context["judge"]
    generated_at = context["generated_at"]
    institution_name = _institution_name()
    act_number = f"{judge.id}-{generated_at.year}"

    def draw_act_header(page_title: str):
        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawCentredString(width / 2, height - 54, _pdf_normalize_text(page_title))
        pdf.setFont("Helvetica", 11)
        pdf.drawCentredString(width / 2, height - 72, _pdf_normalize_text(institution_name))
        return height - 104

    y = draw_act_header(f"Acta N° {act_number}")
    opening_text = (
        f"Al ser las {generated_at.strftime('%I:%M %p').lower()}, del día {_long_date_es(generated_at)}, "
        f"se hace constar que la persona juez {judge.full_name}, integrante del Comité de Juzgamiento del centro "
        f"educativo {institution_name}, registró evaluación en los siguientes proyectos activos de la etapa "
        f"institucional de la ExpoTÉCNICA, celebrada el {_long_date_es(generated_at)}."
    )
    y = _pdf_draw_wrapped_line_set(pdf, opening_text, 46, y, width - 92, "Helvetica", 10, 15)
    y -= 8

    for section in context["category_sections"]:
        projects = section["projects"]
        if not projects:
            continue

        if y < 170:
            pdf.showPage()
            y = draw_act_header(f"Acta N° {act_number} - Continuación")

        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(46, y, _pdf_normalize_text(section["title"]))
        y -= 18

        for project_data in projects:
            project = project_data["project"]
            if y < 170:
                pdf.showPage()
                y = draw_act_header(f"Acta N° {act_number} - Continuación")
                pdf.setFont("Helvetica-Bold", 11)
                pdf.drawString(46, y, _pdf_normalize_text(section["title"]))
                y -= 18

            pdf.setFont("Helvetica", 10)
            y = _pdf_draw_wrapped_line_set(pdf, f"Nombre del proyecto: {project.title}", 52, y, width - 104, "Helvetica", 10, 14)
            y -= 8

    signature_y = 96
    pdf.setStrokeColor(colors.black)
    pdf.line(70, signature_y, 250, signature_y)
    pdf.line(width - 250, signature_y, width - 70, signature_y)
    pdf.setFont("Helvetica", 10)
    pdf.setFillColor(colors.black)
    pdf.drawCentredString(160, signature_y - 14, "Director(a) del centro educativo")
    pdf.drawCentredString(width - 160, signature_y - 14, "Presidente Comité de Juzgamiento")

    pdf.save()
    buffer.seek(0)
    return buffer


def _render_participation_certificates_pdf(context):
    buffer = BytesIO()
    page_size = (720, 540)  # 10 x 7.5 in, same 4:3 canvas as the official PowerPoint template.
    pdf = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size
    institution_name = context["institution_name"]
    generated_at = context["generated_at"]
    director_name = context.get("director_name") or "MSc. __________________"
    coordinator_name = context.get("technical_coordinator_name") or "MSc. __________________"
    script_font = _certificate_script_font()
    template_path = os.path.join(current_app.static_folder, "certificates", "institucional_2026_bg.jpg")

    for index, recipient in enumerate(context["recipients"]):
        if index:
            pdf.showPage()

        member = recipient["member"]

        if os.path.exists(template_path):
            pdf.drawImage(
                ImageReader(template_path),
                25.5,
                101.4,
                width=682.8,
                height=384.1,
                mask="auto",
                preserveAspectRatio=False,
            )

        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica", 25)
        pdf.drawCentredString(width / 2, 475, _pdf_normalize_text("Ministerio de Educaci\u00f3n P\u00fablica"))
        pdf.setFont("Helvetica", 14)
        pdf.drawCentredString(width / 2, 421, _pdf_normalize_text(institution_name))

        pdf.setFont("Helvetica-Bold", 15)
        pdf.drawCentredString(width / 2, 369, _pdf_normalize_text("Otorga el presente certificado a:"))

        pdf.setFont(script_font, 42)
        pdf.drawCentredString(width / 2, 312, _pdf_normalize_text(member.full_name))

        pdf.setFont("Helvetica", 18)
        pdf.drawCentredString(width / 2, 247, _pdf_normalize_text("Por su participaci\u00f3n en la:"))

        pdf.setFont(script_font, 23)
        pdf.drawCentredString(width / 2, 194, _pdf_normalize_text("Etapa institucional de ExpoT\u00c9CNICA"))

        date_line = (
            f"Realizada el {generated_at.day} del mes de {_month_name_es(generated_at.month)} "
            f"del a\u00f1o {generated_at.year} en el Colegio T\u00e9cnico Profesional {institution_name}."
        )
        _pdf_draw_centered_wrapped_line_set(
            pdf,
            date_line,
            width / 2,
            130,
            width - 160,
            "Helvetica",
            8.8,
            11,
        )

        pdf.setFont("Helvetica", 8.8)
        pdf.drawCentredString(
            width / 2,
            105,
            _pdf_normalize_text(
                f"Dado a los {generated_at.day} d\u00edas del mes de {_month_name_es(generated_at.month)} de {generated_at.year}."
            ),
        )

        pdf.setStrokeColor(colors.black)
        pdf.setLineWidth(0.75)
        pdf.line(108, 78, 280, 78)
        pdf.line(440, 78, 612, 78)
        pdf.setFont("Helvetica-Bold", 8.5)
        pdf.drawCentredString(194, 61, _pdf_normalize_text(director_name))
        pdf.drawCentredString(526, 61, _pdf_normalize_text(coordinator_name))
        pdf.drawCentredString(194, 47, "Director del centro educativo")
        pdf.drawCentredString(526, 47, _pdf_normalize_text("Coordinador Institucional ExpoT\u00c9CNICA"))

    pdf.save()
    buffer.seek(0)
    return buffer

def _get_extension(filename):
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _save_member_photo(photo_file):
    original_name = secure_filename(photo_file.filename or "")
    extension = _get_extension(original_name)
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("Formato de imagen invalido. Usa PNG, JPG, JPEG, WEBP o GIF.")

    relative_dir = os.path.join("uploads", "members")
    absolute_dir = os.path.join(current_app.static_folder, relative_dir)
    os.makedirs(absolute_dir, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}.{extension}"
    absolute_path = os.path.join(absolute_dir, unique_name)
    photo_file.save(absolute_path)
    return f"{relative_dir}/{unique_name}".replace("\\", "/")


def _save_project_logo(photo_file):
    original_name = secure_filename(photo_file.filename or "")
    extension = _get_extension(original_name)
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("Formato de logo invalido. Usa PNG, JPG, JPEG, WEBP o GIF.")

    relative_dir = os.path.join("uploads", "projects", "logos")
    absolute_dir = os.path.join(current_app.static_folder, relative_dir)
    os.makedirs(absolute_dir, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}.{extension}"
    absolute_path = os.path.join(absolute_dir, unique_name)
    photo_file.save(absolute_path)
    return f"{relative_dir}/{unique_name}".replace("\\", "/")


def _save_project_document(document_file):
    original_name = secure_filename(document_file.filename or "")
    extension = _get_extension(original_name)
    if extension not in ALLOWED_PROJECT_DOCUMENT_EXTENSIONS:
        raise ValueError("Formato de documento invalido. Usa unicamente PDF.")

    relative_dir = os.path.join("uploads", "projects", "documents")
    absolute_dir = os.path.join(current_app.static_folder, relative_dir)
    os.makedirs(absolute_dir, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}.pdf"
    absolute_path = os.path.join(absolute_dir, unique_name)
    document_file.save(absolute_path)
    return f"{relative_dir}/{unique_name}".replace("\\", "/")


def _save_institution_logo(photo_file):
    original_name = secure_filename(photo_file.filename or "")
    extension = _get_extension(original_name)
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("Formato de logo institucional invalido. Usa PNG, JPG, JPEG, WEBP o GIF.")

    relative_dir = os.path.join("uploads", "institution")
    absolute_dir = os.path.join(current_app.static_folder, relative_dir)
    os.makedirs(absolute_dir, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}.{extension}"
    absolute_path = os.path.join(absolute_dir, unique_name)
    photo_file.save(absolute_path)
    return f"{relative_dir}/{unique_name}".replace("\\", "/")


def _save_maintenance_image(photo_file):
    original_name = secure_filename(photo_file.filename or "")
    extension = _get_extension(original_name)
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("Formato de imagen invalido. Usa PNG, JPG, JPEG, WEBP o GIF.")

    relative_dir = os.path.join("uploads", "maintenance")
    absolute_dir = os.path.join(current_app.static_folder, relative_dir)
    os.makedirs(absolute_dir, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}.{extension}"
    absolute_path = os.path.join(absolute_dir, unique_name)
    photo_file.save(absolute_path)
    return f"{relative_dir}/{unique_name}".replace("\\", "/")


def _cleanup_expotecnica_counts():
    users_to_delete = Judge.query.filter(
        Judge.is_admin.is_(False),
        Judge.role.notin_(list(Judge.ADMIN_ROLES)),
    ).count()
    return {
        "projects": Project.query.count(),
        "members": ProjectMember.query.count(),
        "member_changes": ProjectMemberChange.query.count(),
        "assignments": Assignment.query.count(),
        "users": users_to_delete,
        "evaluations": Evaluation.query.count(),
        "evaluation_scores": EvaluationScore.query.count(),
    }


def _database_backup_dir() -> Path:
    backup_dir = Path(current_app.instance_path).resolve() / "database_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    return backup_dir


def _database_restore_jobs_dir() -> Path:
    jobs_dir = Path(current_app.instance_path).resolve() / "database_restore_jobs"
    jobs_dir.mkdir(parents=True, exist_ok=True)
    return jobs_dir


def _safe_backup_filename(filename: str) -> str:
    filename = (filename or "").strip()
    if not re.fullmatch(r"expotecnica_db_\d{8}_\d{6}_[a-z0-9_-]+\.sql", filename):
        raise ValueError("Nombre de respaldo invalido.")
    backup_dir = _database_backup_dir()
    candidate = (backup_dir / filename).resolve()
    if backup_dir not in candidate.parents:
        raise ValueError("Ruta de respaldo no permitida.")
    if not candidate.exists() or not candidate.is_file():
        raise ValueError("El respaldo seleccionado no existe.")
    return filename


def _database_url_config():
    url = make_url(current_app.config["SQLALCHEMY_DATABASE_URI"])
    if not url.drivername.startswith("mysql"):
        raise RuntimeError("Los respaldos automaticos estan disponibles solo para bases MySQL/MariaDB.")
    if not url.database:
        raise RuntimeError("DATABASE_URL no define el nombre de la base de datos.")
    return {
        "host": url.host or "localhost",
        "port": str(url.port or 3306),
        "user": url.username or "",
        "password": url.password or "",
        "database": url.database,
    }


def _mysql_env(db_config: dict) -> dict:
    env = os.environ.copy()
    if db_config.get("password"):
        env["MYSQL_PWD"] = db_config["password"]
    return env


def _mysql_base_args(binary: str, db_config: dict) -> list[str]:
    args = [
        binary,
        "--host",
        db_config["host"],
        "--port",
        db_config["port"],
        "--user",
        db_config["user"],
        "--default-character-set=utf8mb4",
    ]
    return args


def _create_database_backup(reason: str = "manual") -> dict:
    db_config = _database_url_config()
    backup_dir = _database_backup_dir()
    safe_reason = re.sub(r"[^a-z0-9_-]+", "_", (reason or "manual").strip().lower()).strip("_") or "manual"
    filename = f"expotecnica_db_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_reason}.sql"
    backup_path = backup_dir / filename

    command = _mysql_base_args("mysqldump", db_config) + [
        "--single-transaction",
        "--quick",
        "--no-tablespaces",
        "--routines",
        "--triggers",
        "--events",
        db_config["database"],
    ]
    try:
        with backup_path.open("wb") as backup_file:
            result = subprocess.run(
                command,
                stdout=backup_file,
                stderr=subprocess.PIPE,
                env=_mysql_env(db_config),
                timeout=300,
                check=False,
            )
    except FileNotFoundError as error:
        raise RuntimeError("No se encontro mysqldump en el servidor.") from error
    except subprocess.TimeoutExpired as error:
        backup_path.unlink(missing_ok=True)
        raise RuntimeError("El respaldo excedio el tiempo maximo permitido.") from error

    if result.returncode != 0:
        backup_path.unlink(missing_ok=True)
        error_text = (result.stderr or b"").decode("utf-8", "replace").strip()
        raise RuntimeError(error_text or "No se pudo crear el respaldo de base de datos.")

    return {
        "filename": filename,
        "path": str(backup_path),
        "size_bytes": backup_path.stat().st_size,
        "created_at": datetime.fromtimestamp(backup_path.stat().st_mtime),
    }


def _list_database_backups() -> list[dict]:
    backups = []
    backup_dir = _database_backup_dir()
    for path in backup_dir.glob("expotecnica_db_*.sql"):
        if not path.is_file():
            continue
        stat = path.stat()
        backups.append(
            {
                "filename": path.name,
                "size_bytes": stat.st_size,
                "size_label": _format_bytes(stat.st_size),
                "created_at": datetime.fromtimestamp(stat.st_mtime),
            }
        )
    backups.sort(key=lambda item: item["created_at"], reverse=True)
    return backups


def _delete_database_backup(filename: str) -> dict:
    safe_filename = _safe_backup_filename(filename)
    backup_path = _database_backup_dir() / safe_filename
    size_bytes = backup_path.stat().st_size
    backup_path.unlink()
    return {"filename": safe_filename, "size_bytes": size_bytes}


def _restore_database_backup(filename: str) -> dict:
    safe_filename = _safe_backup_filename(filename)
    db_config = _database_url_config()
    backup_path = _database_backup_dir() / safe_filename
    command = _mysql_base_args("mysql", db_config) + [db_config["database"]]
    try:
        with backup_path.open("rb") as backup_file:
            result = subprocess.run(
                command,
                stdin=backup_file,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                env=_mysql_env(db_config),
                timeout=300,
                check=False,
            )
    except FileNotFoundError as error:
        raise RuntimeError("No se encontro mysql en el servidor.") from error
    except subprocess.TimeoutExpired as error:
        raise RuntimeError("La restauracion excedio el tiempo maximo permitido.") from error

    if result.returncode != 0:
        error_text = (result.stderr or b"").decode("utf-8", "replace").strip()
        raise RuntimeError(error_text or "No se pudo restaurar el respaldo.")

    return {
        "filename": safe_filename,
        "size_bytes": backup_path.stat().st_size,
    }


def _restore_job_paths(job_id: str) -> dict:
    jobs_dir = _database_restore_jobs_dir()
    safe_job_id = re.sub(r"[^a-z0-9_-]+", "", (job_id or "").lower())
    if not safe_job_id:
        raise ValueError("Identificador de restauracion invalido.")
    return {
        "status": jobs_dir / f"{safe_job_id}.json",
        "log": jobs_dir / f"{safe_job_id}.log",
    }


def _write_restore_job_status(job_id: str, payload: dict) -> None:
    paths = _restore_job_paths(job_id)
    paths["status"].write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _tail_text_file(path: Path, max_lines: int = 12) -> str:
    if not path.exists() or not path.is_file():
        return ""
    try:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError:
        return ""
    return "\n".join(lines[-max_lines:])


def _list_restore_jobs(limit: int = 8) -> list[dict]:
    jobs = []
    jobs_dir = _database_restore_jobs_dir()
    for status_path in jobs_dir.glob("*.json"):
        try:
            payload = json.loads(status_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        job_id = payload.get("job_id") or status_path.stem
        log_path = jobs_dir / f"{job_id}.log"
        payload["job_id"] = job_id
        payload["log_tail"] = _tail_text_file(log_path)
        payload["log_path"] = str(log_path)
        payload["is_running"] = payload.get("status") == "running"
        jobs.append(payload)
    jobs.sort(key=lambda item: item.get("updated_at") or item.get("started_at") or "", reverse=True)
    return jobs[:limit]


def _start_database_restore_job(filename: str) -> dict:
    safe_filename = _safe_backup_filename(filename)
    db_config = _database_url_config()
    backup_path = _database_backup_dir() / safe_filename
    job_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    safe_reason = f"antes_restaurar_{job_id}".lower()
    safety_filename = f"expotecnica_db_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_reason}.sql"
    safety_path = _database_backup_dir() / safety_filename
    paths = _restore_job_paths(job_id)

    initial_payload = {
        "job_id": job_id,
        "status": "queued",
        "filename": safe_filename,
        "safety_backup": safety_filename,
        "message": "Restauracion en cola.",
        "started_at": started_at,
        "updated_at": started_at,
        "finished_at": "",
    }
    _write_restore_job_status(job_id, initial_payload)
    paths["log"].write_text(f"[{started_at}] Restauracion solicitada para {safe_filename}\n", encoding="utf-8")

    script = r'''
set +e
log_line() {
  printf '[%s] %s\n' "$(date '+%Y-%m-%d %H:%M:%S')" "$1" >> "$EXPOTEC_LOG_FILE"
}
write_status() {
  status="$1"
  message="$2"
  finished_at="$3"
  updated_at="$(date '+%Y-%m-%d %H:%M:%S')"
  cat > "$EXPOTEC_STATUS_FILE" <<EOF
{
  "job_id": "$EXPOTEC_JOB_ID",
  "status": "$status",
  "filename": "$EXPOTEC_BACKUP_NAME",
  "safety_backup": "$EXPOTEC_SAFETY_NAME",
  "message": "$message",
  "started_at": "$EXPOTEC_STARTED_AT",
  "updated_at": "$updated_at",
  "finished_at": "$finished_at"
}
EOF
}
fail_job() {
  log_line "$1"
  write_status "failed" "$1" "$(date '+%Y-%m-%d %H:%M:%S')"
  exit 1
}

write_status "running" "Creando respaldo preventivo antes de restaurar..." ""
log_line "Verificando herramientas mysql y mysqldump."
command -v mysqldump >/dev/null 2>&1 || fail_job "No se encontro mysqldump en el servidor."
command -v mysql >/dev/null 2>&1 || fail_job "No se encontro mysql en el servidor."

log_line "Creando respaldo preventivo: $EXPOTEC_SAFETY_NAME"
mysqldump \
  --host "$EXPOTEC_DB_HOST" \
  --port "$EXPOTEC_DB_PORT" \
  --user "$EXPOTEC_DB_USER" \
  --default-character-set=utf8mb4 \
  --single-transaction \
  --quick \
  --no-tablespaces \
  --routines \
  --triggers \
  --events \
  "$EXPOTEC_DB_NAME" > "$EXPOTEC_SAFETY_FILE" 2>> "$EXPOTEC_LOG_FILE"
backup_code=$?
if [ "$backup_code" -ne 0 ]; then
  rm -f "$EXPOTEC_SAFETY_FILE"
  fail_job "Fallo el respaldo preventivo. Codigo $backup_code. Revisa el log."
fi

write_status "running" "Respaldo preventivo creado. Restaurando base de datos..." ""
log_line "Restaurando desde: $EXPOTEC_BACKUP_NAME"
mysql \
  --host "$EXPOTEC_DB_HOST" \
  --port "$EXPOTEC_DB_PORT" \
  --user "$EXPOTEC_DB_USER" \
  --default-character-set=utf8mb4 \
  "$EXPOTEC_DB_NAME" < "$EXPOTEC_BACKUP_FILE" >> "$EXPOTEC_LOG_FILE" 2>&1
restore_code=$?
if [ "$restore_code" -ne 0 ]; then
  fail_job "Fallo la restauracion. Codigo $restore_code. Se conserva el respaldo preventivo."
fi

log_line "Restauracion completada correctamente."
write_status "success" "Restauracion completada correctamente." "$(date '+%Y-%m-%d %H:%M:%S')"
exit 0
'''
    env = _mysql_env(db_config)
    env.update(
        {
            "EXPOTEC_JOB_ID": job_id,
            "EXPOTEC_BACKUP_NAME": safe_filename,
            "EXPOTEC_BACKUP_FILE": str(backup_path),
            "EXPOTEC_SAFETY_NAME": safety_filename,
            "EXPOTEC_SAFETY_FILE": str(safety_path),
            "EXPOTEC_STATUS_FILE": str(paths["status"]),
            "EXPOTEC_LOG_FILE": str(paths["log"]),
            "EXPOTEC_STARTED_AT": started_at,
            "EXPOTEC_DB_HOST": db_config["host"],
            "EXPOTEC_DB_PORT": db_config["port"],
            "EXPOTEC_DB_USER": db_config["user"],
            "EXPOTEC_DB_NAME": db_config["database"],
        }
    )
    try:
        process = subprocess.Popen(
            ["/bin/sh", "-c", script],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            env=env,
            start_new_session=True,
        )
    except FileNotFoundError as error:
        _write_restore_job_status(
            job_id,
            {
                **initial_payload,
                "status": "failed",
                "message": "No se encontro /bin/sh para iniciar la restauracion.",
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
        )
        raise RuntimeError("No se encontro /bin/sh para iniciar la restauracion.") from error

    queued_payload = {
        **initial_payload,
        "status": "running",
        "pid": process.pid,
        "message": "Restauracion iniciada en segundo plano.",
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _write_restore_job_status(job_id, queued_payload)
    return queued_payload


def _format_bytes(size: int) -> str:
    value = float(size or 0)
    for unit in ["B", "KB", "MB", "GB"]:
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{value:.1f} GB"


def _database_backup_storage_summary(backups: list[dict]) -> dict:
    total_size = sum(item.get("size_bytes", 0) for item in backups)
    protected = [item for item in backups if "antes_restaurar" in item["filename"] or "antes_limpiar" in item["filename"]]
    return {
        "count": len(backups),
        "total_size": total_size,
        "total_size_label": _format_bytes(total_size),
        "latest": backups[0] if backups else None,
        "protected_count": len(protected),
    }


def _database_required_tables() -> list[str]:
    return sorted(
        {
            Assignment.__tablename__,
            Campaign.__tablename__,
            Category.__tablename__,
            Evaluation.__tablename__,
            EvaluationScore.__tablename__,
            EvaluationType.__tablename__,
            Judge.__tablename__,
            Level.__tablename__,
            Project.__tablename__,
            ProjectMember.__tablename__,
            ProjectMemberChange.__tablename__,
            RubricCriterion.__tablename__,
            Section.__tablename__,
            Specialty.__tablename__,
            SystemAuditLog.__tablename__,
            SystemSetting.__tablename__,
            Workshop.__tablename__,
        }
    )


def _database_diagnostics() -> dict:
    db_config = _database_url_config()
    required_tables = _database_required_tables()
    result = {
        "ok": False,
        "database": db_config["database"],
        "host": db_config["host"],
        "port": db_config["port"],
        "version": "N/D",
        "size_label": "N/D",
        "table_count": 0,
        "missing_tables": required_tables,
        "tables": [],
        "checks": [],
        "error": "",
    }
    try:
        version_row = db.session.execute(text("SELECT VERSION() AS version")).mappings().first()
        result["version"] = version_row["version"] if version_row else "N/D"
        summary = db.session.execute(
            text(
                """
                SELECT
                    COUNT(*) AS table_count,
                    COALESCE(SUM(data_length + index_length), 0) AS total_bytes
                FROM information_schema.tables
                WHERE table_schema = :schema
                """
            ),
            {"schema": db_config["database"]},
        ).mappings().first()
        result["table_count"] = int(summary["table_count"] or 0) if summary else 0
        result["size_label"] = _format_bytes(int(summary["total_bytes"] or 0)) if summary else "0 B"

        rows = db.session.execute(
            text(
                """
                SELECT
                    TABLE_NAME AS table_name,
                    TABLE_ROWS AS table_rows,
                    DATA_LENGTH AS data_length,
                    INDEX_LENGTH AS index_length,
                    UPDATE_TIME AS update_time
                FROM information_schema.tables
                WHERE TABLE_SCHEMA = :schema
                ORDER BY (DATA_LENGTH + INDEX_LENGTH) DESC, TABLE_NAME ASC
                LIMIT 40
                """
            ),
            {"schema": db_config["database"]},
        ).mappings().all()
        all_tables = db.session.execute(
            text("SELECT TABLE_NAME AS table_name FROM information_schema.tables WHERE TABLE_SCHEMA = :schema"),
            {"schema": db_config["database"]},
        ).scalars().all()
        existing_all = set(all_tables)
        result["missing_tables"] = [name for name in required_tables if name not in existing_all]
        result["tables"] = [
            {
                "name": row["table_name"],
                "rows": int(row["table_rows"] or 0),
                "data_label": _format_bytes(int(row["data_length"] or 0)),
                "index_label": _format_bytes(int(row["index_length"] or 0)),
                "total_label": _format_bytes(int(row["data_length"] or 0) + int(row["index_length"] or 0)),
                "update_time": row["update_time"],
                "required": row["table_name"] in required_tables,
            }
            for row in rows
        ]
        result["checks"] = [
            {"label": "Conexion MySQL", "ok": True, "detail": f"{db_config['host']}:{db_config['port']}"},
            {"label": "Tablas requeridas", "ok": not result["missing_tables"], "detail": f"Faltantes: {len(result['missing_tables'])}"},
            {"label": "Respaldos disponibles", "ok": bool(_list_database_backups()), "detail": "Ver seccion Respaldos"},
        ]
        result["ok"] = not result["missing_tables"]
    except Exception as error:
        db.session.rollback()
        result["error"] = str(error)
        result["checks"] = [
            {"label": "Conexion MySQL", "ok": False, "detail": str(error)},
            {"label": "Tablas requeridas", "ok": False, "detail": "No se pudo diagnosticar"},
            {"label": "Respaldos disponibles", "ok": bool(_list_database_backups()), "detail": "Ver seccion Respaldos"},
        ]
    return result


def _database_operational_counts() -> list[dict]:
    counters = [
        ("Proyectos", Project),
        ("Integrantes", ProjectMember),
        ("Jueces / usuarios", Judge),
        ("Asignaciones", Assignment),
        ("Evaluaciones", Evaluation),
        ("Puntajes", EvaluationScore),
        ("Bitacora", SystemAuditLog),
    ]
    rows = []
    for label, model in counters:
        try:
            rows.append({"label": label, "value": model.query.count(), "ok": True})
        except Exception as error:
            db.session.rollback()
            rows.append({"label": label, "value": "N/D", "ok": False, "detail": str(error)})
    return rows


def _safe_cleanup_expotecnica_counts() -> dict:
    try:
        return _cleanup_expotecnica_counts()
    except Exception:
        db.session.rollback()
        return {
            "projects": "N/D",
            "members": "N/D",
            "member_changes": "N/D",
            "assignments": "N/D",
            "users": "N/D",
            "evaluations": "N/D",
            "evaluation_scores": "N/D",
        }


def _database_audit_logs(limit: int = 30):
    try:
        return (
            SystemAuditLog.query.filter(
                or_(
                    SystemAuditLog.action.ilike("%database%"),
                    SystemAuditLog.action.ilike("%backup%"),
                    SystemAuditLog.action.ilike("%restore%"),
                    SystemAuditLog.action.ilike("%cleanup_expotecnica%"),
                )
            )
            .order_by(SystemAuditLog.created_at.desc())
            .limit(limit)
            .all()
        )
    except Exception:
        db.session.rollback()
        return []


def _clear_static_upload_dir(relative_dir: str) -> int:
    static_root = Path(current_app.static_folder).resolve()
    uploads_root = (static_root / "uploads").resolve()
    target = (static_root / relative_dir).resolve()
    if uploads_root not in target.parents and target != uploads_root:
        raise ValueError(f"Ruta de limpieza no permitida: {relative_dir}")

    target.mkdir(parents=True, exist_ok=True)
    deleted = 0
    for item in target.iterdir():
        if item.is_dir():
            shutil.rmtree(item)
        else:
            item.unlink()
        deleted += 1
    return deleted


def _run_expotecnica_cleanup():
    before = _cleanup_expotecnica_counts()
    deleted_files = 0
    for relative_dir in [
        "uploads/projects/documents",
        "uploads/projects/temp_documents",
        "uploads/projects/logos",
        "uploads/members",
    ]:
        deleted_files += _clear_static_upload_dir(relative_dir)

    Evaluation.query.update(
        {
            Evaluation.judge_id: None,
            Evaluation.project_id: None,
        },
        synchronize_session=False,
    )
    Assignment.query.delete(synchronize_session=False)
    Judge.query.filter(
        Judge.is_admin.is_(False),
        Judge.role.notin_(list(Judge.ADMIN_ROLES)),
    ).delete(synchronize_session=False)
    ProjectMemberChange.query.delete(synchronize_session=False)
    ProjectMember.query.delete(synchronize_session=False)
    Project.query.delete(synchronize_session=False)
    return before, deleted_files


def _send_judge_credentials_email(judge: Judge, plain_password: str):
    if not smtp_is_configured():
        return False

    login_url = url_for("auth.login", _external=True)
    school_name = SystemSetting.get_value("school_name", "CTP Roberto Gamboa Valverde")
    school_logo = SystemSetting.get_value("school_logo_path", "")
    expo_logo = SystemSetting.get_value("expo_logo_path", "")
    school_logo_url = url_for("static", filename=school_logo, _external=True) if school_logo else ""
    expo_logo_url = url_for("static", filename=expo_logo, _external=True) if expo_logo else ""
    subject = "Credenciales de acceso - ExpoTécnica"
    body = (
        f"Hola {judge.full_name},\n\n"
        "Gracias por confirmar tu participación como juez de ExpoTécnica.\n"
        f"Disponibilidad registrada: {judge.evaluation_scope_label}.\n"
        "La organización usará esta información para asignarte evaluaciones según la necesidad de cada proyecto.\n\n"
        "Se ha creado/actualizado tu acceso al portal de juez.\n"
        f"Portal: {login_url}\n"
        f"Usuario: {judge.email}\n"
        f"Contraseña temporal: {plain_password}\n\n"
        "Por seguridad, cambia esta contraseña al ingresar.\n"
    )
    html_body = _build_judge_credentials_email_html(
        judge=judge,
        plain_password=plain_password,
        login_url=login_url,
        school_name=school_name,
        school_logo_url=school_logo_url,
        expo_logo_url=expo_logo_url,
    )
    ok, error = send_email(judge.email, subject, body, html_body=html_body)
    if not ok:
        flash(f"No se pudo enviar correo de credenciales: {error}", "error")
        return False
    return True


def _build_judge_credentials_email_html(
    *,
    judge: Judge,
    plain_password: str,
    login_url: str,
    school_name: str,
    school_logo_url: str = "",
    expo_logo_url: str = "",
) -> str:
    logo_cells = ""
    if school_logo_url:
        logo_cells += (
            f'<img src="{escape(school_logo_url)}" alt="{escape(school_name)}" '
            'style="height:70px;max-width:90px;object-fit:contain;margin-right:14px;">'
        )
    if expo_logo_url:
        logo_cells += (
            f'<img src="{escape(expo_logo_url)}" alt="ExpoTécnica" '
            'style="height:70px;max-width:150px;object-fit:contain;">'
        )
    if not logo_cells:
        logo_cells = '<strong style="font-size:22px;color:#ffffff;">ExpoTécnica</strong>'

    return f"""\
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Credenciales ExpoTécnica</title>
</head>
<body style="margin:0;padding:0;background:#eef5fb;font-family:Arial,Helvetica,sans-serif;color:#123f6b;">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#eef5fb;padding:28px 12px;">
    <tr>
      <td align="center">
        <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:680px;background:#ffffff;border:1px solid #cfe0f1;border-radius:22px;overflow:hidden;box-shadow:0 18px 38px rgba(18,63,107,.12);">
          <tr>
            <td style="background:#123f6b;padding:22px 26px;">
              <table role="presentation" width="100%" cellspacing="0" cellpadding="0">
                <tr>
                  <td style="vertical-align:middle;">{logo_cells}</td>
                  <td align="right" style="vertical-align:middle;color:#ffffff;">
                    <div style="font-size:12px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;">ExpoTécnica 2026</div>
                    <div style="font-size:18px;font-weight:800;margin-top:4px;">Acceso de juez</div>
                  </td>
                </tr>
              </table>
            </td>
          </tr>
          <tr>
            <td style="padding:30px 30px 18px;">
              <p style="margin:0 0 8px;font-size:14px;font-weight:800;letter-spacing:.1em;text-transform:uppercase;color:#1f8fb5;">Registro confirmado</p>
              <h1 style="margin:0 0 14px;font-size:30px;line-height:1.15;color:#123f6b;">Gracias, {escape(judge.full_name)}</h1>
              <p style="margin:0 0 22px;font-size:16px;line-height:1.55;color:#4f6680;">Tu acceso como juez quedó registrado en el sistema de ExpoTécnica de {escape(school_name)}.</p>
              <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f7fbff;border:1px solid #cfe0f1;border-radius:16px;margin:0 0 22px;">
                <tr>
                  <td style="padding:18px 20px;">
                    <p style="margin:0 0 10px;font-size:13px;font-weight:800;letter-spacing:.08em;text-transform:uppercase;color:#607998;">Datos de acceso</p>
                    <p style="margin:0 0 8px;font-size:16px;color:#123f6b;"><strong>Usuario:</strong> {escape(judge.email)}</p>
                    <p style="margin:0 0 8px;font-size:16px;color:#123f6b;"><strong>Contraseña temporal:</strong></p>
                    <div style="display:inline-block;background:#effaf3;border:1px solid #b7d7c4;border-radius:12px;padding:12px 16px;font-family:Consolas,Monaco,monospace;font-size:20px;font-weight:800;color:#135d37;">{escape(plain_password)}</div>
                    <p style="margin:14px 0 0;font-size:14px;line-height:1.45;color:#5b7189;">Por seguridad, el sistema solicitará cambiar esta contraseña al iniciar sesión.</p>
                  </td>
                </tr>
              </table>
              <p style="margin:0 0 24px;font-size:15px;line-height:1.55;color:#4f6680;">Disponibilidad registrada: <strong>{escape(judge.evaluation_scope_label)}</strong>. La organización usará esta información para asignarte evaluaciones según la necesidad de cada proyecto.</p>
              <p style="margin:0 0 26px;text-align:center;">
                <a href="{escape(login_url)}" style="display:inline-block;background:#f5a11a;color:#123f6b;text-decoration:none;font-size:16px;font-weight:900;padding:14px 26px;border-radius:999px;border:1px solid #da8a0d;">Ingresar al portal</a>
              </p>
            </td>
          </tr>
          <tr>
            <td style="background:#eaf4fb;border-top:1px solid #cfe0f1;padding:16px 30px;color:#58708a;font-size:13px;line-height:1.45;">
              Si no solicitaste este acceso, comunícate con la organización de ExpoTécnica.
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""


def _get_judge_form_secret():
    return (SystemSetting.get_value("judge_form_webhook_secret", "") or "").strip()


def _judge_form_settings():
    secret = _get_judge_form_secret()
    return {
        "form_url": SystemSetting.get_value("judge_form_url", ""),
        "enabled": SystemSetting.get_value("judge_form_enabled", "0") == "1",
        "auto_send_email": SystemSetting.get_value("judge_form_auto_send_email", "1") == "1",
        "has_secret": bool(secret),
        "secret": secret,
        "secret_preview": f"{secret[:6]}...{secret[-4:]}" if len(secret) >= 12 else "",
    }


def _extract_json_value(payload: dict, *keys: str) -> str:
    normalized = {}
    for key, value in (payload or {}).items():
        cleaned_key = re.sub(r"[^a-z0-9]", "", str(key).strip().lower())
        normalized[cleaned_key] = value

    for key in keys:
        cleaned_key = re.sub(r"[^a-z0-9]", "", key.strip().lower())
        value = normalized.get(cleaned_key)
        if value is not None:
            return str(value).strip()
    return ""


def _yes_no_value(value: str) -> str:
    normalized = (value or "").strip().lower()
    if normalized in {"si", "s", "yes", "y", "true", "1", "sí"}:
        return "Si"
    if normalized in {"no", "n", "false", "0"}:
        return "No"
    return value.strip()


def _judge_scope_from_value(value: str) -> tuple[bool, bool, str]:
    normalized = re.sub(r"[^a-z0-9]", "", (value or "").strip().lower())
    if normalized in {"documento", "documentacionescrita", "documentoescrito", "escrita", "virtual"}:
        return True, False, "Solo documento"
    if normalized in {"exposicion", "exposicionpresencial", "oral", "presencial"}:
        return False, True, "Solo exposición"
    return True, True, "Documento y exposición"


def _judge_category_scope_from_value(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]", "", (value or "").strip().lower())
    has_steam = "steam" in normalized
    has_entrepreneurship = "emprend" in normalized or "innovacion" in normalized
    if "ambas" in normalized or (has_steam and has_entrepreneurship):
        return "ambas"
    if has_steam:
        return "steam"
    if has_entrepreneurship:
        return "emprendimiento"
    return "ambas"


def _join_payload_values(*values: str) -> str:
    return " ".join(value.strip() for value in values if value and value.strip())


def _extract_judge_form_payload(payload: dict):
    first_last_name = _extract_json_value(
        payload,
        "primer_apellido",
        "primer apellido",
        "indique su primer apellido",
        "indique su primer apellido 1 apellido",
        "apellido1",
        "apellido_1",
    )
    second_last_name = _extract_json_value(
        payload,
        "segundo_apellido",
        "segundo apellido",
        "indique su segundo apellido",
        "indique su segundo apellido 2 apellido",
        "apellido2",
        "apellido_2",
    )
    given_names = _extract_json_value(
        payload,
        "nombres",
        "nombre",
        "nombre_completo",
        "nombre completo",
        "indique su nombre completo",
        "indique su nombre completo nombre",
        "name",
    )
    full_name = _extract_json_value(payload, "full_name", "full name")
    if not full_name:
        full_name = _join_payload_values(given_names, first_last_name, second_last_name)
    email = _extract_json_value(payload, "email", "correo", "correo electronico", "correo_electronico", "mail")
    phone = _extract_json_value(payload, "phone", "telefono", "teléfono", "celular")
    identity = _extract_json_value(payload, "cedula", "cédula", "identificacion", "identificación", "documento")
    job_title = _extract_json_value(payload, "job_title", "cargo", "puesto", "profesion", "profesión", "especialidad")
    institution = _extract_json_value(payload, "institucion", "institución", "empresa", "lugar de trabajo")
    previous_expo = _yes_no_value(
        _extract_json_value(
            payload,
            "ha_participado",
            "a participado en otras ferias de expotecnica",
            "ha participado en otras ferias de expotecnica",
            "experiencia expotecnica",
        )
    )
    accepts_participation = _yes_no_value(
        _extract_json_value(
            payload,
            "acepta_participar",
            "esta de acuerdo en participar",
            "esta de acuerdo en participar como juez",
            "está de acuerdo en participar",
        )
    )
    modality = (
        _extract_json_value(
            payload,
            "modalidad",
            "alcance_evaluacion",
            "disponibilidad_evaluacion",
            "participacion",
            "participación",
            "tipo de evaluacion",
        )
        or "ambas"
    )
    can_documentation, can_exposition, scope_label = _judge_scope_from_value(modality)
    evaluation_areas = _extract_json_value(payload, "areas", "areas de evaluacion", "categorias", "categorías")
    english_available = _yes_no_value(_extract_json_value(payload, "ingles", "inglés", "evalua ingles", "dominio ingles"))
    can_evaluate_english = english_available == "Si"
    category_scope = _judge_category_scope_from_value(evaluation_areas)
    notes = _extract_json_value(payload, "notes", "observaciones", "comentarios")
    detail_parts = []
    if identity:
        detail_parts.append(f"cedula={identity}")
    if institution:
        detail_parts.append(f"institucion={institution}")
    if previous_expo:
        detail_parts.append(f"participacion_previa={previous_expo}")
    if accepts_participation:
        detail_parts.append(f"acepta_participar={accepts_participation}")
    if modality:
        detail_parts.append(f"disponibilidad={scope_label}")
    if evaluation_areas:
        detail_parts.append(f"areas={evaluation_areas}")
    if english_available:
        detail_parts.append(f"ingles={english_available}")
    if notes:
        detail_parts.append(f"notas={notes}")
    return {
        "full_name": full_name,
        "email": email.lower(),
        "phone": phone,
        "identity": identity,
        "job_title": job_title,
        "institution": institution,
        "previous_expo": previous_expo,
        "notes": "; ".join(detail_parts),
        "accepts_participation": accepts_participation,
        "can_evaluate_documentation": can_documentation,
        "can_evaluate_exposition": can_exposition,
        "can_evaluate_english": can_evaluate_english,
        "category_scope": category_scope,
        "scope_label": scope_label,
    }


def _request_judge_form_token(payload: dict):
    return (
        request.headers.get("X-Expotecnica-Token")
        or request.headers.get("X-Forms-Token")
        or request.args.get("token")
        or (payload or {}).get("token")
        or ""
    ).strip()


def _create_or_update_judge_from_form(payload: dict):
    data = _extract_judge_form_payload(payload)
    if not data["full_name"] or not data["email"]:
        return None, None, "Nombre y correo son obligatorios."
    if "@" not in data["email"]:
        return None, None, "El correo recibido no es valido."
    if data.get("accepts_participation") != "Si":
        return None, None, "La persona no confirmo participacion."

    judge = Judge.query.filter_by(email=data["email"]).first()
    created = judge is None
    temporary_password = ""
    if created:
        temporary_password = secrets.token_urlsafe(10)
        judge = Judge(
            full_name=data["full_name"],
            email=data["email"],
            department="",
            job_title=data["job_title"],
            identity=data["identity"],
            institution=data["institution"],
            previous_expo=data["previous_expo"],
            phone=data["phone"],
            can_evaluate_documentation=data["can_evaluate_documentation"],
            can_evaluate_exposition=data["can_evaluate_exposition"],
            can_evaluate_english=data["can_evaluate_english"],
            category_scope=data["category_scope"],
            registration_notes=data["notes"],
            registered_from_public_form=True,
            role=Judge.ROLE_JUDGE,
            is_admin=False,
            is_active_user=True,
            must_change_password=True,
        )
        judge.set_password(temporary_password)
        db.session.add(judge)
        db.session.flush()
    else:
        temporary_password = secrets.token_urlsafe(10)
        judge.full_name = data["full_name"]
        judge.job_title = data["job_title"] or judge.job_title
        judge.identity = data["identity"] or judge.identity
        judge.institution = data["institution"] or judge.institution
        judge.previous_expo = data["previous_expo"] or judge.previous_expo
        judge.phone = data["phone"] or judge.phone
        judge.can_evaluate_documentation = data["can_evaluate_documentation"]
        judge.can_evaluate_exposition = data["can_evaluate_exposition"]
        judge.can_evaluate_english = data["can_evaluate_english"]
        judge.category_scope = data["category_scope"]
        judge.registration_notes = data["notes"] or judge.registration_notes
        judge.registered_from_public_form = True
        judge.role = Judge.ROLE_JUDGE
        judge.is_admin = False
        judge.is_active_user = True
        judge.must_change_password = True
        judge.set_password(temporary_password)

    detail = f"Solicitud de juez: {judge.full_name} <{judge.email}>"
    if data["notes"]:
        detail = f"{detail}; notas={data['notes'][:300]}"
    log_event("forms.judge_access.created" if created else "forms.judge_access.updated", "judge", entity_id=judge.id, detail=detail)
    db.session.commit()

    credentials_email_sent = False
    if SystemSetting.get_value("judge_form_auto_send_email", "1") == "1":
        credentials_email_sent = _send_judge_credentials_email(judge, temporary_password)
    judge._credentials_email_sent = credentials_email_sent

    return judge, temporary_password, ""


def _send_assignment_email(judge: Judge, project: Project, assignment: Assignment | None = None):
    if not smtp_is_configured():
        return

    scope_label = assignment.scope_label if assignment else "Documento y exposición"
    panel_url = url_for("judge.dashboard", _external=True)
    school_name = SystemSetting.get_value("school_name", "CTP Roberto Gamboa Valverde")
    school_logo = SystemSetting.get_value("school_logo_path", "")
    expo_logo = SystemSetting.get_value("expo_logo_path", "")
    school_logo_url = url_for("static", filename=school_logo, _external=True) if school_logo else ""
    expo_logo_url = url_for("static", filename=expo_logo, _external=True) if expo_logo else ""
    category_label = (project.category or "").strip().title()
    axis_label = project.thematic_axis.name if project.thematic_axis else ""
    project_type_label = project.project_type.name if project.project_type else ""
    subject = "Nuevo proyecto asignado - ExpoTécnica"
    body = (
        f"Hola {judge.full_name},\n\n"
        "Tienes un nuevo proyecto asignado para evaluación:\n"
        f"Proyecto: {project.title}\n"
        f"Categoría: {category_label}\n"
        f"Eje temático: {axis_label or 'No definido'}\n"
        f"Tipo de proyecto: {project_type_label or 'No definido'}\n"
        f"Equipo: {project.team_name}\n"
        f"Alcance: {scope_label}\n\n"
        f"Panel de juez: {panel_url}\n\n"
        "Ingresa al panel de juez para completar la evaluación.\n"
    )
    html_body = _build_assignment_email_html(
        judge=judge,
        project=project,
        scope_label=scope_label,
        panel_url=panel_url,
        school_name=school_name,
        school_logo_url=school_logo_url,
        expo_logo_url=expo_logo_url,
        category_label=category_label,
        axis_label=axis_label,
        project_type_label=project_type_label,
    )
    ok, error = send_email(judge.email, subject, body, html_body=html_body)
    if not ok:
        flash(f"No se pudo enviar correo de asignación: {error}", "error")


def _build_assignment_email_html(
    *,
    judge: Judge,
    project: Project,
    scope_label: str,
    panel_url: str,
    school_name: str,
    school_logo_url: str = "",
    expo_logo_url: str = "",
    category_label: str = "",
    axis_label: str = "",
    project_type_label: str = "",
) -> str:
    logo_cells = ""
    if school_logo_url:
        logo_cells += (
            f'<img src="{escape(school_logo_url)}" alt="{escape(school_name)}" '
            'style="height:64px;max-width:86px;object-fit:contain;margin-right:14px;">'
        )
    if expo_logo_url:
        logo_cells += (
            f'<img src="{escape(expo_logo_url)}" alt="ExpoTécnica" '
            'style="height:64px;max-width:150px;object-fit:contain;">'
        )
    if not logo_cells:
        logo_cells = '<strong style="font-size:22px;color:#ffffff;">ExpoTécnica</strong>'

    axis_html = (
        f"""
                    <tr>
                      <td style="padding:10px 0;color:#607998;font-size:13px;font-weight:800;text-transform:uppercase;letter-spacing:.06em;">Eje temático</td>
                      <td style="padding:10px 0;color:#123f6b;font-size:15px;font-weight:700;text-align:right;">{escape(axis_label)}</td>
                    </tr>"""
        if axis_label
        else ""
    )
    project_type_html = (
        f"""
                    <tr>
                      <td style="padding:10px 0;color:#607998;font-size:13px;font-weight:800;text-transform:uppercase;letter-spacing:.06em;">Tipo</td>
                      <td style="padding:10px 0;color:#123f6b;font-size:15px;font-weight:700;text-align:right;">{escape(project_type_label)}</td>
                    </tr>"""
        if project_type_label
        else ""
    )

    return f"""\
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Nuevo proyecto asignado</title>
</head>
<body style="margin:0;padding:0;background:#eef5fb;font-family:Arial,Helvetica,sans-serif;color:#123f6b;">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#eef5fb;padding:28px 12px;">
    <tr>
      <td align="center">
        <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:700px;background:#ffffff;border:1px solid #cfe0f1;border-radius:22px;overflow:hidden;box-shadow:0 18px 38px rgba(18,63,107,.12);">
          <tr>
            <td style="background:#123f6b;padding:22px 26px;">
              <table role="presentation" width="100%" cellspacing="0" cellpadding="0">
                <tr>
                  <td style="vertical-align:middle;">{logo_cells}</td>
                  <td align="right" style="vertical-align:middle;color:#ffffff;">
                    <div style="font-size:12px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;">ExpoTécnica 2026</div>
                    <div style="font-size:18px;font-weight:800;margin-top:4px;">Nueva asignación</div>
                  </td>
                </tr>
              </table>
            </td>
          </tr>
          <tr>
            <td style="padding:30px;">
              <p style="margin:0 0 8px;font-size:14px;font-weight:800;letter-spacing:.1em;text-transform:uppercase;color:#1f8fb5;">Panel de evaluación</p>
              <h1 style="margin:0 0 12px;font-size:30px;line-height:1.15;color:#123f6b;">Hola, {escape(judge.full_name)}</h1>
              <p style="margin:0 0 22px;font-size:16px;line-height:1.55;color:#4f6680;">Se te asignó un nuevo proyecto para evaluar en el sistema de ExpoTécnica de {escape(school_name)}.</p>

              <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f7fbff;border:1px solid #cfe0f1;border-radius:16px;margin:0 0 22px;">
                <tr>
                  <td style="padding:20px;">
                    <div style="display:inline-block;background:#eaf6ff;border:1px solid #c9dff4;border-radius:999px;color:#123f6b;font-size:12px;font-weight:900;letter-spacing:.08em;text-transform:uppercase;padding:6px 12px;margin-bottom:12px;">{escape(category_label or 'Proyecto')}</div>
                    <h2 style="margin:0 0 6px;font-size:25px;color:#123f6b;line-height:1.15;">{escape(project.title)}</h2>
                    <p style="margin:0 0 18px;color:#5b7189;font-size:15px;">{escape(project.team_name or 'Equipo ExpoTEC')}</p>
                    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-top:1px solid #dce8f5;">
                      <tr>
                        <td style="padding:12px 0;color:#607998;font-size:13px;font-weight:800;text-transform:uppercase;letter-spacing:.06em;">Alcance asignado</td>
                        <td style="padding:12px 0;color:#8b1024;font-size:15px;font-weight:900;text-align:right;">{escape(scope_label)}</td>
                      </tr>{axis_html}{project_type_html}
                    </table>
                  </td>
                </tr>
              </table>

              <p style="margin:0 0 24px;font-size:15px;line-height:1.55;color:#4f6680;">Ingresa al panel de juez para revisar el documento del proyecto y completar las evaluaciones que correspondan.</p>
              <p style="margin:0;text-align:center;">
                <a href="{escape(panel_url)}" style="display:inline-block;background:#f5a11a;color:#123f6b;text-decoration:none;font-size:16px;font-weight:900;padding:14px 28px;border-radius:999px;border:1px solid #da8a0d;">Abrir panel de juez</a>
              </p>
            </td>
          </tr>
          <tr>
            <td style="background:#eaf4fb;border-top:1px solid #cfe0f1;padding:16px 30px;color:#58708a;font-size:13px;line-height:1.45;">
              Si esta asignación no corresponde, comunícate con la organización de ExpoTécnica.
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""


def _delete_project_member_photos(project: Project):
    for member in project.members:
        if not member.photo_url:
            continue
        if member.photo_url.startswith("http://") or member.photo_url.startswith("https://"):
            continue
        try:
            full_path = os.path.join(current_app.static_folder, member.photo_url.replace("/", os.sep))
            if os.path.exists(full_path):
                os.remove(full_path)
        except Exception:  # noqa: BLE001
            continue

    if project.has_real_logo:
        try:
            logo_path = os.path.join(current_app.static_folder, project.project_logo_path.replace("/", os.sep))
            if os.path.exists(logo_path):
                os.remove(logo_path)
        except Exception:  # noqa: BLE001
            pass


def _delete_member_photo_file(member: ProjectMember):
    if not member.photo_url:
        return
    if member.photo_url.startswith("http://") or member.photo_url.startswith("https://"):
        return
    try:
        full_path = os.path.join(current_app.static_folder, member.photo_url.replace("/", os.sep))
        if os.path.exists(full_path):
            os.remove(full_path)
    except Exception:  # noqa: BLE001
        return


def _delete_project_logo_file(project: Project):
    if not project.has_real_logo:
        return
    if project.project_logo_path.startswith("http://") or project.project_logo_path.startswith("https://"):
        return
    try:
        full_path = os.path.join(current_app.static_folder, project.project_logo_path.replace("/", os.sep))
        if os.path.exists(full_path):
            os.remove(full_path)
    except Exception:  # noqa: BLE001
        return


def _delete_project_document_file(project: Project):
    if not project.project_document_path:
        return
    if project.project_document_path.startswith("http://") or project.project_document_path.startswith("https://"):
        return
    try:
        static_root = Path(current_app.static_folder).resolve()
        documents_root = (static_root / "uploads" / "projects" / "documents").resolve()
        full_path = (static_root / project.project_document_path.replace("/", os.sep)).resolve()
        if documents_root not in full_path.parents:
            return
        if full_path.exists():
            full_path.unlink()
    except Exception:  # noqa: BLE001
        return


def _delete_institution_logo_file(relative_path: str):
    if not relative_path:
        return
    if relative_path.startswith("http://") or relative_path.startswith("https://"):
        return
    try:
        full_path = os.path.join(current_app.static_folder, relative_path.replace("/", os.sep))
        if os.path.exists(full_path):
            os.remove(full_path)
    except Exception:  # noqa: BLE001
        return


def _add_member_change(project_id, member_id, action: str, details: str):
    db.session.add(
        ProjectMemberChange(
            project_id=project_id,
            member_id=member_id,
            action=action,
            details=details,
        )
    )


def _resolve_member_academic_fields(form_data):
    section = Section.query.options(joinedload(Section.level)).get(form_data.get("member_section_id", type=int))
    if not section:
        return None, None, "Debes seleccionar una seccion valida."

    level_code = (section.level.code or "").strip() if section.level else ""
    if level_code not in {"10", "11", "12"}:
        return None, None, "La ExpoTécnica solo permite estudiantes de especialidad tecnica (niveles 10, 11 y 12)."

    specialty = Specialty.query.get(form_data.get("member_specialty_id", type=int))
    if not specialty:
        return None, None, "Debes seleccionar la especialidad tecnica del integrante."
    focus_name = specialty.name

    return section.name, focus_name, None


def _handle_action(action: str):
    if action == "create_campaign":
        name = request.form.get("campaign_name", "").strip()
        start_date = _parse_date(request.form.get("campaign_start_date"))
        end_date = _parse_date(request.form.get("campaign_end_date"))
        is_active = _str_to_bool(request.form.get("campaign_is_active"))
        notes = request.form.get("campaign_notes", "").strip()

        if not name or not start_date or not end_date:
            flash("Nombre y fechas de campana son obligatorios.", "error")
        elif start_date > end_date:
            flash("La fecha de inicio no puede ser mayor a la fecha final.", "error")
        elif Campaign.query.filter_by(name=name).first():
            flash("Ya existe una campana con ese nombre.", "error")
        else:
            if is_active:
                Campaign.query.update({"is_active": False})
            campaign = Campaign(name=name, start_date=start_date, end_date=end_date, is_active=is_active, notes=notes)
            db.session.add(campaign)
            log_event("admin.campaign.create", "campaign", detail=f"Campana creada: {name} ({start_date} a {end_date})")
            db.session.commit()
            flash("Campana creada.", "success")

    elif action == "update_campaign":
        campaign_id = request.form.get("campaign_id", type=int)
        campaign = Campaign.query.get(campaign_id) if campaign_id else None
        if not campaign:
            flash("Campana no encontrada.", "error")
        else:
            name = request.form.get("campaign_name", "").strip()
            start_date = _parse_date(request.form.get("campaign_start_date"))
            end_date = _parse_date(request.form.get("campaign_end_date"))
            is_active = _str_to_bool(request.form.get("campaign_is_active"))
            notes = request.form.get("campaign_notes", "").strip()
            duplicate = Campaign.query.filter(Campaign.name == name, Campaign.id != campaign.id).first()
            if not name or not start_date or not end_date:
                flash("Nombre y fechas de campana son obligatorios.", "error")
            elif start_date > end_date:
                flash("La fecha de inicio no puede ser mayor a la fecha final.", "error")
            elif duplicate:
                flash("El nombre de campana ya esta en uso.", "error")
            else:
                if is_active:
                    Campaign.query.update({"is_active": False})
                campaign.name = name
                campaign.start_date = start_date
                campaign.end_date = end_date
                campaign.is_active = is_active
                campaign.notes = notes
                log_event("admin.campaign.update", "campaign", entity_id=campaign.id, detail=f"Campana actualizada: {name}")
                db.session.commit()
                flash("Campana actualizada.", "success")

    elif action == "delete_campaign":
        campaign_id = request.form.get("campaign_id", type=int)
        campaign = Campaign.query.get(campaign_id) if campaign_id else None
        if not campaign:
            flash("Campana no encontrada.", "error")
        elif Project.query.filter_by(campaign_id=campaign.id).count() > 0:
            flash("No puedes eliminar una campana con proyectos asociados.", "error")
        else:
            log_event("admin.campaign.delete", "campaign", entity_id=campaign.id, detail=f"Campana eliminada: {campaign.name}")
            db.session.delete(campaign)
            db.session.commit()
            flash("Campana eliminada.", "success")

    elif action == "activate_campaign":
        campaign_id = request.form.get("campaign_id", type=int)
        campaign = Campaign.query.get(campaign_id) if campaign_id else None
        if not campaign:
            flash("Campana no encontrada.", "error")
        else:
            Campaign.query.update({"is_active": False})
            campaign.is_active = True
            log_event("admin.campaign.activate", "campaign", entity_id=campaign.id, detail=f"Campana activa: {campaign.name}")
            db.session.commit()
            flash("Campana activada.", "success")

    elif action == "deactivate_campaign":
        campaign_id = request.form.get("campaign_id", type=int)
        campaign = Campaign.query.get(campaign_id) if campaign_id else None
        if not campaign:
            flash("Campana no encontrada.", "error")
        else:
            campaign.is_active = False
            log_event("admin.campaign.deactivate", "campaign", entity_id=campaign.id, detail=f"Campana desactivada: {campaign.name}")
            db.session.commit()
            flash("Campana desactivada.", "success")

    elif action == "create_assignment":
        judge_id = request.form.get("judge_id", type=int)
        project_ids = request.form.getlist("project_ids")
        can_documentation, can_exposition = _assignment_scope_from_form()
        if not project_ids:
            single_project_id = request.form.get("project_id", type=int)
            if single_project_id:
                project_ids = [str(single_project_id)]

        judge = Judge.query.get(judge_id) if judge_id else None
        selected_project_ids = []
        for raw_id in project_ids:
            try:
                project_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if project_id not in selected_project_ids:
                selected_project_ids.append(project_id)

        projects = (
            Project.query.options(joinedload(Project.members))
            .filter(Project.id.in_(selected_project_ids))
            .all()
            if selected_project_ids
            else []
        )
        project_map = {project.id: project for project in projects}

        if not judge or not selected_project_ids:
            flash("Debes seleccionar un juez y al menos un proyecto.", "error")
        elif not _assignment_scope_valid(can_documentation, can_exposition):
            flash("Selecciona si el juez evaluará documento, exposición o ambos.", "error")
        elif len(project_map) != len(selected_project_ids):
            flash("Hay proyectos inválidos en la selección.", "error")
        else:
            compatibility_errors = []
            for project_id in selected_project_ids:
                error = _assignment_compatibility_error(
                    judge,
                    project_map[project_id],
                    can_documentation,
                    can_exposition,
                )
                if error:
                    compatibility_errors.append(error)
            if compatibility_errors:
                flash(compatibility_errors[0], "error")
                return

            created_assignments = []
            skipped_projects = []
            for project_id in selected_project_ids:
                project = project_map[project_id]
                if Assignment.query.filter_by(judge_id=judge_id, project_id=project_id).first():
                    skipped_projects.append(project.title)
                    continue

                assignment = Assignment(judge_id=judge_id, project_id=project_id)
                _apply_assignment_scope(assignment, can_documentation, can_exposition)
                db.session.add(assignment)
                created_assignments.append(assignment)
                log_event(
                    "admin.assignment.create",
                    "assignment",
                    detail=(
                        f"Asignacion creada: juez={judge.full_name} <{judge.email}> "
                        f"=> proyecto=#{project.id} '{project.title}', alcance={assignment.scope_label}"
                    ),
                )

            if created_assignments:
                db.session.commit()
                for assignment in created_assignments:
                    _send_assignment_email(judge, assignment.project, assignment)
                flash(f"Asignaciones creadas: {len(created_assignments)}.", "success")
            elif skipped_projects:
                flash("Las asignaciones seleccionadas ya existian.", "error")

            if skipped_projects:
                flash(f"Se omitieron {len(skipped_projects)} proyectos ya asignados.", "error")

    elif action == "delete_assignment":
        assignment_id = request.form.get("assignment_id", type=int)
        assignment = Assignment.query.get(assignment_id) if assignment_id else None
        if not assignment:
            flash("Asignacion no encontrada.", "error")
        else:
            log_event(
                "admin.assignment.delete",
                "assignment",
                entity_id=assignment.id,
                detail=(
                    f"Asignacion eliminada: juez={assignment.judge.full_name} <{assignment.judge.email}> "
                    f"de proyecto=#{assignment.project.id} '{assignment.project.title}'"
                ),
            )
            db.session.delete(assignment)
            db.session.commit()
            flash("Asignacion eliminada.", "success")

    elif action == "update_advisor":
        old_key = request.form.get("old_advisor_key", "").strip()
        merge_key = request.form.get("merge_advisor_key", "").strip()
        new_name = request.form.get("advisor_name", "").strip()
        new_identity = request.form.get("advisor_identity", "").strip()
        new_email = request.form.get("advisor_email", "").strip().lower()
        new_phone = request.form.get("advisor_phone", "").strip()
        if not old_key:
            flash("Clave de tutor no especificada.", "error")
        else:
            def _projects_for_key(key):
                return Project.query.filter(
                    db.or_(
                        Project.advisor_identity == key,
                        db.and_(
                            db.or_(Project.advisor_identity == None, Project.advisor_identity == ""),  # noqa: E711
                            Project.advisor_name == key,
                        ),
                    )
                ).all()

            affected = _projects_for_key(old_key)
            if merge_key and merge_key != old_key:
                affected += _projects_for_key(merge_key)
            for p in affected:
                p.advisor_name = new_name or p.advisor_name
                p.advisor_identity = new_identity
                p.advisor_email = new_email
                if new_phone:
                    p.advisor_phone = new_phone
            db.session.commit()
            flash(f"Tutor actualizado en {len(affected)} proyecto(s).", "ok")

    elif action == "install_package":
        if not current_user.is_superadmin:
            flash("Solo el superadministrador puede instalar dependencias.", "error")
        else:
            package_spec = request.form.get("package_spec", "").strip()
            install_from_req = request.form.get("install_from_requirements") == "1"
            if install_from_req:
                req_path = Path(current_app.root_path).parent / "requirements.txt"
                package_spec = f"-r {req_path}"
            if not package_spec:
                flash("Especifica un paquete para instalar.", "error")
            else:
                try:
                    import sys
                    cmd = [sys.executable, "-m", "pip", "install"] + package_spec.split()
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
                    output = (result.stdout + result.stderr).strip()
                    status = "ok" if result.returncode == 0 else "error"
                    session["dep_last_output"] = output[-4000:] if len(output) > 4000 else output
                    session["dep_last_status"] = status
                    if status == "ok":
                        flash(f"Instalación completada: {package_spec}", "ok")
                    else:
                        flash(f"Error al instalar: {package_spec}", "error")
                except subprocess.TimeoutExpired:
                    flash("Tiempo de espera agotado (120s). El proceso puede seguir corriendo.", "error")
                except Exception as exc:
                    flash(f"Error inesperado: {exc}", "error")

    elif action == "auto_assign":
        max_per_project = request.form.get("max_per_project", type=int) or 2
        max_per_project = max(1, min(max_per_project, 10))
        replace_drafts = request.form.get("replace_drafts") == "1"
        created, skipped = _auto_assign_judges(max_per_project, replace_drafts)
        if created:
            log_event(
                "admin.assignment.auto_assign",
                "assignment",
                detail=f"Auto-asignacion generada: {created} borradores creados, {skipped} proyectos sin espacio, max={max_per_project}",
            )
            flash(
                f"Se generaron {created} asignación(es) en borrador. "
                "Revísalas y confirma para notificar a los jueces.",
                "success",
            )
        else:
            flash("No se generaron nuevas asignaciones. Verifica que haya jueces activos y proyectos disponibles.", "error")

    elif action == "confirm_draft_assignments":
        drafts = Assignment.query.filter_by(status=Assignment.STATUS_DRAFT).all()
        if not drafts:
            flash("No hay asignaciones en borrador para confirmar.", "error")
        else:
            for assignment in drafts:
                assignment.status = Assignment.STATUS_CONFIRMED
            db.session.commit()
            log_event(
                "admin.assignment.confirm_drafts",
                "assignment",
                detail=f"Confirmadas {len(drafts)} asignaciones en borrador. Se envían correos.",
            )
            for assignment in drafts:
                _send_assignment_email(assignment.judge, assignment.project, assignment)
            flash(f"{len(drafts)} asignación(es) confirmadas. Se notificó a los jueces por correo.", "success")

    elif action == "discard_draft_assignments":
        deleted = Assignment.query.filter_by(status=Assignment.STATUS_DRAFT).delete()
        db.session.commit()
        log_event(
            "admin.assignment.discard_drafts",
            "assignment",
            detail=f"Descartadas {deleted} asignaciones en borrador.",
        )
        flash(f"{deleted} asignación(es) en borrador descartadas. No se enviaron correos.", "success")

    elif action == "replace_assignment":
        assignment_id = request.form.get("assignment_id", type=int)
        new_judge_id = request.form.get("judge_id", type=int)
        can_documentation, can_exposition = _assignment_scope_from_form()
        assignment = (
            Assignment.query.options(
                joinedload(Assignment.project).joinedload(Project.members),
                joinedload(Assignment.judge),
            ).get(assignment_id)
            if assignment_id
            else None
        )
        judge = Judge.query.get(new_judge_id) if new_judge_id else None
        if not assignment:
            flash("Asignacion no encontrada.", "error")
        elif not _assignment_scope_valid(can_documentation, can_exposition):
            flash("Selecciona si el juez evaluará documento, exposición o ambos.", "error")
        elif new_judge_id and not judge:
            flash("Debes seleccionar un juez valido.", "error")
        elif judge and assignment.judge_id != judge.id and Assignment.query.filter_by(project_id=assignment.project_id, judge_id=judge.id).first():
            flash("El juez seleccionado ya esta asignado a este proyecto.", "error")
        elif _assignment_compatibility_error(judge if judge else assignment.judge, assignment.project, can_documentation, can_exposition):
            flash(
                _assignment_compatibility_error(
                    judge if judge else assignment.judge,
                    assignment.project,
                    can_documentation,
                    can_exposition,
                ),
                "error",
            )
        else:
            previous_judge = assignment.judge
            target_judge = judge if judge else previous_judge
            if judge and assignment.judge_id != judge.id:
                assignment.judge_id = judge.id
            _apply_assignment_scope(assignment, can_documentation, can_exposition)
            log_event(
                "admin.assignment.replace",
                "assignment",
                entity_id=assignment.id,
                detail=(
                    f"Asignacion reasignada: proyecto=#{assignment.project.id} '{assignment.project.title}' "
                    f"{previous_judge.full_name} <{previous_judge.email}> => {target_judge.full_name} <{target_judge.email}>, "
                    f"alcance={assignment.scope_label}"
                ),
            )
            db.session.commit()
            if judge and previous_judge.id != judge.id:
                _send_assignment_email(judge, assignment.project, assignment)
            flash("Asignacion actualizada correctamente.", "success")

    elif action == "quick_create_assignment_judge":
        full_name = request.form.get("quick_judge_full_name", "").strip()
        email = request.form.get("quick_judge_email", "").strip().lower()
        phone = request.form.get("quick_judge_phone", "").strip()
        manual_password = request.form.get("quick_judge_password", "")
        category_scope = _judge_category_scope_from_value(request.form.get("quick_judge_category_scope", "ambas"))
        can_evaluate_english = _str_to_bool(request.form.get("quick_judge_can_english"))
        project_id = request.form.get("project_id", type=int)
        can_documentation, can_exposition = _assignment_scope_from_form()
        project = Project.query.options(joinedload(Project.members)).get(project_id) if project_id else None

        if not project:
            flash("Proyecto no encontrado para la asignacion.", "error")
        elif not _assignment_scope_valid(can_documentation, can_exposition):
            flash("Selecciona si el juez evaluará documento, exposición o ambos.", "error")
        elif not full_name or not email:
            flash("Nombre y correo son obligatorios para crear el juez.", "error")
        elif Judge.query.filter_by(email=email).first():
            flash("Ya existe un usuario con ese correo.", "error")
        elif manual_password and len(manual_password) < 8:
            flash("La contrasena manual debe tener al menos 8 caracteres.", "error")
        else:
            password_value = manual_password if manual_password else secrets.token_urlsafe(8)
            judge = Judge(
                full_name=full_name,
                email=email,
                department="",
                job_title="",
                phone=phone,
                can_evaluate_documentation=can_documentation,
                can_evaluate_exposition=can_exposition,
                can_evaluate_english=can_evaluate_english,
                category_scope=category_scope,
                role=Judge.ROLE_JUDGE,
                is_admin=False,
                is_active_user=True,
                must_change_password=not bool(manual_password),
            )
            compatibility_error = _assignment_compatibility_error(judge, project, can_documentation, can_exposition)
            if compatibility_error:
                flash(compatibility_error, "error")
                return
            judge.set_password(password_value)
            db.session.add(judge)
            db.session.flush()
            assignment = Assignment(judge_id=judge.id, project_id=project.id)
            _apply_assignment_scope(assignment, can_documentation, can_exposition)
            db.session.add(assignment)
            log_event(
                "admin.user.create_and_assign",
                "judge",
                entity_id=judge.id,
                detail=(
                    f"Juez rapido creado: {judge.full_name} <{judge.email}> "
                    f"para proyecto=#{project.id} '{project.title}'"
                ),
            )
            log_event(
                "admin.assignment.create",
                "assignment",
                detail=(
                    f"Asignacion creada: juez={judge.full_name} <{judge.email}> "
                    f"=> proyecto=#{project.id} '{project.title}', alcance={assignment.scope_label}"
                ),
            )
            db.session.commit()
            _send_judge_credentials_email(judge, password_value)
            _send_assignment_email(judge, project, assignment)
            flash("Juez creado y asignado correctamente.", "success")

    elif action == "create_judge":
        full_name = request.form.get("judge_full_name", "").strip()
        email = request.form.get("judge_email", "").strip().lower()
        role = _valid_role(request.form.get("judge_role"))
        department = _normalize_department_for_role(
            role,
            _valid_department(request.form.get("judge_department")),
        )
        job_title = request.form.get("judge_job_title", "").strip()
        phone = request.form.get("judge_phone", "").strip()
        institution = request.form.get("judge_institution", "").strip()
        identity = request.form.get("judge_identity", "").strip()
        previous_expo = _yes_no_value(request.form.get("judge_previous_expo", "")).strip()
        category_scope = _judge_category_scope_from_value(request.form.get("judge_category_scope", "ambas"))
        can_evaluate_english = _str_to_bool(request.form.get("judge_can_evaluate_english"))
        can_documentation, can_exposition, _scope_label = _judge_scope_from_value(request.form.get("judge_evaluation_scope", "ambas"))
        manual_password = request.form.get("judge_password", "")
        if role != Judge.ROLE_JUDGE:
            identity = ""
            institution = ""
            previous_expo = ""
            category_scope = "ambas"
            can_evaluate_english = False
            can_documentation = False
            can_exposition = False
        if not full_name or not email:
            flash("Nombre y correo son obligatorios.", "error")
        elif _role_requires_department(role) and not department:
            flash("El departamento es obligatorio para usuarios administrativos.", "error")
        elif role == Judge.ROLE_SUPERADMIN and not current_user.is_superadmin:
            flash("Solo un superadministrador puede crear otro superadministrador.", "error")
        elif Judge.query.filter_by(email=email).first():
            flash("Ya existe un usuario con ese correo.", "error")
        elif manual_password and len(manual_password) < 8:
            flash("La contrasena manual debe tener al menos 8 caracteres.", "error")
        else:
            password_value = manual_password if manual_password else secrets.token_urlsafe(8)
            judge = Judge(
                full_name=full_name,
                email=email,
                department=department,
                job_title=job_title,
                identity=identity,
                institution=institution,
                previous_expo=previous_expo,
                phone=phone,
                can_evaluate_documentation=can_documentation,
                can_evaluate_exposition=can_exposition,
                can_evaluate_english=can_evaluate_english,
                category_scope=category_scope,
                role=role,
                is_admin=role in Judge.ADMIN_ROLES,
                is_active_user=True,
                must_change_password=not bool(manual_password),
            )
            judge.set_password(password_value)
            db.session.add(judge)
            log_event(
                "admin.user.create",
                "judge",
                detail=f"Nuevo usuario creado: {full_name} <{email}> role={role} departamento={department}",
            )
            db.session.commit()
            flash("Usuario creado correctamente.", "success")
            if not manual_password:
                _send_judge_credentials_email(judge, password_value)

    elif action == "update_judge":
        judge_id = request.form.get("judge_id", type=int)
        judge = Judge.query.get(judge_id) if judge_id else None
        if not judge:
            flash("Usuario no encontrado.", "error")
        else:
            full_name = request.form.get("judge_full_name", "").strip()
            email = request.form.get("judge_email", "").strip().lower()
            role = _valid_role(request.form.get("judge_role"))
            department = _normalize_department_for_role(
                role,
                _valid_department(request.form.get("judge_department")),
            )
            job_title = request.form.get("judge_job_title", "").strip()
            phone = request.form.get("judge_phone", "").strip()
            institution = request.form.get("judge_institution", "").strip()
            identity = request.form.get("judge_identity", "").strip()
            previous_expo = _yes_no_value(request.form.get("judge_previous_expo", "")).strip()
            category_scope = _judge_category_scope_from_value(request.form.get("judge_category_scope", "ambas"))
            can_evaluate_english = _str_to_bool(request.form.get("judge_can_evaluate_english"))
            can_documentation, can_exposition, _scope_label = _judge_scope_from_value(request.form.get("judge_evaluation_scope", "ambas"))
            is_active_user = _str_to_bool(request.form.get("judge_is_active_user", "1"))
            if role != Judge.ROLE_JUDGE:
                identity = ""
                institution = ""
                previous_expo = ""
                category_scope = "ambas"
                can_evaluate_english = False
                can_documentation = False
                can_exposition = False
            duplicate = Judge.query.filter(Judge.email == email, Judge.id != judge.id).first()
            if not full_name or not email:
                flash("Nombre y correo son obligatorios.", "error")
            elif _role_requires_department(role) and not department:
                flash("El departamento es obligatorio para usuarios administrativos.", "error")
            elif duplicate:
                flash("Ya existe otro usuario con ese correo.", "error")
            elif judge.id == current_user.id and not is_active_user:
                flash("No puedes desactivarte a ti mismo.", "error")
            elif role == Judge.ROLE_SUPERADMIN and not current_user.is_superadmin:
                flash("Solo un superadministrador puede asignar ese rol.", "error")
            elif judge.is_superadmin and not current_user.is_superadmin:
                flash("No puedes modificar un superadministrador.", "error")
            elif judge.id == current_user.id and judge.has_admin_access and role == Judge.ROLE_JUDGE:
                flash("No puedes remover tu propio acceso admin.", "error")
            else:
                judge.full_name = full_name
                judge.email = email
                judge.department = department
                judge.job_title = job_title
                judge.identity = identity
                judge.institution = institution
                judge.previous_expo = previous_expo
                judge.phone = phone
                judge.can_evaluate_documentation = can_documentation
                judge.can_evaluate_exposition = can_exposition
                judge.can_evaluate_english = can_evaluate_english
                judge.category_scope = category_scope
                judge.role = role
                judge.is_admin = role in Judge.ADMIN_ROLES
                judge.is_active_user = is_active_user
                log_event(
                    "admin.user.update",
                    "judge",
                    entity_id=judge.id,
                    detail=(
                        f"Usuario actualizado: {judge.full_name} <{judge.email}> "
                        f"role={judge.role} departamento={judge.department}"
                    ),
                )
                db.session.commit()
                flash("Usuario actualizado.", "success")

    elif action == "reset_judge_password":
        judge_id = request.form.get("judge_id", type=int)
        judge = Judge.query.get(judge_id) if judge_id else None
        if not judge:
            flash("Usuario no encontrado.", "error")
        else:
            temp_password = secrets.token_urlsafe(8)
            judge.set_password(temp_password)
            judge.must_change_password = True
            log_event(
                "admin.user.password.reset",
                "judge",
                entity_id=judge.id,
                detail=f"Contrasena reiniciada para usuario: {judge.full_name} <{judge.email}>",
            )
            db.session.commit()
            flash("Contrasena reiniciada con clave temporal.", "success")
            _send_judge_credentials_email(judge, temp_password)

    elif action == "set_judge_password":
        judge_id = request.form.get("judge_id", type=int)
        judge = Judge.query.get(judge_id) if judge_id else None
        new_password = request.form.get("judge_new_password", "")
        confirm_password = request.form.get("judge_confirm_password", "")
        if not judge:
            flash("Usuario no encontrado.", "error")
        elif len(new_password) < 8:
            flash("La contrasena manual debe tener al menos 8 caracteres.", "error")
        elif new_password != confirm_password:
            flash("La confirmacion de la contrasena no coincide.", "error")
        else:
            judge.set_password(new_password)
            judge.must_change_password = False
            log_event(
                "admin.user.password.set",
                "judge",
                entity_id=judge.id,
                detail=f"Contrasena asignada manualmente a usuario: {judge.full_name} <{judge.email}>",
            )
            db.session.commit()
            flash("Contrasena actualizada manualmente.", "success")

    elif action == "toggle_judge_active":
        judge_id = request.form.get("judge_id", type=int)
        judge = Judge.query.get(judge_id) if judge_id else None
        if not judge:
            flash("Usuario no encontrado.", "error")
        elif judge.id == current_user.id:
            flash("No puedes desactivarte a ti mismo.", "error")
        else:
            judge.is_active_user = not judge.is_active_user
            log_event(
                "admin.user.active.toggle",
                "judge",
                entity_id=judge.id,
                detail=f"Estado activo de usuario {judge.full_name} <{judge.email}> => {judge.is_active_user}",
            )
            db.session.commit()
            flash("Estado de usuario actualizado.", "success")

    elif action == "toggle_judge_admin":
        judge_id = request.form.get("judge_id", type=int)
        judge = Judge.query.get(judge_id) if judge_id else None
        if not judge:
            flash("Usuario no encontrado.", "error")
        elif judge.is_superadmin and not current_user.is_superadmin:
            flash("No puedes cambiar el rol de un superadministrador.", "error")
        elif judge.id == current_user.id and judge.has_admin_access:
            flash("No puedes remover tu propio acceso admin.", "error")
        else:
            judge.role = Judge.ROLE_JUDGE if judge.has_admin_access else Judge.ROLE_ADMIN
            judge.is_admin = judge.role in Judge.ADMIN_ROLES
            log_event(
                "admin.user.role.toggle",
                "judge",
                entity_id=judge.id,
                detail=f"Rol de usuario {judge.full_name} <{judge.email}> => {judge.role}",
            )
            db.session.commit()
            flash("Rol de usuario actualizado.", "success")

    elif action == "delete_judge":
        judge_id = request.form.get("judge_id", type=int)
        judge = Judge.query.get(judge_id) if judge_id else None
        if not judge:
            flash("Usuario no encontrado.", "error")
        elif judge.id == current_user.id:
            flash("No puedes eliminar tu propio usuario.", "error")
        elif judge.is_superadmin and not current_user.is_superadmin:
            flash("No puedes eliminar un superadministrador.", "error")
        else:
            log_event("admin.user.delete", "judge", entity_id=judge.id, detail=f"Usuario eliminado: {judge.full_name} <{judge.email}>")
            db.session.delete(judge)
            db.session.commit()
            flash("Usuario eliminado.", "success")

    elif action == "save_judge_form_settings":
        form_url = request.form.get("judge_form_url", "").strip()
        enabled = _str_to_bool(request.form.get("judge_form_enabled"))
        auto_send_email = _str_to_bool(request.form.get("judge_form_auto_send_email", "1"))
        manual_secret = request.form.get("judge_form_secret", "").strip()

        SystemSetting.set_value("judge_form_url", form_url)
        SystemSetting.set_value("judge_form_enabled", "1" if enabled else "0")
        SystemSetting.set_value("judge_form_auto_send_email", "1" if auto_send_email else "0")
        if manual_secret:
            SystemSetting.set_value("judge_form_webhook_secret", manual_secret)
        elif not _get_judge_form_secret():
            SystemSetting.set_value("judge_form_webhook_secret", secrets.token_urlsafe(32))
        log_event(
            "admin.forms.judge_settings.save",
            "system_setting",
            detail=f"Integracion Forms jueces actualizada: enabled={enabled}, auto_email={auto_send_email}",
        )
        db.session.commit()
        flash("Integracion de Microsoft Forms actualizada.", "success")

    elif action == "rotate_judge_form_secret":
        SystemSetting.set_value("judge_form_webhook_secret", secrets.token_urlsafe(32))
        log_event("admin.forms.judge_secret.rotate", "system_setting", detail="Token webhook Microsoft Forms rotado")
        db.session.commit()
        flash("Token del webhook rotado correctamente.", "success")

    elif action == "update_project":
        project_id = request.form.get("project_id", type=int)
        project = Project.query.get(project_id) if project_id else None
        if not project:
            flash("Proyecto no encontrado.", "error")
        else:
            thematic_axis_id = request.form.get("project_thematic_axis_id", type=int)
            project_type_id = request.form.get("project_project_type_id", type=int)
            thematic_axis = ThematicAxis.query.get(thematic_axis_id) if thematic_axis_id else None
            project_type = ProjectType.query.get(project_type_id) if project_type_id else None
            project.title = request.form.get("project_title", "").strip()
            project.team_name = request.form.get("project_team_name", "").strip()
            project.representative_name = request.form.get("project_representative_name", "").strip()
            project.representative_email = request.form.get("project_representative_email", "").strip().lower()
            project.description = request.form.get("project_description", "").strip()
            if not all([project.title, project.team_name, project.representative_name, project.representative_email]):
                flash("Campos obligatorios incompletos en proyecto.", "error")
            elif not thematic_axis or not thematic_axis.is_active:
                flash("Debes seleccionar un eje tematico valido.", "error")
            elif not project_type or not project_type.is_active:
                flash("Debes seleccionar un tipo de proyecto valido.", "error")
            else:
                project.thematic_axis_id = thematic_axis.id
                project.project_type_id = project_type.id
                log_event(
                    "admin.project.update",
                    "project",
                    entity_id=project.id,
                    detail=(
                        f"Proyecto actualizado: #{project.id} '{project.title}' "
                        f"(equipo: {project.team_name}, eje={thematic_axis.name}, tipo={project_type.name})"
                    ),
                )
                db.session.commit()
                flash("Proyecto actualizado.", "success")

    elif action == "update_project_logistics":
        project_id = request.form.get("project_id", type=int)
        project = Project.query.get(project_id) if project_id else None
        if not project:
            flash("Proyecto no encontrado.", "error")
        else:
            status = request.form.get("logistics_status", "").strip()
            valid_status = {code for code, _ in LOGISTICS_STATUSES}
            if status not in valid_status:
                flash("Estado logistico invalido.", "error")
            else:
                project.is_active = _str_to_bool(request.form.get("project_is_active", "1"))
                project.logistics_document_ok = _str_to_bool(request.form.get("logistics_document_ok"))
                project.logistics_logo_ok = _str_to_bool(request.form.get("logistics_logo_ok"))
                project.logistics_photos_ok = _str_to_bool(request.form.get("logistics_photos_ok"))
                project.logistics_registration_form_signed_ok = _str_to_bool(request.form.get("logistics_registration_form_signed_ok"))
                project.logistics_student_consents_signed_ok = _str_to_bool(request.form.get("logistics_student_consents_signed_ok"))
                project.logistics_requirements_reviewed_ok = _str_to_bool(request.form.get("logistics_requirements_reviewed_ok"))
                missing_items = _project_logistics_missing_items(project)
                forced_incomplete = False
                if status == "completo" and missing_items:
                    project.logistics_status = "incompleto"
                    forced_incomplete = True
                    flash("No se puede marcar como completo. Pendientes: " + ", ".join(missing_items) + ".", "error")
                else:
                    project.logistics_status = status
                project.logistics_notes = request.form.get("logistics_notes", "").strip()
                log_event(
                    "admin.project.logistics.update",
                    "project",
                    entity_id=project.id,
                    detail=(
                        f"Proyecto #{project.id} '{project.title}' => activo={project.is_active}, status={project.logistics_status}, "
                        f"doc={project.logistics_document_ok}, logo={project.logistics_logo_ok}, "
                        f"fotos={project.logistics_photos_ok}, formulario_fisico={project.logistics_registration_form_signed_ok}, "
                        f"consentimientos={project.logistics_student_consents_signed_ok}, requisitos={project.logistics_requirements_reviewed_ok}"
                    ),
                )
                db.session.commit()
                if not forced_incomplete:
                    flash("Control logistico actualizado.", "success")

    elif action == "replace_project_document":
        project_id = request.form.get("project_id", type=int)
        project = Project.query.get(project_id) if project_id else None
        document_file = request.files.get("project_document")
        if not project:
            flash("Proyecto no encontrado.", "error")
        elif not document_file or not document_file.filename:
            flash("Debes seleccionar un archivo PDF.", "error")
        else:
            try:
                old_path = project.project_document_path
                new_path = _save_project_document(document_file)
                _delete_project_document_file(project)
                project.project_document_path = new_path
                project.logistics_document_ok = False
                project.logistics_status = "pendiente_revision"
                log_event(
                    "admin.project.document.replace",
                    "project",
                    entity_id=project.id,
                    detail=(
                        f"Documento reemplazado para proyecto #{project.id} '{project.title}'. "
                        f"Anterior={old_path or 'sin documento'} Nuevo={new_path}"
                    ),
                )
                db.session.commit()
                flash("Documento del proyecto reemplazado. Queda pendiente de revision logistica.", "success")
            except ValueError as error:
                flash(str(error), "error")

    elif action == "approve_document_revision":
        revision_id = request.form.get("revision_id", type=int)
        revision = ProjectDocumentRevision.query.get(revision_id) if revision_id else None
        if not revision:
            flash("Solicitud no encontrada.", "error")
        elif revision.status != ProjectDocumentRevision.STATUS_PENDING:
            flash("Esta solicitud ya fue procesada.", "error")
        else:
            project = revision.project
            old_path = project.project_document_path
            revision.replaced_document_path = old_path
            project.project_document_path = revision.document_path
            project.logistics_document_ok = True
            project.logistics_status = "pendiente_revision"
            revision.status = ProjectDocumentRevision.STATUS_APPROVED
            revision.reviewed_by_id = current_user.id
            revision.reviewed_at = datetime.now()
            log_event(
                "admin.project.document_revision.approve",
                "project",
                entity_id=project.id,
                detail=(
                    f"Revision #{revision.id} aprobada para proyecto #{project.id} '{project.title}'. "
                    f"Enviada por '{revision.submitted_by_name}'. "
                    f"Anterior={old_path or 'ninguno'} Nuevo={revision.document_path}"
                ),
            )
            db.session.commit()
            flash(f"Documento aprobado y activado como version oficial del proyecto '{project.title}'.", "success")

    elif action == "reject_document_revision":
        revision_id = request.form.get("revision_id", type=int)
        revision = ProjectDocumentRevision.query.get(revision_id) if revision_id else None
        admin_notes = (request.form.get("admin_notes") or "").strip()
        if not revision:
            flash("Solicitud no encontrada.", "error")
        elif revision.status != ProjectDocumentRevision.STATUS_PENDING:
            flash("Esta solicitud ya fue procesada.", "error")
        else:
            project = revision.project
            revision.status = ProjectDocumentRevision.STATUS_REJECTED
            revision.admin_notes = admin_notes or None
            revision.reviewed_by_id = current_user.id
            revision.reviewed_at = datetime.now()
            log_event(
                "admin.project.document_revision.reject",
                "project",
                entity_id=project.id,
                detail=(
                    f"Revision #{revision.id} rechazada para proyecto #{project.id} '{project.title}'. "
                    f"Enviada por '{revision.submitted_by_name}'. Motivo: {admin_notes or 'sin motivo indicado'}"
                ),
            )
            db.session.commit()
            flash(f"Solicitud rechazada. El documento oficial del proyecto '{project.title}' no fue modificado.", "success")

    elif action == "approve_member_edit":
        import json as _json
        req_id = request.form.get("edit_request_id", type=int)
        edit_req = ProjectMemberEditRequest.query.get(req_id) if req_id else None
        if not edit_req:
            flash("Solicitud no encontrada.", "error")
        elif edit_req.status != ProjectMemberEditRequest.STATUS_PENDING:
            flash("Esta solicitud ya fue procesada.", "error")
        else:
            member = edit_req.member
            if not member:
                flash("El integrante ya no existe en el sistema.", "error")
            else:
                new_vals = _json.loads(edit_req.changes_json)
                member.full_name = new_vals.get("full_name") or member.full_name
                member.identity_number = new_vals.get("identity_number") or None
                if new_vals.get("birth_date"):
                    from datetime import date as _date
                    try:
                        member.birth_date = _date.fromisoformat(new_vals["birth_date"])
                    except (ValueError, TypeError):
                        pass
                member.gender = new_vals.get("gender") or None
                member.specialty = new_vals.get("specialty") or None
                member.section_name = new_vals.get("section_name") or None
                member.has_dining_scholarship = bool(new_vals.get("has_dining_scholarship"))
                member.participates_in_english = bool(new_vals.get("participates_in_english"))
                member.phone = new_vals.get("phone") or None
                member.email = new_vals.get("email") or None
                member.role = new_vals.get("role") or None
                edit_req.status = ProjectMemberEditRequest.STATUS_APPROVED
                edit_req.reviewed_by_id = current_user.id
                edit_req.reviewed_at = datetime.now()
                log_event(
                    "admin.project.member_edit.approve",
                    "project_member",
                    entity_id=member.id,
                    detail=f"Solicitud #{edit_req.id} aprobada para integrante '{member.full_name}' del proyecto #{edit_req.project_id}",
                )
                db.session.commit()
                flash(f"Cambios aprobados y aplicados a '{member.full_name}'.", "success")

    elif action == "reject_member_edit":
        req_id = request.form.get("edit_request_id", type=int)
        edit_req = ProjectMemberEditRequest.query.get(req_id) if req_id else None
        admin_notes = (request.form.get("admin_notes") or "").strip()
        if not edit_req:
            flash("Solicitud no encontrada.", "error")
        elif edit_req.status != ProjectMemberEditRequest.STATUS_PENDING:
            flash("Esta solicitud ya fue procesada.", "error")
        else:
            edit_req.status = ProjectMemberEditRequest.STATUS_REJECTED
            edit_req.admin_notes = admin_notes or None
            edit_req.reviewed_by_id = current_user.id
            edit_req.reviewed_at = datetime.now()
            log_event(
                "admin.project.member_edit.reject",
                "project_member",
                entity_id=edit_req.member_id,
                detail=f"Solicitud #{edit_req.id} rechazada. Motivo: {admin_notes or 'sin motivo indicado'}",
            )
            db.session.commit()
            flash("Solicitud rechazada. Los datos del integrante no fueron modificados.", "success")

    elif action == "send_logistics_reminder":
        if not smtp_is_configured():
            flash("El servidor SMTP no está configurado. Ve a Ajustes → SMTP antes de enviar correos.", "error")
        else:
            from datetime import timedelta
            active_campaign = Campaign.query.filter_by(is_active=True).first()
            deadline = None
            if active_campaign and active_campaign.end_date:
                deadline = active_campaign.end_date - timedelta(days=1)
            institution_name = SystemSetting.get_value("school_name", "ExpoTécnica")
            active_projects = (
                Project.query
                .options(joinedload(Project.members))
                .filter_by(is_active=True)
                .all()
            )
            sent = 0
            failed = 0
            skipped = 0
            for project in active_projects:
                missing_group = _project_logistics_group_missing(project)
                for member in project.members:
                    missing_individual = []
                    if not member.photo_url:
                        missing_individual.append("Foto de perfil")
                    if not missing_group and not missing_individual:
                        continue
                    if not member.email or not member.email.strip():
                        skipped += 1
                        continue
                    subject = f"Recordatorio: Requisitos pendientes — {project.title}"
                    html_body = _render_logistics_reminder_email(
                        member=member,
                        project=project,
                        missing_group=missing_group,
                        missing_individual=missing_individual,
                        deadline=deadline,
                        institution_name=institution_name,
                    )
                    plain_body = (
                        f"Estimado/a {member.full_name},\n\n"
                        f"Tienes requisitos pendientes en el proyecto '{project.title}'.\n\n"
                        f"Grupales: {', '.join(missing_group) if missing_group else 'Ninguno'}\n"
                        f"Individuales: {', '.join(missing_individual) if missing_individual else 'Ninguno'}\n\n"
                        f"{'Fecha límite: ' + deadline.strftime('%d/%m/%Y') if deadline else ''}\n\n"
                        "Organización ExpoTécnica"
                    )
                    ok, _err = send_email(member.email.strip(), subject, plain_body, html_body=html_body)
                    if ok:
                        sent += 1
                    else:
                        failed += 1
            parts = [f"{sent} correo(s) enviado(s)"]
            if failed:
                parts.append(f"{failed} fallo(s)")
            if skipped:
                parts.append(f"{skipped} integrante(s) sin correo registrado omitido(s)")
            log_event(
                "admin.logistics.reminder_sent",
                "projects",
                detail=f"Recordatorio logístico: {sent} enviados, {failed} fallos, {skipped} sin correo",
            )
            flash(". ".join(parts) + ".", "success" if not failed else "warning")

    elif action == "upload_project_logo":
        project_id = request.form.get("project_id", type=int)
        project = Project.query.get(project_id) if project_id else None
        logo_file = request.files.get("project_logo")
        if not project:
            flash("Proyecto no encontrado.", "error")
        elif not logo_file or not logo_file.filename:
            flash("Debes seleccionar un archivo de logo.", "error")
        else:
            try:
                new_path = _save_project_logo(logo_file)
                _delete_project_logo_file(project)
                project.project_logo_path = new_path
                project.logistics_logo_ok = True
                log_event(
                    "admin.project.logo.upload",
                    "project",
                    entity_id=project.id,
                    detail=f"Logo actualizado para proyecto #{project.id} '{project.title}'",
                )
                db.session.commit()
                flash("Logo del proyecto actualizado.", "success")
            except ValueError as error:
                flash(str(error), "error")

    elif action == "delete_project":
        project_id = request.form.get("project_id", type=int)
        project = Project.query.get(project_id) if project_id else None
        if not project:
            flash("Proyecto no encontrado.", "error")
        else:
            _delete_project_member_photos(project)
            log_event("admin.project.delete", "project", entity_id=project.id, detail=f"Proyecto eliminado: {project.title}")
            db.session.delete(project)
            db.session.commit()
            flash("Proyecto eliminado.", "success")

    elif action == "upload_member_photo":
        member_id = request.form.get("member_id", type=int)
        member = ProjectMember.query.get(member_id) if member_id else None
        photo_file = request.files.get("member_photo")
        if not member:
            flash("Integrante no encontrado.", "error")
        elif not photo_file or not photo_file.filename:
            flash("Debes seleccionar una foto para cargar.", "error")
        else:
            try:
                new_path = _save_member_photo(photo_file)
                _delete_member_photo_file(member)
                member.photo_url = new_path
                log_event(
                    "admin.member.photo_upload",
                    "project_member",
                    entity_id=member.id,
                    detail=(
                        f"Foto actualizada de integrante #{member.student_number} '{member.full_name}' "
                        f"en proyecto #{member.project_id} '{member.project.title if member.project else 'N/D'}'"
                    ),
                )
                db.session.commit()
                flash("Foto del integrante actualizada.", "success")
            except ValueError as error:
                flash(str(error), "error")

    elif action == "create_project_member":
        project_id = request.form.get("project_id", type=int)
        project = Project.query.options(joinedload(Project.members)).get(project_id) if project_id else None
        if not project:
            flash("Proyecto no encontrado.", "error")
        else:
            full_name = request.form.get("member_full_name", "").strip()
            identity_number = request.form.get("member_identity_number", "").strip()
            birth_date = _parse_date(request.form.get("member_birth_date"))
            gender = request.form.get("member_gender", "").strip().lower()
            phone = request.form.get("member_phone", "").strip()
            email = request.form.get("member_email", "").strip().lower()
            has_dining_scholarship = _str_to_bool(request.form.get("member_has_dining_scholarship"))
            participates_in_english = _str_to_bool(request.form.get("member_participates_in_english"))
            photo_file = request.files.get("member_photo")
            section_name, specialty, academic_error = _resolve_member_academic_fields(request.form)

            if not full_name:
                flash("El nombre del integrante es obligatorio.", "error")
            elif gender not in {"masculino", "femenino"}:
                flash("Genero invalido. Usa Masculino o Femenino.", "error")
            elif academic_error:
                flash(academic_error, "error")
            elif len(project.members) >= 3:
                flash("Cada proyecto permite un máximo de 3 integrantes.", "error")
            else:
                used_numbers = {member.student_number for member in project.members}
                number = request.form.get("member_student_number", type=int)
                if not number:
                    for candidate in [1, 2, 3]:
                        if candidate not in used_numbers:
                            number = candidate
                            break
                if not number or number < 1 or number > 3:
                    flash("Numero de estudiante invalido. Usa 1, 2 o 3.", "error")
                elif number in used_numbers:
                    flash("Ese numero de estudiante ya esta asignado en el proyecto.", "error")
                else:
                    new_member = ProjectMember(
                        project_id=project.id,
                        student_number=number,
                        full_name=full_name,
                        identity_number=identity_number,
                        birth_date=birth_date,
                        gender=gender,
                        specialty=specialty,
                        section_name=section_name,
                        has_dining_scholarship=has_dining_scholarship,
                        participates_in_english=participates_in_english,
                        phone=phone,
                        email=email,
                    )
                    db.session.add(new_member)
                    db.session.flush()
                    if photo_file and photo_file.filename:
                        try:
                            new_member.photo_url = _save_member_photo(photo_file)
                        except ValueError as error:
                            db.session.rollback()
                            flash(str(error), "error")
                            return
                    _add_member_change(
                        project.id,
                        new_member.id,
                        "created",
                        f"Integrante agregado: #{number} {new_member.full_name}",
                    )
                    log_event(
                        "admin.member.create",
                        "project_member",
                        entity_id=new_member.id,
                        detail=(
                            f"Integrante agregado: #{new_member.student_number} '{new_member.full_name}' "
                            f"en proyecto #{project.id} '{project.title}'"
                        ),
                    )
                    db.session.commit()
                    flash("Integrante agregado.", "success")

    elif action == "update_project_member":
        member_id = request.form.get("member_id", type=int)
        member = ProjectMember.query.options(joinedload(ProjectMember.project).joinedload(Project.members)).get(member_id) if member_id else None
        if not member:
            flash("Integrante no encontrado.", "error")
        else:
            full_name = request.form.get("member_full_name", "").strip()
            number = request.form.get("member_student_number", type=int)
            gender = request.form.get("member_gender", "").strip().lower()
            photo_file = request.files.get("member_photo")
            section_name, specialty, academic_error = _resolve_member_academic_fields(request.form)
            if not full_name:
                flash("El nombre del integrante es obligatorio.", "error")
            elif gender not in {"masculino", "femenino"}:
                flash("Genero invalido. Usa Masculino o Femenino.", "error")
            elif academic_error:
                flash(academic_error, "error")
            elif not number or number < 1 or number > 3:
                flash("Numero de estudiante invalido. Usa 1, 2 o 3.", "error")
            else:
                duplicate = None
                if number != member.student_number:
                    duplicate = ProjectMember.query.filter(
                        ProjectMember.project_id == member.project_id,
                        ProjectMember.student_number == number,
                        ProjectMember.id != member.id,
                    ).first()
                if duplicate:
                    flash("Ese numero de estudiante ya esta en uso.", "error")
                else:
                    before = (
                        f"#{member.student_number} {member.full_name} / {member.section_name or 'N/D'} / "
                        f"{member.specialty or 'N/D'}"
                    )
                    member.student_number = number
                    member.full_name = full_name
                    member.identity_number = request.form.get("member_identity_number", "").strip()
                    member.birth_date = _parse_date(request.form.get("member_birth_date"))
                    member.gender = gender
                    member.specialty = specialty
                    member.section_name = section_name
                    member.has_dining_scholarship = _str_to_bool(request.form.get("member_has_dining_scholarship"))
                    if "member_participates_in_english" in request.form:
                        member.participates_in_english = _str_to_bool(request.form.get("member_participates_in_english"))
                    member.phone = request.form.get("member_phone", "").strip()
                    member.email = request.form.get("member_email", "").strip().lower()
                    if photo_file and photo_file.filename:
                        try:
                            new_path = _save_member_photo(photo_file)
                            _delete_member_photo_file(member)
                            member.photo_url = new_path
                        except ValueError as error:
                            flash(str(error), "error")
                            return
                    after = (
                        f"#{member.student_number} {member.full_name} / {member.section_name or 'N/D'} / "
                        f"{member.specialty or 'N/D'}"
                    )
                    _add_member_change(member.project_id, member.id, "updated", f"{before} => {after}")
                    log_event(
                        "admin.member.update",
                        "project_member",
                        entity_id=member.id,
                        detail=(
                            f"Integrante actualizado: #{member.student_number} '{member.full_name}' "
                            f"en proyecto #{member.project_id} '{member.project.title if member.project else 'N/D'}'"
                        ),
                    )
                    db.session.commit()
                    flash("Integrante actualizado.", "success")

    elif action == "delete_project_member":
        member_id = request.form.get("member_id", type=int)
        member = ProjectMember.query.get(member_id) if member_id else None
        if not member:
            flash("Integrante no encontrado.", "error")
        else:
            details = f"Integrante eliminado: #{member.student_number} {member.full_name}"
            _add_member_change(member.project_id, member.id, "deleted", details)
            _delete_member_photo_file(member)
            log_event(
                "admin.member.delete",
                "project_member",
                entity_id=member.id,
                detail=(
                    f"Integrante eliminado: #{member.student_number} '{member.full_name}' "
                    f"de proyecto #{member.project_id} '{member.project.title if member.project else 'N/D'}'"
                ),
            )
            db.session.delete(member)
            db.session.commit()
            flash("Integrante eliminado.", "success")

    elif action == "create_category":
        code = _normalize_code(request.form.get("category_code", ""))
        name = request.form.get("category_name", "").strip()
        sort_order = request.form.get("category_sort_order", type=int) or 0
        exposition_evaluation_type_id = request.form.get("category_rubric_1_evaluation_type_id", type=int)
        documentation_evaluation_type_id = request.form.get("category_rubric_2_evaluation_type_id", type=int)
        exposition_eval_type = EvaluationType.query.get(exposition_evaluation_type_id) if exposition_evaluation_type_id else None
        documentation_eval_type = EvaluationType.query.get(documentation_evaluation_type_id) if documentation_evaluation_type_id else None
        if not code or not name:
            flash("Codigo y nombre de categoria son obligatorios.", "error")
        elif Category.query.filter_by(code=code).first():
            flash("El codigo de categoria ya existe.", "error")
        elif any(
            item and item.code == ENGLISH_EVAL_TYPE_CODE
            for item in [exposition_eval_type, documentation_eval_type]
        ):
            flash("La rubrica de ingles se activa automaticamente por proyecto y no por categoria.", "error")
        elif (validation_error := _validate_category_evaluation_types(exposition_eval_type, documentation_eval_type)):
            flash(validation_error, "error")
        else:
            db.session.add(
                Category(
                    code=code,
                    name=name,
                    sort_order=sort_order,
                    rubric_1_evaluation_type_id=exposition_eval_type.id if exposition_eval_type else None,
                    rubric_2_evaluation_type_id=documentation_eval_type.id if documentation_eval_type else None,
                    is_active=True,
                )
            )
            log_event(
                "admin.category.create",
                "category",
                detail=(
                    f"Categoria creada: code={code} nombre='{name}' "
                    f"rubrica1={exposition_eval_type.code if exposition_eval_type else 'N/D'} "
                    f"rubrica2={documentation_eval_type.code if documentation_eval_type else 'N/D'}"
                ),
            )
            db.session.commit()
            flash("Categoria creada.", "success")

    elif action == "update_category":
        category_id = request.form.get("category_id", type=int)
        category = Category.query.get(category_id) if category_id else None
        if not category:
            flash("Categoria no encontrada.", "error")
        else:
            code = _normalize_code(request.form.get("category_code", ""))
            name = request.form.get("category_name", "").strip()
            sort_order = request.form.get("category_sort_order", type=int) or 0
            is_active = _str_to_bool(request.form.get("category_is_active"))
            exposition_evaluation_type_id = request.form.get("category_rubric_1_evaluation_type_id", type=int)
            documentation_evaluation_type_id = request.form.get("category_rubric_2_evaluation_type_id", type=int)
            exposition_eval_type = EvaluationType.query.get(exposition_evaluation_type_id) if exposition_evaluation_type_id else None
            documentation_eval_type = EvaluationType.query.get(documentation_evaluation_type_id) if documentation_evaluation_type_id else None
            duplicate = Category.query.filter(Category.code == code, Category.id != category.id).first()
            if not code or not name:
                flash("Codigo y nombre de categoria son obligatorios.", "error")
            elif duplicate:
                flash("Codigo de categoria ya en uso.", "error")
            elif any(
                item and item.code == ENGLISH_EVAL_TYPE_CODE
                for item in [exposition_eval_type, documentation_eval_type]
            ):
                flash("La rubrica de ingles se activa automaticamente por proyecto y no por categoria.", "error")
            elif (validation_error := _validate_category_evaluation_types(exposition_eval_type, documentation_eval_type)):
                flash(validation_error, "error")
            else:
                category.code = code
                category.name = name
                category.sort_order = sort_order
                category.is_active = is_active
                category.rubric_1_evaluation_type_id = exposition_eval_type.id if exposition_eval_type else None
                category.rubric_2_evaluation_type_id = documentation_eval_type.id if documentation_eval_type else None
                log_event(
                    "admin.category.update",
                    "category",
                    entity_id=category.id,
                    detail=(
                        f"Categoria actualizada: code={category.code} nombre='{category.name}' "
                        f"activa={category.is_active} "
                        f"rubrica1={exposition_eval_type.code if exposition_eval_type else 'N/D'} "
                        f"rubrica2={documentation_eval_type.code if documentation_eval_type else 'N/D'}"
                    ),
                )
                db.session.commit()
                flash("Categoria actualizada.", "success")

    elif action == "delete_category":
        category_id = request.form.get("category_id", type=int)
        category = Category.query.get(category_id) if category_id else None
        if not category:
            flash("Categoria no encontrada.", "error")
        elif Project.query.filter_by(category=category.code).count() > 0:
            flash("No puedes eliminar una categoria con proyectos asociados.", "error")
        else:
            log_event(
                "admin.category.delete",
                "category",
                entity_id=category.id,
                detail=f"Categoria eliminada: code={category.code} nombre='{category.name}'",
            )
            db.session.delete(category)
            db.session.commit()
            flash("Categoria eliminada.", "success")

    elif action == "create_level":
        code = _normalize_code(request.form.get("level_code", ""))
        name = request.form.get("level_name", "").strip()
        sort_order = request.form.get("level_sort_order", type=int) or 0
        if not code or not name:
            flash("Codigo y nombre de nivel son obligatorios.", "error")
        elif Level.query.filter_by(code=code).first():
            flash("Ese codigo de nivel ya existe.", "error")
        else:
            db.session.add(Level(code=code, name=name, sort_order=sort_order, is_active=True))
            log_event(
                "admin.level.create",
                "level",
                detail=f"Nivel creado: code={code} nombre='{name}' orden={sort_order}",
            )
            db.session.commit()
            flash("Nivel creado.", "success")

    elif action == "update_level":
        level_id = request.form.get("level_id", type=int)
        level = Level.query.get(level_id) if level_id else None
        if not level:
            flash("Nivel no encontrado.", "error")
        else:
            code = _normalize_code(request.form.get("level_code", ""))
            name = request.form.get("level_name", "").strip()
            sort_order = request.form.get("level_sort_order", type=int) or 0
            is_active = _str_to_bool(request.form.get("level_is_active"))
            duplicate = Level.query.filter(Level.code == code, Level.id != level.id).first()
            if not code or not name:
                flash("Codigo y nombre de nivel son obligatorios.", "error")
            elif duplicate:
                flash("Codigo de nivel ya en uso.", "error")
            else:
                level.code = code
                level.name = name
                level.sort_order = sort_order
                level.is_active = is_active
                log_event(
                    "admin.level.update",
                    "level",
                    entity_id=level.id,
                    detail=f"Nivel actualizado: code={level.code} nombre='{level.name}' activo={level.is_active}",
                )
                db.session.commit()
                flash("Nivel actualizado.", "success")

    elif action == "create_section":
        level_id = request.form.get("section_level_id", type=int)
        name = request.form.get("section_name", "").strip()
        sort_order = request.form.get("section_sort_order", type=int) or 0
        level = Level.query.get(level_id) if level_id else None
        if not level or not name:
            flash("Nivel y nombre de seccion son obligatorios.", "error")
        else:
            exists = Section.query.filter_by(level_id=level.id, name=name).first()
            if exists:
                flash("La seccion ya existe en ese nivel.", "error")
            else:
                db.session.add(Section(level_id=level.id, name=name, sort_order=sort_order, is_active=True))
                log_event(
                    "admin.section.create",
                    "section",
                    detail=f"Seccion creada: nivel={level.name} nombre='{name}' orden={sort_order}",
                )
                db.session.commit()
                flash("Seccion creada.", "success")

    elif action == "update_section":
        section_id = request.form.get("section_id", type=int)
        section = Section.query.get(section_id) if section_id else None
        if not section:
            flash("Seccion no encontrada.", "error")
        else:
            level_id = request.form.get("section_level_id", type=int)
            name = request.form.get("section_name", "").strip()
            sort_order = request.form.get("section_sort_order", type=int) or 0
            is_active = _str_to_bool(request.form.get("section_is_active"))
            level = Level.query.get(level_id) if level_id else None
            if not level or not name:
                flash("Nivel y nombre de seccion son obligatorios.", "error")
            else:
                section.level_id = level.id
                section.name = name
                section.sort_order = sort_order
                section.is_active = is_active
                log_event(
                    "admin.section.update",
                    "section",
                    entity_id=section.id,
                    detail=f"Seccion actualizada: nivel={level.name} nombre='{section.name}' activo={section.is_active}",
                )
                db.session.commit()
                flash("Seccion actualizada.", "success")

    elif action == "delete_section":
        section_id = request.form.get("section_id", type=int)
        section = Section.query.get(section_id) if section_id else None
        if not section:
            flash("Seccion no encontrada.", "error")
        elif Project.query.filter_by(section_id=section.id).count() > 0:
            flash("No puedes eliminar una seccion con proyectos asociados.", "error")
        else:
            log_event(
                "admin.section.delete",
                "section",
                entity_id=section.id,
                detail=f"Seccion eliminada: nombre='{section.name}'",
            )
            db.session.delete(section)
            db.session.commit()
            flash("Seccion eliminada.", "success")

    elif action == "create_specialty":
        name = request.form.get("specialty_name", "").strip()
        sort_order = request.form.get("specialty_sort_order", type=int) or 0
        if not name:
            flash("Nombre de especialidad obligatorio.", "error")
        elif Specialty.query.filter_by(name=name).first():
            flash("La especialidad ya existe.", "error")
        else:
            db.session.add(Specialty(name=name, sort_order=sort_order, is_active=True))
            log_event(
                "admin.specialty.create",
                "specialty",
                detail=f"Especialidad creada: nombre='{name}' orden={sort_order}",
            )
            db.session.commit()
            flash("Especialidad creada.", "success")

    elif action == "update_specialty":
        specialty_id = request.form.get("specialty_id", type=int)
        specialty = Specialty.query.get(specialty_id) if specialty_id else None
        if not specialty:
            flash("Especialidad no encontrada.", "error")
        else:
            specialty.name = request.form.get("specialty_name", "").strip()
            specialty.sort_order = request.form.get("specialty_sort_order", type=int) or 0
            specialty.is_active = _str_to_bool(request.form.get("specialty_is_active"))
            if not specialty.name:
                flash("Nombre de especialidad obligatorio.", "error")
            else:
                log_event(
                    "admin.specialty.update",
                    "specialty",
                    entity_id=specialty.id,
                    detail=f"Especialidad actualizada: nombre='{specialty.name}' activa={specialty.is_active}",
                )
                db.session.commit()
                flash("Especialidad actualizada.", "success")

    elif action == "delete_specialty":
        specialty_id = request.form.get("specialty_id", type=int)
        specialty = Specialty.query.get(specialty_id) if specialty_id else None
        if not specialty:
            flash("Especialidad no encontrada.", "error")
        elif Project.query.filter_by(specialty_id=specialty.id).count() > 0:
            flash("No puedes eliminar una especialidad con proyectos asociados.", "error")
        else:
            log_event(
                "admin.specialty.delete",
                "specialty",
                entity_id=specialty.id,
                detail=f"Especialidad eliminada: nombre='{specialty.name}'",
            )
            db.session.delete(specialty)
            db.session.commit()
            flash("Especialidad eliminada.", "success")

    elif action == "create_workshop":
        name = request.form.get("workshop_name", "").strip()
        sort_order = request.form.get("workshop_sort_order", type=int) or 0
        if not name:
            flash("Nombre de taller obligatorio.", "error")
        elif Workshop.query.filter_by(name=name).first():
            flash("El taller ya existe.", "error")
        else:
            db.session.add(Workshop(name=name, sort_order=sort_order, is_active=True))
            log_event(
                "admin.workshop.create",
                "workshop",
                detail=f"Taller creado: nombre='{name}' orden={sort_order}",
            )
            db.session.commit()
            flash("Taller creado.", "success")

    elif action == "update_workshop":
        workshop_id = request.form.get("workshop_id", type=int)
        workshop = Workshop.query.get(workshop_id) if workshop_id else None
        if not workshop:
            flash("Taller no encontrado.", "error")
        else:
            workshop.name = request.form.get("workshop_name", "").strip()
            workshop.sort_order = request.form.get("workshop_sort_order", type=int) or 0
            workshop.is_active = _str_to_bool(request.form.get("workshop_is_active"))
            if not workshop.name:
                flash("Nombre de taller obligatorio.", "error")
            else:
                log_event(
                    "admin.workshop.update",
                    "workshop",
                    entity_id=workshop.id,
                    detail=f"Taller actualizado: nombre='{workshop.name}' activo={workshop.is_active}",
                )
                db.session.commit()
                flash("Taller actualizado.", "success")

    elif action == "delete_workshop":
        workshop_id = request.form.get("workshop_id", type=int)
        workshop = Workshop.query.get(workshop_id) if workshop_id else None
        if not workshop:
            flash("Taller no encontrado.", "error")
        elif Project.query.filter_by(workshop_id=workshop.id).count() > 0:
            flash("No puedes eliminar un taller con proyectos asociados.", "error")
        else:
            log_event(
                "admin.workshop.delete",
                "workshop",
                entity_id=workshop.id,
                detail=f"Taller eliminado: nombre='{workshop.name}'",
            )
            db.session.delete(workshop)
            db.session.commit()
            flash("Taller eliminado.", "success")

    elif action == "create_thematic_axis":
        code = _normalize_code(request.form.get("thematic_axis_code", "")) or _normalize_code(request.form.get("thematic_axis_name", ""))
        name = request.form.get("thematic_axis_name", "").strip()
        description = request.form.get("thematic_axis_description", "").strip()
        sort_order = request.form.get("thematic_axis_sort_order", type=int) or 0
        if not code or not name:
            flash("Codigo y nombre del eje tematico son obligatorios.", "error")
        elif ThematicAxis.query.filter_by(code=code).first():
            flash("El eje tematico ya existe.", "error")
        else:
            db.session.add(ThematicAxis(code=code, name=name, description=description, sort_order=sort_order, is_active=True))
            log_event("admin.thematic_axis.create", "thematic_axis", detail=f"Eje tematico creado: code={code} nombre='{name}'")
            db.session.commit()
            flash("Eje tematico creado.", "success")

    elif action == "update_thematic_axis":
        axis_id = request.form.get("thematic_axis_id", type=int)
        axis = ThematicAxis.query.get(axis_id) if axis_id else None
        if not axis:
            flash("Eje tematico no encontrado.", "error")
        else:
            code = _normalize_code(request.form.get("thematic_axis_code", "")) or _normalize_code(request.form.get("thematic_axis_name", ""))
            name = request.form.get("thematic_axis_name", "").strip()
            duplicate = ThematicAxis.query.filter(ThematicAxis.code == code, ThematicAxis.id != axis.id).first()
            if not code or not name:
                flash("Codigo y nombre del eje tematico son obligatorios.", "error")
            elif duplicate:
                flash("Codigo de eje tematico ya en uso.", "error")
            else:
                axis.code = code
                axis.name = name
                axis.description = request.form.get("thematic_axis_description", "").strip()
                axis.sort_order = request.form.get("thematic_axis_sort_order", type=int) or 0
                axis.is_active = _str_to_bool(request.form.get("thematic_axis_is_active"))
                log_event("admin.thematic_axis.update", "thematic_axis", entity_id=axis.id, detail=f"Eje tematico actualizado: code={axis.code} nombre='{axis.name}' activo={axis.is_active}")
                db.session.commit()
                flash("Eje tematico actualizado.", "success")

    elif action == "delete_thematic_axis":
        axis_id = request.form.get("thematic_axis_id", type=int)
        axis = ThematicAxis.query.get(axis_id) if axis_id else None
        if not axis:
            flash("Eje tematico no encontrado.", "error")
        elif Project.query.filter_by(thematic_axis_id=axis.id).count() > 0:
            flash("No puedes eliminar un eje tematico con proyectos asociados.", "error")
        else:
            log_event("admin.thematic_axis.delete", "thematic_axis", entity_id=axis.id, detail=f"Eje tematico eliminado: nombre='{axis.name}'")
            db.session.delete(axis)
            db.session.commit()
            flash("Eje tematico eliminado.", "success")

    elif action == "create_project_type":
        code = _normalize_code(request.form.get("project_type_code", "")) or _normalize_code(request.form.get("project_type_name", ""))
        name = request.form.get("project_type_name", "").strip()
        description = request.form.get("project_type_description", "").strip()
        sort_order = request.form.get("project_type_sort_order", type=int) or 0
        if not code or not name:
            flash("Codigo y nombre del tipo de proyecto son obligatorios.", "error")
        elif ProjectType.query.filter_by(code=code).first():
            flash("El tipo de proyecto ya existe.", "error")
        else:
            db.session.add(ProjectType(code=code, name=name, description=description, sort_order=sort_order, is_active=True))
            log_event("admin.project_type.create", "project_type", detail=f"Tipo de proyecto creado: code={code} nombre='{name}'")
            db.session.commit()
            flash("Tipo de proyecto creado.", "success")

    elif action == "update_project_type":
        project_type_id = request.form.get("project_type_id", type=int)
        project_type = ProjectType.query.get(project_type_id) if project_type_id else None
        if not project_type:
            flash("Tipo de proyecto no encontrado.", "error")
        else:
            code = _normalize_code(request.form.get("project_type_code", "")) or _normalize_code(request.form.get("project_type_name", ""))
            name = request.form.get("project_type_name", "").strip()
            duplicate = ProjectType.query.filter(ProjectType.code == code, ProjectType.id != project_type.id).first()
            if not code or not name:
                flash("Codigo y nombre del tipo de proyecto son obligatorios.", "error")
            elif duplicate:
                flash("Codigo de tipo de proyecto ya en uso.", "error")
            else:
                project_type.code = code
                project_type.name = name
                project_type.description = request.form.get("project_type_description", "").strip()
                project_type.sort_order = request.form.get("project_type_sort_order", type=int) or 0
                project_type.is_active = _str_to_bool(request.form.get("project_type_is_active"))
                log_event("admin.project_type.update", "project_type", entity_id=project_type.id, detail=f"Tipo de proyecto actualizado: code={project_type.code} nombre='{project_type.name}' activo={project_type.is_active}")
                db.session.commit()
                flash("Tipo de proyecto actualizado.", "success")

    elif action == "delete_project_type":
        project_type_id = request.form.get("project_type_id", type=int)
        project_type = ProjectType.query.get(project_type_id) if project_type_id else None
        if not project_type:
            flash("Tipo de proyecto no encontrado.", "error")
        elif Project.query.filter_by(project_type_id=project_type.id).count() > 0:
            flash("No puedes eliminar un tipo de proyecto con proyectos asociados.", "error")
        else:
            log_event("admin.project_type.delete", "project_type", entity_id=project_type.id, detail=f"Tipo de proyecto eliminado: nombre='{project_type.name}'")
            db.session.delete(project_type)
            db.session.commit()
            flash("Tipo de proyecto eliminado.", "success")

    elif action == "create_evaluation_type":
        name = request.form.get("eval_type_name", "").strip()
        description = request.form.get("eval_type_description", "").strip()
        raw_code = request.form.get("eval_type_code", "").strip()
        code = _normalize_code(raw_code) if raw_code else _normalize_code(name)
        sort_order = request.form.get("eval_type_sort_order", type=int) or 0
        if not code or not name:
            flash("Nombre del tipo de evaluacion es obligatorio.", "error")
        elif EvaluationType.query.filter_by(code=code).first():
            flash("El codigo del tipo de evaluacion ya existe.", "error")
        else:
            db.session.add(
                EvaluationType(
                    code=code,
                    name=name,
                    description=description or name,
                    sort_order=sort_order,
                    is_active=True,
                )
            )
            log_event(
                "admin.evaluation_type.create",
                "evaluation_type",
                detail=f"Tipo de evaluacion creado: code={code} nombre='{name}' orden={sort_order}",
            )
            db.session.commit()
            flash("Tipo de evaluacion creado.", "success")

    elif action == "update_evaluation_type":
        eval_type_id = request.form.get("eval_type_id", type=int)
        eval_type = EvaluationType.query.get(eval_type_id) if eval_type_id else None
        if not eval_type:
            flash("Tipo de evaluacion no encontrado.", "error")
        else:
            raw_code = request.form.get("eval_type_code", "").strip()
            name = request.form.get("eval_type_name", "").strip()
            description = request.form.get("eval_type_description", "").strip()
            code = _normalize_code(raw_code) if raw_code else _normalize_code(name)
            sort_order = request.form.get("eval_type_sort_order", type=int) or 0
            is_active = _str_to_bool(request.form.get("eval_type_is_active"))
            duplicate = EvaluationType.query.filter(EvaluationType.code == code, EvaluationType.id != eval_type.id).first()
            if not code or not name:
                flash("Nombre del tipo de evaluacion es obligatorio.", "error")
            elif duplicate:
                flash("Codigo del tipo de evaluacion ya en uso.", "error")
            else:
                eval_type.code = code
                eval_type.name = name
                eval_type.description = description or name
                eval_type.sort_order = sort_order
                eval_type.is_active = is_active
                log_event(
                    "admin.evaluation_type.update",
                    "evaluation_type",
                    entity_id=eval_type.id,
                    detail=f"Tipo de evaluacion actualizado: code={eval_type.code} nombre='{eval_type.name}' activo={eval_type.is_active}",
                )
                db.session.commit()
                flash("Tipo de evaluacion actualizado.", "success")

    elif action == "delete_evaluation_type":
        eval_type_id = request.form.get("eval_type_id", type=int)
        eval_type = EvaluationType.query.get(eval_type_id) if eval_type_id else None
        if not eval_type:
            flash("Tipo de evaluacion no encontrado.", "error")
        elif RubricCriterion.query.filter_by(evaluation_type_id=eval_type.id).count() > 0:
            flash("Elimina primero las rubricas asociadas.", "error")
        else:
            log_event(
                "admin.evaluation_type.delete",
                "evaluation_type",
                entity_id=eval_type.id,
                detail=f"Tipo de evaluacion eliminado: code={eval_type.code} nombre='{eval_type.name}'",
            )
            db.session.delete(eval_type)
            db.session.commit()
            flash("Tipo de evaluacion eliminado.", "success")

    elif action == "create_rubric":
        evaluation_type_id = request.form.get("rubric_evaluation_type_id", type=int)
        eval_type = EvaluationType.query.get(evaluation_type_id) if evaluation_type_id else None
        if not eval_type:
            flash("Tipo de evaluacion invalido.", "error")
        else:
            section_name = request.form.get("rubric_section_name", "").strip()
            section_sort_order = request.form.get("rubric_section_sort_order", type=int) or 0
            name = request.form.get("rubric_name", "").strip()
            min_score = request.form.get("rubric_min_score", type=int)
            max_score = request.form.get("rubric_max_score", type=int)
            sort_order = request.form.get("rubric_sort_order", type=int) or 0
            if not name or min_score is None or max_score is None or min_score > max_score:
                flash("Datos invalidos para rubrica.", "error")
            else:
                db.session.add(
                    RubricCriterion(
                        evaluation_type_id=eval_type.id,
                        section_name=section_name or None,
                        section_sort_order=section_sort_order,
                        name=name,
                        min_score=min_score,
                        max_score=max_score,
                        sort_order=sort_order,
                        is_active=True,
                    )
                )
                log_event(
                    "admin.rubric.create",
                    "rubric",
                    detail=(
                        f"Rubrica creada: tipo={eval_type.code} seccion='{section_name or 'General'}' "
                        f"nombre='{name}' rango={min_score}-{max_score}"
                    ),
                )
                db.session.commit()
                flash("Rubrica creada.", "success")

    elif action == "update_rubric":
        rubric_id = request.form.get("rubric_id", type=int)
        rubric = RubricCriterion.query.get(rubric_id) if rubric_id else None
        if not rubric:
            flash("Rubrica no encontrada.", "error")
        else:
            rubric.section_name = request.form.get("rubric_section_name", "").strip() or None
            rubric.section_sort_order = request.form.get("rubric_section_sort_order", type=int) or 0
            rubric.name = request.form.get("rubric_name", "").strip()
            rubric.min_score = request.form.get("rubric_min_score", type=int) or 0
            rubric.max_score = request.form.get("rubric_max_score", type=int) or 0
            rubric.sort_order = request.form.get("rubric_sort_order", type=int) or 0
            rubric.is_active = _str_to_bool(request.form.get("rubric_is_active"))
            if not rubric.name or rubric.min_score > rubric.max_score:
                flash("Datos invalidos para rubrica.", "error")
            else:
                log_event(
                    "admin.rubric.update",
                    "rubric",
                    entity_id=rubric.id,
                    detail=(
                        f"Rubrica actualizada: nombre='{rubric.name}' "
                        f"seccion='{rubric.section_name or 'General'}' "
                        f"rango={rubric.min_score}-{rubric.max_score} activa={rubric.is_active}"
                    ),
                )
                db.session.commit()
                flash("Rubrica actualizada.", "success")

    elif action == "delete_rubric":
        rubric_id = request.form.get("rubric_id", type=int)
        rubric = RubricCriterion.query.get(rubric_id) if rubric_id else None
        if not rubric:
            flash("Rubrica no encontrada.", "error")
        else:
            log_event(
                "admin.rubric.delete",
                "rubric",
                entity_id=rubric.id,
                detail=f"Rubrica eliminada: nombre='{rubric.name}' seccion='{rubric.section_name or 'General'}'",
            )
            db.session.delete(rubric)
            db.session.commit()
            flash("Rubrica eliminada.", "success")

    elif action == "delete_evaluation":
        evaluation_id = request.form.get("evaluation_id", type=int)
        evaluation = Evaluation.query.options(joinedload(Evaluation.judge), joinedload(Evaluation.project)).get(evaluation_id) if evaluation_id else None
        if not evaluation:
            flash("Evaluacion no encontrada.", "error")
        else:
            log_event(
                "admin.evaluation.delete",
                "evaluation",
                entity_id=evaluation.id,
                detail=(
                    f"Evaluacion eliminada: proyecto=#{evaluation.project_id} "
                    f"'{evaluation.project.title if evaluation.project else 'N/D'}', "
                    f"juez={evaluation.judge.full_name if evaluation.judge else 'N/D'}, "
                    f"tipo={evaluation.evaluation_type}"
                ),
            )
            db.session.delete(evaluation)
            db.session.commit()
            flash("Evaluacion eliminada.", "success")

    elif action == "save_smtp":
        SystemSetting.set_value("smtp_host", request.form.get("smtp_host", "").strip())
        SystemSetting.set_value("smtp_port", request.form.get("smtp_port", "587").strip() or "587")
        SystemSetting.set_value("smtp_username", request.form.get("smtp_username", "").strip())
        if request.form.get("smtp_password"):
            SystemSetting.set_value("smtp_password", request.form.get("smtp_password", ""))
        SystemSetting.set_value("smtp_from_email", request.form.get("smtp_from_email", "").strip())
        SystemSetting.set_value("smtp_use_tls", "1" if _str_to_bool(request.form.get("smtp_use_tls")) else "0")
        SystemSetting.set_value("smtp_use_ssl", "1" if _str_to_bool(request.form.get("smtp_use_ssl")) else "0")
        log_event("admin.smtp.save", "smtp", detail="Configuracion SMTP actualizada")
        db.session.commit()
        flash("Configuracion SMTP actualizada.", "success")

    elif action == "test_smtp":
        target_email = request.form.get("smtp_test_email", "").strip()
        if not target_email:
            flash("Ingresa un correo destino para prueba SMTP.", "error")
        else:
            ok, error = send_email(target_email, "Prueba SMTP - ExpoTécnica", "Este es un correo de prueba del modulo SMTP de ExpoTécnica.")
            if ok:
                log_event("admin.smtp.test", "smtp", detail=f"Prueba enviada a {target_email}")
                flash("Correo de prueba enviado.", "success")
            else:
                log_event("admin.smtp.test.fail", "smtp", detail=f"Error al enviar a {target_email}: {error}")
                flash(f"No se pudo enviar correo de prueba: {error}", "error")

    elif action == "save_institution":
        name = request.form.get("school_name", "").strip()
        address = request.form.get("school_address", "").strip()
        phone = request.form.get("school_phone", "").strip()
        email = request.form.get("school_email", "").strip()
        logo_file = request.files.get("school_logo")
        expo_logo_file = request.files.get("expo_logo")
        if not name or not email:
            flash("Nombre y correo institucional son obligatorios.", "error")
        else:
            SystemSetting.set_value("school_name", name)
            SystemSetting.set_value("school_address", address)
            SystemSetting.set_value("school_phone", phone)
            SystemSetting.set_value("school_email", email)
            SystemSetting.set_value("expotec_stage", "Institucional")
            for setting_key in [
                "expotec_school_year",
                "expotec_service_type",
                "expotec_program_office",
                "expotec_director_name",
                "expotec_director_email",
                "expotec_technical_coordinator_name",
                "expotec_technical_coordinator_email",
            ]:
                SystemSetting.set_value(setting_key, request.form.get(setting_key, "").strip())
            if logo_file and logo_file.filename:
                try:
                    old_logo = SystemSetting.get_value("school_logo_path", "")
                    new_logo = _save_institution_logo(logo_file)
                    SystemSetting.set_value("school_logo_path", new_logo)
                    _delete_institution_logo_file(old_logo)
                except ValueError as error:
                    flash(str(error), "error")
                    return
            if expo_logo_file and expo_logo_file.filename:
                try:
                    old_expo_logo = SystemSetting.get_value("expo_logo_path", "")
                    new_expo_logo = _save_institution_logo(expo_logo_file)
                    SystemSetting.set_value("expo_logo_path", new_expo_logo)
                    _delete_institution_logo_file(old_expo_logo)
                except ValueError as error:
                    flash(str(error), "error")
                    return
            log_event("admin.institution.save", "institution", detail=f"Datos institucionales actualizados: {name}")
            db.session.commit()
            flash("Informacion institucional actualizada.", "success")

    elif action == "save_maintenance_settings":
        maintenance_enabled = "1" if _str_to_bool(request.form.get("maintenance_enabled")) else "0"
        maintenance_message = request.form.get("maintenance_message", "").strip()
        maintenance_image = request.files.get("maintenance_image")
        if not maintenance_message:
            maintenance_message = "Estamos cargando informacion de proyectos. Vuelve pronto."

        SystemSetting.set_value("maintenance_enabled", maintenance_enabled)
        SystemSetting.set_value("maintenance_message", maintenance_message)
        if maintenance_image and maintenance_image.filename:
            try:
                old_image = SystemSetting.get_value("maintenance_image_path", "")
                new_image = _save_maintenance_image(maintenance_image)
                SystemSetting.set_value("maintenance_image_path", new_image)
                _delete_institution_logo_file(old_image)
            except ValueError as error:
                flash(str(error), "error")
                return
        log_event(
            "admin.maintenance.save",
            "system_setting",
            detail=(
                "Mantenimiento actualizado: "
                f"maintenance_enabled={maintenance_enabled}"
            ),
        )
        db.session.commit()
        flash("Configuracion de mantenimiento actualizada.", "success")

    elif action == "cleanup_expotecnica":
        confirmation = (request.form.get("cleanup_confirmation", "") or "").strip().upper()
        if confirmation != "LIMPIAR EXPOTECNICA":
            flash("Para limpiar ExpoTécnica debes escribir exactamente: LIMPIAR EXPOTECNICA.", "error")
            return

        try:
            backup = _create_database_backup("antes_limpiar")
        except RuntimeError as error:
            flash(f"No se ejecuto la limpieza porque fallo el respaldo: {error}", "error")
            return

        before, deleted_files = _run_expotecnica_cleanup()
        SystemSetting.set_value("maintenance_enabled", "1")
        log_event(
            "admin.maintenance.cleanup_expotecnica",
            "system",
            detail=(
                "Limpieza anual ExpoTécnica ejecutada: "
                f"projects={before['projects']}, members={before['members']}, "
                f"assignments={before['assignments']}, users={before['users']}, "
                f"evaluations_preserved={before['evaluations']}, "
                f"evaluation_scores_preserved={before['evaluation_scores']}, "
                f"member_changes={before['member_changes']}, "
                f"archivos={deleted_files}, respaldo={backup['filename']}"
            ),
        )
        db.session.commit()
        flash(
            f"Respaldo creado: {backup['filename']}. ExpoTécnica limpiada. Se eliminaron "
            f"{before['projects']} proyectos, {before['members']} integrantes, "
            f"{before['assignments']} asignaciones, {before['users']} usuarios no admin "
            f"y {deleted_files} archivo(s). Evaluaciones y rubricas se conservaron. "
            "El sitio quedo en mantenimiento.",
            "success",
        )

    elif action == "backup_database":
        try:
            backup = _create_database_backup("manual")
        except RuntimeError as error:
            flash(f"No se pudo crear el respaldo: {error}", "error")
            return
        log_event(
            "admin.maintenance.backup_database",
            "system",
            detail=f"Respaldo manual creado: {backup['filename']} ({_format_bytes(backup['size_bytes'])})",
        )
        db.session.commit()
        flash(f"Respaldo creado correctamente: {backup['filename']}.", "success")

    elif action == "restore_database":
        filename = (request.form.get("backup_filename") or "").strip()
        confirmation = (request.form.get("restore_confirmation") or "").strip().upper()
        if confirmation not in {"SI", "SÍ", "RESTAURAR RESPALDO"}:
            flash("Para restaurar debes confirmar con SI.", "error")
            return
        try:
            restore_job = _start_database_restore_job(filename)
            log_event(
                "admin.maintenance.restore_database_started",
                "system",
                detail=(
                    f"Restauracion iniciada: {restore_job['filename']} "
                    f"(job={restore_job['job_id']}, respaldo preventivo={restore_job['safety_backup']})"
                ),
            )
            db.session.commit()
        except (RuntimeError, ValueError) as error:
            db.session.rollback()
            flash(f"No se pudo iniciar la restauracion: {error}", "error")
            return
        flash(
            f"Restauracion iniciada en segundo plano: {restore_job['filename']}. "
            "Puedes seguir el avance en la traza visual de esta pantalla.",
            "success",
        )

    elif action == "delete_database_backup":
        filename = (request.form.get("backup_filename") or "").strip()
        confirmation = (request.form.get("delete_backup_confirmation") or "").strip().upper()
        if confirmation not in {"SI", "SÍ", "ELIMINAR RESPALDO"}:
            flash("Para eliminar el respaldo debes confirmar con SI.", "error")
            return
        try:
            deleted = _delete_database_backup(filename)
            log_event(
                "admin.database.backup.delete",
                "system",
                detail=f"Respaldo eliminado: {deleted['filename']} ({_format_bytes(deleted['size_bytes'])})",
            )
            db.session.commit()
        except (RuntimeError, ValueError, OSError) as error:
            db.session.rollback()
            flash(f"No se pudo eliminar el respaldo: {error}", "error")
            return
        flash(f"Respaldo eliminado: {deleted['filename']}.", "success")

    elif action == "database_service_reload":
        result = _gitops_reload_service()
        log_event(
            "admin.database.service.reload" if result["ok"] else "admin.database.service.reload_fail",
            "system",
            detail=result.get("out") or result.get("err") or "Recarga de servicio solicitada desde modulo BD",
        )
        db.session.commit()
        flash(
            "Servicio de aplicacion recargado correctamente." if result["ok"] else f"No se pudo recargar el servicio: {result.get('err') or 'sin detalle'}",
            "success" if result["ok"] else "error",
        )

    elif action == "gitops_refresh":
        result = {"ok": True, "out": "Estado actualizado.", "err": "", "code": 0}
        _save_gitops_result("refresh", result)
        log_event("admin.git.refresh", "system", detail="Refresco de estado Git solicitado")
        db.session.commit()
        flash("Estado Git actualizado.", "success")

    elif action == "gitops_fetch":
        result = _run_git_remote_command(["git", "fetch", "--all", "--prune"], timeout=180)
        _save_gitops_result("fetch", result)
        if result["ok"]:
            log_event("admin.git.fetch", "system", detail="Fetch ejecutado correctamente")
            flash("Fetch completado correctamente.", "success")
        else:
            log_event("admin.git.fetch.fail", "system", detail=result.get("err") or "Error en fetch")
            flash(f"Fetch falló: {result.get('err') or 'sin detalle'}", "error")
        db.session.commit()

    elif action == "gitops_pull_ff":
        branch_result = _run_git_command(["git", "rev-parse", "--abbrev-ref", "HEAD"], timeout=20)
        branch = branch_result["out"] if branch_result["ok"] and branch_result["out"] else "main"
        fetch_result = _run_git_remote_command(["git", "fetch", "--all", "--prune"], timeout=180)
        if fetch_result["ok"]:
            result = _run_git_remote_command(["git", "pull", "--ff-only", "origin", branch], timeout=300)
        else:
            result = fetch_result
        _save_gitops_result("pull_ff", result)
        if result["ok"]:
            log_event("admin.git.pull", "system", detail=f"Pull ff-only aplicado en rama {branch}")
            flash("Pull ff-only completado correctamente.", "success")
        else:
            log_event("admin.git.pull.fail", "system", detail=result.get("err") or "Error en pull")
            flash(f"Pull falló: {result.get('err') or 'sin detalle'}", "error")
        db.session.commit()

    elif action == "gitops_pull_apply":
        branch_result = _run_git_command(["git", "rev-parse", "--abbrev-ref", "HEAD"], timeout=20)
        branch = branch_result["out"] if branch_result["ok"] and branch_result["out"] else "main"
        fetch_result = _run_git_remote_command(["git", "fetch", "--all", "--prune"], timeout=180)
        if fetch_result["ok"]:
            pull_result = _run_git_remote_command(["git", "pull", "--ff-only", "origin", branch], timeout=300)
        else:
            pull_result = fetch_result

        if pull_result["ok"]:
            reload_result = _gitops_reload_service()
            combined = {
                "ok": reload_result["ok"],
                "code": reload_result["code"],
                "out": "\n\n".join(
                    [
                        f"Fetch:\n{fetch_result.get('out') or '(sin salida)'}",
                        f"Pull:\n{pull_result.get('out') or '(sin salida)'}",
                        f"Servicio:\n{reload_result.get('out') or reload_result.get('err') or '(sin salida)'}",
                    ]
                ),
                "err": reload_result.get("err", ""),
            }
            _save_gitops_result("pull_apply_reload", combined)
            if reload_result["ok"]:
                log_event("admin.git.apply", "system", detail=f"Pull aplicado y servicio recargado en rama {branch}")
                flash("Cambios aplicados y servicio recargado.", "success")
            else:
                log_event("admin.git.apply.reload_fail", "system", detail=reload_result.get("err") or "Fallo recargando servicio")
                flash(f"Pull aplicado, pero fallo la recarga: {reload_result.get('err') or 'sin detalle'}", "error")
        else:
            _save_gitops_result("pull_apply_reload", pull_result)
            log_event("admin.git.apply.fail", "system", detail=pull_result.get("err") or "Error en pull")
            flash(f"No se pudieron aplicar cambios: {pull_result.get('err') or 'sin detalle'}", "error")
        db.session.commit()

    elif action == "gitops_revert_commit":
        target_commit = (request.form.get("gitops_target_commit", "") or "").strip()
        confirmation = (request.form.get("gitops_revert_confirm", "") or "").strip().upper()
        reload_after = _str_to_bool(request.form.get("gitops_revert_reload"))
        if confirmation != "REVERTIR":
            flash("Escribe REVERTIR para confirmar el regreso a una versión previa.", "error")
            return
        if not re.fullmatch(r"[0-9a-fA-F]{6,40}", target_commit):
            flash("Indica un commit válido para revertir.", "error")
            return
        status_result = _run_git_command(["git", "status", "--porcelain"], timeout=20)
        if status_result["ok"] and (status_result.get("out") or "").strip():
            result = {
                "ok": False,
                "code": -10,
                "out": "",
                "err": "Hay cambios locales. Limpia o respalda esos cambios antes de revertir.",
            }
            _save_gitops_result("revert_blocked_dirty", result)
            log_event("admin.git.revert.blocked", "system", detail=result["err"])
            db.session.commit()
            flash(result["err"], "error")
            return
        validate_result = _run_git_command(["git", "cat-file", "-e", f"{target_commit}^{{commit}}"], timeout=20)
        if not validate_result["ok"]:
            _save_gitops_result("revert_invalid_commit", validate_result)
            log_event("admin.git.revert.invalid", "system", detail=validate_result.get("err") or target_commit)
            db.session.commit()
            flash("El commit indicado no existe en este repositorio.", "error")
            return
        reset_result = _run_git_command(["git", "reset", "--hard", target_commit], timeout=120)
        if reset_result["ok"] and reload_after:
            reload_result = _gitops_reload_service()
            result = {
                "ok": reload_result["ok"],
                "code": reload_result["code"],
                "out": "\n\n".join(
                    [
                        f"Reset:\n{reset_result.get('out') or '(sin salida)'}",
                        f"Servicio:\n{reload_result.get('out') or reload_result.get('err') or '(sin salida)'}",
                    ]
                ),
                "err": reload_result.get("err", ""),
            }
        else:
            result = reset_result
        _save_gitops_result("revert_commit", result)
        if result["ok"]:
            log_event("admin.git.revert", "system", detail=f"Repositorio regresado al commit {target_commit}")
            flash("Repositorio revertido a la versión seleccionada." + (" Servicio recargado." if reload_after else ""), "success")
        else:
            log_event("admin.git.revert.fail", "system", detail=result.get("err") or "Error en reset")
            flash(f"No se pudo revertir: {result.get('err') or 'sin detalle'}", "error")
        db.session.commit()

    elif action == "gitops_service_check":
        status = _gitops_service_status()
        result = {
            "ok": status["running"] and status["health_ok"],
            "code": 0 if status["running"] and status["health_ok"] else -1,
            "out": (
                f"Estado: {status['status_label']}\n"
                f"PID: {status.get('pid') or 'N/D'}\n"
                f"Bind: {status['bind']}\n"
                f"Health URL: {status['health_url']}\n"
                f"HTTP: {status['http_code']}"
            ),
            "err": "",
        }
        _save_gitops_result("service_check", result)
        log_event("admin.git.service.check", "system", detail=result["out"])
        db.session.commit()
        flash("Estado del servicio actualizado.", "success" if result["ok"] else "error")

    elif action == "gitops_service_reload":
        result = _gitops_reload_service()
        _save_gitops_result("service_reload", result)
        log_event("admin.git.service.reload" if result["ok"] else "admin.git.service.reload_fail", "system", detail=result.get("out") or result.get("err"))
        db.session.commit()
        flash("Servicio recargado correctamente." if result["ok"] else f"No se pudo recargar: {result.get('err') or 'sin detalle'}", "success" if result["ok"] else "error")

    elif action == "gitops_service_restart":
        result = _gitops_restart_service()
        _save_gitops_result("service_restart", result)
        log_event("admin.git.service.restart" if result["ok"] else "admin.git.service.restart_fail", "system", detail=result.get("out") or result.get("err"))
        db.session.commit()
        flash("Servicio reiniciado correctamente." if result["ok"] else f"No se pudo reiniciar: {result.get('err') or 'sin detalle'}", "success" if result["ok"] else "error")

    elif action == "save_gitops_remote":
        remote_url = (request.form.get("gitops_remote_url", "") or "").strip()
        username = (request.form.get("gitops_username", "") or "").strip() or "x-access-token"
        token = (request.form.get("gitops_private_token", "") or "").strip()
        persist_token = _str_to_bool(request.form.get("gitops_persist_token"))

        if not remote_url.startswith("https://") and not remote_url.startswith("git@"):
            flash("La URL remota debe iniciar con https:// o git@.", "error")
            return

        SystemSetting.set_value("gitops_remote_url", remote_url)
        SystemSetting.set_value("gitops_username", username)
        if persist_token and token:
            SystemSetting.set_value("gitops_private_token", token)
        elif not persist_token:
            SystemSetting.set_value("gitops_private_token", "")

        set_url_result = _run_git_command(["git", "remote", "set-url", "origin", remote_url], timeout=40)
        _save_gitops_result("save_remote", set_url_result)
        if set_url_result["ok"]:
            log_event("admin.git.remote.save", "system", detail=f"Remote origin configurado: {remote_url}")
            flash("Configuración remota guardada.", "success")
        else:
            log_event("admin.git.remote.save.fail", "system", detail=set_url_result.get("err") or "Error set-url")
            flash(f"No se pudo configurar origin: {set_url_result.get('err') or 'sin detalle'}", "error")
        db.session.commit()

    elif action == "gitops_test_remote":
        result = _run_git_remote_command(["git", "ls-remote", "--heads", "origin"], timeout=60)
        _save_gitops_result("test_remote", result)
        if result["ok"]:
            log_event("admin.git.remote.test", "system", detail="Conexión remota verificada correctamente")
            flash("Conexión con repositorio remoto verificada.", "success")
        else:
            log_event("admin.git.remote.test.fail", "system", detail=result.get("err") or "Error ls-remote")
            flash(f"Error de conexión remota: {result.get('err') or 'sin detalle'}", "error")
        db.session.commit()

    elif action == "save_permissions_matrix":
        if not current_user.is_superadmin:
            flash("Solo superadministrador puede modificar permisos.", "error")
            return

        updated_map = {}
        for dept_code, _dept_name in USER_DEPARTMENTS:
            selected = ["overview"]
            for module_key in PERMISSION_MANAGEABLE_MODULES:
                if _str_to_bool(request.form.get(f"perm_{dept_code}_{module_key}")):
                    selected.append(module_key)
            updated_map[dept_code] = sorted(set(selected))

        _save_department_module_access(updated_map)
        log_event("admin.permissions.save", "system_setting", detail="Matriz de permisos por departamento actualizada")
        db.session.commit()
        flash("Permisos actualizados correctamente.", "success")


def _base_context(active_page: str, **kwargs):
    restore_jobs = _list_restore_jobs()
    restore_job_running = any(job.get("is_running") for job in restore_jobs)
    restore_safe_mode = active_page in {"maintenance", "database"} and restore_job_running
    database_light_mode = active_page == "database"
    allowed_modules = _allowed_modules_for_current_user()
    admin_menu_items = [
        {"key": key, "endpoint": endpoint, "label": label}
        for key, endpoint, label in ADMIN_MENU_ITEMS
        if key in allowed_modules
    ]
    menu_lookup = {item["key"]: item for item in admin_menu_items}
    admin_menu_groups = []
    for group_label, group_keys in ADMIN_MENU_GROUPS:
        entries = []
        for key in group_keys:
            item = menu_lookup.get(key)
            if not item:
                continue
            entries.append(
                {
                    **item,
                    "icon": ADMIN_MENU_ICONS.get(key, "settings"),
                }
            )
        if entries:
            admin_menu_groups.append({"label": group_label, "items": entries})

    if restore_safe_mode or database_light_mode:
        permission_access_map = {}
        permission_modules = [
            {"key": key, "label": label}
            for key, _endpoint, label in ADMIN_MENU_ITEMS
            if key in PERMISSION_MANAGEABLE_MODULES
        ]
        permission_matrix = []
        judges = []
        campaigns = []
        categories = []
        levels = []
        sections = []
        specialties = []
        workshops = []
        thematic_axes = []
        project_types = []
        projects = []
        assignments = []
        evaluation_types = []
        exposition_evaluation_types = []
        documentation_evaluation_types = []
        pending_document_revisions = []
        pending_member_edit_requests = []
        smtp_settings = {"host": "", "port": "587", "username": "", "from_email": "", "use_tls": True, "use_ssl": False}
        institution_settings = {
            "name": "CTP Roberto Gamboa Valverde",
            "address": "",
            "phone": "",
            "email": "",
            "logo_path": "",
            "expo_logo_path": "",
            "expotec_stage": "Institucional",
            "expotec_school_year": "2026",
            "expotec_service_type": "Tecnico profesional",
            "expotec_program_office": "Direccion de Educacion Tecnica y Capacidades Emprendedoras",
            "expotec_director_name": "",
            "expotec_director_email": "",
            "expotec_technical_coordinator_name": "",
            "expotec_technical_coordinator_email": "",
        }
        maintenance_settings = {
            "maintenance_enabled": False,
            "maintenance_message": "Restauracion de base de datos en proceso.",
            "maintenance_image_path": "",
        }
        cleanup_stats = _safe_cleanup_expotecnica_counts() if database_light_mode and not restore_safe_mode else {
            "projects": 0,
            "members": 0,
            "member_changes": 0,
            "assignments": 0,
            "users": 0,
            "evaluations": 0,
            "evaluation_scores": 0,
        }
    else:
        permission_access_map = _load_department_module_access()
        permission_modules = [
            {"key": key, "label": label}
            for key, _endpoint, label in ADMIN_MENU_ITEMS
            if key in PERMISSION_MANAGEABLE_MODULES
        ]
        permission_matrix = []
        for dept_code, dept_name in USER_DEPARTMENTS:
            enabled = set(permission_access_map.get(dept_code, ["overview"]))
            row = {"code": dept_code, "name": dept_name, "modules": {}}
            for module in permission_modules:
                row["modules"][module["key"]] = module["key"] in enabled
            permission_matrix.append(row)

        judges = Judge.query.order_by(Judge.full_name.asc()).all()
        campaigns = Campaign.query.order_by(Campaign.start_date.desc(), Campaign.id.desc()).all()
        categories = Category.query.order_by(Category.sort_order.asc(), Category.name.asc()).all()
        levels = Level.query.order_by(Level.sort_order.asc()).all()
        sections = Section.query.options(joinedload(Section.level)).order_by(Section.sort_order.asc(), Section.name.asc()).all()
        specialties = Specialty.query.order_by(Specialty.sort_order.asc(), Specialty.name.asc()).all()
        workshops = Workshop.query.order_by(Workshop.sort_order.asc(), Workshop.name.asc()).all()
        thematic_axes = ThematicAxis.query.order_by(ThematicAxis.sort_order.asc(), ThematicAxis.name.asc()).all()
        project_types = ProjectType.query.order_by(ProjectType.sort_order.asc(), ProjectType.name.asc()).all()
        projects = Project.query.options(
            joinedload(Project.members),
            joinedload(Project.assignments),
            joinedload(Project.evaluations),
            joinedload(Project.section),
            joinedload(Project.specialty_ref),
            joinedload(Project.thematic_axis),
            joinedload(Project.project_type),
            joinedload(Project.workshop_ref),
            joinedload(Project.member_changes),
            joinedload(Project.document_revisions),
        ).order_by(Project.created_at.desc()).all()
        pending_document_revisions = (
            ProjectDocumentRevision.query
            .options(joinedload(ProjectDocumentRevision.project))
            .filter(ProjectDocumentRevision.status == ProjectDocumentRevision.STATUS_PENDING)
            .order_by(ProjectDocumentRevision.created_at.asc())
            .all()
        )
        pending_member_edit_requests = (
            ProjectMemberEditRequest.query
            .options(joinedload(ProjectMemberEditRequest.project), joinedload(ProjectMemberEditRequest.member))
            .filter(ProjectMemberEditRequest.status == ProjectMemberEditRequest.STATUS_PENDING)
            .order_by(ProjectMemberEditRequest.created_at.asc())
            .all()
        )
        assignments = Assignment.query.options(joinedload(Assignment.judge), joinedload(Assignment.project)).order_by(Assignment.id.desc()).all()
        evaluation_types = EvaluationType.query.options(joinedload(EvaluationType.rubric_criteria)).order_by(EvaluationType.sort_order.asc(), EvaluationType.name.asc()).all()
        exposition_evaluation_types = [
            eval_type
            for eval_type in evaluation_types
            if eval_type.code != ENGLISH_EVAL_TYPE_CODE and infer_evaluation_type_kind(eval_type) == "exposicion"
        ]
        documentation_evaluation_types = [
            eval_type
            for eval_type in evaluation_types
            if eval_type.code != ENGLISH_EVAL_TYPE_CODE and infer_evaluation_type_kind(eval_type) == "documentacion"
        ]

        smtp_settings = {
            "host": SystemSetting.get_value("smtp_host", ""),
            "port": SystemSetting.get_value("smtp_port", "587"),
            "username": SystemSetting.get_value("smtp_username", ""),
            "from_email": SystemSetting.get_value("smtp_from_email", ""),
            "use_tls": SystemSetting.get_value("smtp_use_tls", "1") == "1",
            "use_ssl": SystemSetting.get_value("smtp_use_ssl", "0") == "1",
        }
        institution_settings = {
            "name": SystemSetting.get_value("school_name", "CTP Roberto Gamboa Valverde"),
            "address": SystemSetting.get_value("school_address", ""),
            "phone": SystemSetting.get_value("school_phone", ""),
            "email": SystemSetting.get_value("school_email", ""),
            "logo_path": SystemSetting.get_value("school_logo_path", ""),
            "expo_logo_path": SystemSetting.get_value("expo_logo_path", ""),
            "expotec_stage": SystemSetting.get_value("expotec_stage", "Institucional"),
            "expotec_school_year": SystemSetting.get_value("expotec_school_year", "2026"),
            "expotec_service_type": SystemSetting.get_value("expotec_service_type", "Tecnico profesional"),
            "expotec_program_office": SystemSetting.get_value(
                "expotec_program_office",
                "Direccion de Educacion Tecnica y Capacidades Emprendedoras",
            ),
            "expotec_director_name": SystemSetting.get_value("expotec_director_name", ""),
            "expotec_director_email": SystemSetting.get_value("expotec_director_email", ""),
            "expotec_technical_coordinator_name": SystemSetting.get_value("expotec_technical_coordinator_name", ""),
            "expotec_technical_coordinator_email": SystemSetting.get_value("expotec_technical_coordinator_email", ""),
        }
        maintenance_settings = {
            "maintenance_enabled": SystemSetting.get_value("maintenance_enabled", "0") == "1",
            "maintenance_message": SystemSetting.get_value(
                "maintenance_message",
                "Estamos cargando informacion de proyectos. Vuelve pronto.",
            ),
            "maintenance_image_path": SystemSetting.get_value("maintenance_image_path", ""),
        }
        cleanup_stats = _cleanup_expotecnica_counts()
    database_backups = _list_database_backups()
    database_backup_summary = _database_backup_storage_summary(database_backups)
    if active_page == "database" and not restore_safe_mode:
        database_diagnostics = _database_diagnostics()
        database_counts = _database_operational_counts()
        database_logs = _database_audit_logs()
    else:
        database_diagnostics = {
            "ok": not restore_job_running,
            "database": "",
            "host": "",
            "port": "",
            "version": "N/D",
            "size_label": "N/D",
            "table_count": 0,
            "missing_tables": [],
            "tables": [],
            "checks": [],
            "error": "",
        }
        database_counts = []
        database_logs = []
    gitops_status = _git_status_snapshot()
    gitops_service = _gitops_service_status()
    if restore_safe_mode or database_light_mode:
        gitops_last = {"action": "", "status": "", "output": "", "ran_at": ""}
        gitops_remote = {
            "remote_url": gitops_status.get("remote") or "",
            "username": "x-access-token",
            "has_token": False,
        }
        gitops_logs = []
        judge_form_logs = []
        judge_form_settings = {
            "enabled": False,
            "auto_create_access": False,
            "default_password": "",
            "review_email": "",
            "webhook_secret": "",
        }
    else:
        gitops_last = {
            "action": SystemSetting.get_value("gitops_last_action", ""),
            "status": SystemSetting.get_value("gitops_last_status", ""),
            "output": SystemSetting.get_value("gitops_last_output", ""),
            "ran_at": SystemSetting.get_value("gitops_last_ran_at", ""),
        }
        gitops_remote = {
            "remote_url": SystemSetting.get_value("gitops_remote_url", "") or (gitops_status.get("remote") or ""),
            "username": SystemSetting.get_value("gitops_username", "x-access-token"),
            "has_token": bool(SystemSetting.get_value("gitops_private_token", "")),
        }
        gitops_logs = (
            SystemAuditLog.query.filter(SystemAuditLog.action.ilike("admin.git%"))
            .order_by(SystemAuditLog.created_at.desc())
            .limit(40)
            .all()
        )
        judge_form_logs = (
            SystemAuditLog.query.filter(SystemAuditLog.action.ilike("%forms.judge%"))
            .order_by(SystemAuditLog.created_at.desc())
            .limit(20)
            .all()
        )
        judge_form_settings = _judge_form_settings()
    judge_form_webhook_url = url_for("public.judge_form_webhook", _external=True)
    judge_public_registration_url = url_for("public.judge_registration_short", _external=True)
    smtp_configured = False if (restore_safe_mode or database_light_mode) else smtp_is_configured()
    overview_metrics = _build_overview_metrics(
        projects,
        assignments,
        logistics_page=kwargs.get("logistics_page", 1),
        pending_revisions=pending_document_revisions if not (restore_safe_mode or database_light_mode) else [],
    )

    return {
        "active_page": active_page,
        "allowed_modules": allowed_modules,
        "admin_menu": admin_menu_items,
        "admin_menu_groups": admin_menu_groups,
        "action_url": url_for("admin.perform_action"),
        "next_url": request.path,
        "judges": judges,
        "campaigns": campaigns,
        "categories": categories,
        "levels": levels,
        "sections": sections,
        "specialties": specialties,
        "workshops": workshops,
        "thematic_axes": thematic_axes,
        "project_types": project_types,
        "category_map": {row.code: row.name for row in categories},
        "projects": projects,
        "assignments": assignments,
        "evaluation_types": evaluation_types,
        "exposition_evaluation_types": exposition_evaluation_types,
        "documentation_evaluation_types": documentation_evaluation_types,
        "user_departments": USER_DEPARTMENTS,
        "user_roles": USER_ROLES,
        "smtp_settings": smtp_settings,
        "smtp_configured": smtp_configured,
        "institution_settings": institution_settings,
        "maintenance_settings": maintenance_settings,
        "cleanup_stats": cleanup_stats,
        "database_backups": database_backups,
        "database_backup_summary": database_backup_summary,
        "database_diagnostics": database_diagnostics,
        "database_counts": database_counts,
        "database_logs": database_logs,
        "restore_jobs": restore_jobs,
        "restore_job_running": restore_job_running,
        "gitops_status": gitops_status,
        "gitops_service": gitops_service,
        "gitops_last": gitops_last,
        "gitops_remote": gitops_remote,
        "gitops_logs": gitops_logs,
        "judge_form_settings": judge_form_settings,
        "judge_form_logs": judge_form_logs,
        "judge_form_webhook_url": judge_form_webhook_url,
        "judge_public_registration_url": judge_public_registration_url,
        "overview_metrics": overview_metrics,
        "logistics_statuses": LOGISTICS_STATUSES,
        "logistics_status_map": {code: label for code, label in LOGISTICS_STATUSES},
        "pending_document_revisions": pending_document_revisions,
        "pending_member_edit_requests": pending_member_edit_requests,
        "permission_modules": permission_modules,
        "permission_matrix": permission_matrix,
        "is_superadmin": current_user.is_superadmin,
    }


def _render(page_template: str, active_page: str, **kwargs):
    return render_template(page_template, **_base_context(active_page, **kwargs))


def _build_judge_pool_context(context: dict) -> dict:
    judge_users = [
        judge
        for judge in context.get("judges", [])
        if judge.effective_role == Judge.ROLE_JUDGE
    ]
    active_judge_users = [judge for judge in judge_users if judge.is_active_user]
    project_map = {project.id: project for project in context.get("projects", [])}
    rows = []
    for judge in judge_users:
        judge_assignments = [
            assignment
            for assignment in context.get("assignments", [])
            if assignment.judge_id == judge.id
        ]
        active_assignments = [
            assignment
            for assignment in judge_assignments
            if project_map.get(assignment.project_id) and project_map[assignment.project_id].is_active
        ]
        warnings = []
        for assignment in active_assignments:
            project = project_map.get(assignment.project_id)
            warning = _assignment_compatibility_error(
                judge,
                project,
                assignment.can_evaluate_documentation,
                assignment.can_evaluate_exposition,
            )
            if warning:
                warnings.append(warning)
        rows.append(
            {
                "judge": judge,
                "assignments": active_assignments,
                "assignment_count": len(active_assignments),
                "warning_count": len(warnings),
                "warnings": warnings[:2],
            }
        )

    def _capability_matrix(judges_subset):
        return {
            "documentacion": {
                "steam": sum(
                    1
                    for judge in judges_subset
                    if judge.can_evaluate_documentation and judge.category_scope_normalized == "steam"
                ),
                "emprendimiento": sum(
                    1
                    for judge in judges_subset
                    if judge.can_evaluate_documentation and judge.category_scope_normalized == "emprendimiento"
                ),
                "ambas": sum(
                    1
                    for judge in judges_subset
                    if judge.can_evaluate_documentation and judge.category_scope_normalized == "ambas"
                ),
            },
            "exposicion": {
                "steam": sum(
                    1
                    for judge in judges_subset
                    if judge.can_evaluate_exposition and judge.category_scope_normalized == "steam"
                ),
                "emprendimiento": sum(
                    1
                    for judge in judges_subset
                    if judge.can_evaluate_exposition and judge.category_scope_normalized == "emprendimiento"
                ),
                "ambas": sum(
                    1
                    for judge in judges_subset
                    if judge.can_evaluate_exposition and judge.category_scope_normalized == "ambas"
                ),
            },
        }

    english_exposition_judges = [
        judge
        for judge in active_judge_users
        if judge.can_evaluate_english and judge.can_evaluate_exposition
    ]

    stats = {
        "total": len(judge_users),
        "active": len(active_judge_users),
        "english": len(english_exposition_judges),
        "steam": sum(1 for judge in active_judge_users if judge.category_scope_normalized == "steam"),
        "entrepreneurship": sum(1 for judge in active_judge_users if judge.category_scope_normalized == "emprendimiento"),
        "both_categories": sum(1 for judge in active_judge_users if judge.category_scope_normalized == "ambas"),
        "documentation_only": sum(1 for judge in active_judge_users if judge.can_evaluate_documentation and not judge.can_evaluate_exposition),
        "exposition_only": sum(1 for judge in active_judge_users if judge.can_evaluate_exposition and not judge.can_evaluate_documentation),
        "both_scopes": sum(1 for judge in active_judge_users if judge.can_evaluate_documentation and judge.can_evaluate_exposition),
    }
    stats["matrix"] = _capability_matrix(active_judge_users)
    stats["english_matrix"] = _capability_matrix(english_exposition_judges)

    def _scope_category_matrix(judges_subset):
        cats = ["steam", "emprendimiento", "ambas"]
        result = {}
        for cat in cats:
            subset = [j for j in judges_subset if j.category_scope_normalized == cat]
            result[cat] = {
                "doc_only": sum(1 for j in subset if j.can_evaluate_documentation and not j.can_evaluate_exposition),
                "expo_only": sum(1 for j in subset if j.can_evaluate_exposition and not j.can_evaluate_documentation),
                "both": sum(1 for j in subset if j.can_evaluate_documentation and j.can_evaluate_exposition),
                "total": len(subset),
            }
        result["totals"] = {
            "doc_only": sum(result[c]["doc_only"] for c in cats),
            "expo_only": sum(result[c]["expo_only"] for c in cats),
            "both": sum(result[c]["both"] for c in cats),
            "total": sum(result[c]["total"] for c in cats),
        }
        return result

    stats["scope_matrix"] = _scope_category_matrix(active_judge_users)
    return {
        "judge_pool_rows": rows,
        "judge_pool_stats": stats,
    }


@admin_required
def perform_action():
    action = request.form.get("action", "").strip()
    if action and _can_perform_action(action):
        _handle_action(action)
    elif action:
        flash("No tienes permisos para ejecutar esta accion.", "error")
    return _redirect_next()


@admin_module_required("overview")
def overview():
    logistics_page = request.args.get("logistics_page", default=1, type=int) or 1
    return _render("admin/overview.html", "overview", logistics_page=logistics_page)


@admin_module_required("assignments")
def assignments_page():
    return _render("admin/assignments.html", "assignments")


@admin_module_required("judge_pool")
def judge_pool_page():
    context = _base_context("judge_pool")
    context.update(_build_judge_pool_context(context))
    return render_template("admin/judge_pool.html", **context)


@admin_module_required("judges")
def judges_page():
    return _render("admin/judges.html", "judges")


@admin_module_required("permissions")
def permissions_page():
    return _render("admin/permissions.html", "permissions")


@admin_module_required("categories")
def categories_page():
    return _render("admin/categories.html", "categories")


@admin_module_required("academic")
def academic_page():
    return _render("admin/academic.html", "academic")


@admin_module_required("rubrics")
def rubrics_page():
    return _render("admin/rubrics.html", "rubrics")


@admin_module_required("projects")
def _build_advisor_stats(projects):
    from collections import defaultdict
    buckets = defaultdict(lambda: {"name": "", "email": "", "steam": 0, "emprendimiento": 0, "total": 0})
    for p in projects:
        key = (p.advisor_identity or "").strip() or (p.advisor_name or "sin_cedula").strip()
        if not key:
            continue
        b = buckets[key]
        b["name"] = " ".join(w.capitalize() for w in (p.advisor_name or "Sin nombre").split())
        b["email"] = p.advisor_email or ""
        b["identity"] = p.advisor_identity or ""
        cat = (p.category or "").lower()
        if "steam" in cat:
            b["steam"] += 1
        elif "emprend" in cat:
            b["emprendimiento"] += 1
        b["total"] += 1
    return sorted(buckets.values(), key=lambda r: -r["total"])


def projects_page():
    context = _base_context("projects")
    context["advisor_stats"] = _build_advisor_stats(context.get("projects", []))
    return render_template("admin/projects.html", **context)


@admin_module_required("projects")
def logistics_reminder_page():
    context = _base_context("projects")
    reminder_data = _build_logistics_reminder_data(context["projects"], context["campaigns"])
    preview_html = ""
    for row in reminder_data["reminder_rows"]:
        for member_row in row["recipients"]:
            preview_html = _render_logistics_reminder_email(
                member=member_row["member"],
                project=row["project"],
                missing_group=row["missing_group"],
                missing_individual=member_row["missing_individual"],
                deadline=reminder_data["deadline"],
                institution_name=reminder_data["institution_name"],
            )
            break
        if preview_html:
            break
    context["reminder_data"] = reminder_data
    context["preview_email_html"] = preview_html
    return render_template("admin/logistics_reminder.html", **context)


@admin_module_required("evaluations")
def evaluations_page():
    context = _base_context("evaluations")
    context.update(build_admin_evaluation_overview())
    return render_template("admin/evaluations.html", **context)


@admin_module_required("documents")
def documents_page():
    context = _base_context("documents")
    context.update(_build_documents_context())
    return render_template("admin/documents.html", **context)


@admin_module_required("evaluations")
def evaluation_report_project_preview(project_id: int):
    acta_data = _build_project_acta_context(project_id)
    if not acta_data:
        abort(404)
    context = _base_context("evaluations")
    context.update(
        {
            "acta_data": acta_data,
            "report_generated_at": datetime.now(),
        }
    )
    return render_template("admin/evaluations_report_project.html", **context)


@admin_module_required("evaluations")
def evaluation_report_project_pdf(project_id: int):
    acta_data = _build_project_acta_context(project_id)
    if not acta_data:
        abort(404)
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.evaluation_report_project_preview", project_id=project_id))
    pdf_bytes = _render_project_acta_pdf(acta_data)
    safe_name = _normalize_code(acta_data["project"].title) or f"proyecto_{project_id}"
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=f"acta_{safe_name}.pdf",
    )


@admin_module_required("evaluations")
def evaluation_report_project_download(project_id: int):
    acta_data = _build_project_acta_context(project_id)
    if not acta_data:
        abort(404)
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.evaluation_report_project_preview", project_id=project_id))
    pdf_bytes = _render_project_acta_pdf(acta_data)
    safe_name = _normalize_code(acta_data["project"].title) or f"proyecto_{project_id}"
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"acta_{safe_name}.pdf",
    )


@admin_module_required("evaluations")
def evaluation_report_all_preview():
    report_context = _build_all_projects_acta_context()
    context = _base_context("evaluations")
    context.update(report_context)
    return render_template("admin/evaluations_report_all.html", **context)


@admin_module_required("evaluations")
def evaluation_report_all_pdf():
    report_context = _build_all_projects_acta_context()
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.evaluation_report_all_preview"))
    pdf_bytes = _render_all_projects_acta_pdf(report_context)
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=False,
        download_name="acta_general_expotecnica.pdf",
    )


@admin_module_required("evaluations")
def evaluation_report_all_download():
    report_context = _build_all_projects_acta_context()
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.evaluation_report_all_preview"))
    pdf_bytes = _render_all_projects_acta_pdf(report_context)
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="acta_general_expotecnica.pdf",
    )


@admin_module_required("documents")
def evaluation_report_judge_preview(judge_id: int):
    report_context = _build_judge_acta_context(judge_id)
    if not report_context:
        abort(404)
    context = _base_context("documents")
    context.update(report_context)
    return render_template("admin/evaluations_report_judge.html", **context)


@admin_module_required("documents")
def evaluation_report_judge_pdf(judge_id: int):
    report_context = _build_judge_acta_context(judge_id)
    if not report_context:
        abort(404)
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.evaluation_report_judge_preview", judge_id=judge_id))
    pdf_bytes = _render_judge_acta_pdf(report_context)
    safe_name = _normalize_code(report_context["judge"].full_name) or f"juez_{judge_id}"
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=f"acta_juez_{safe_name}.pdf",
    )


@admin_module_required("documents")
def evaluation_report_judge_download(judge_id: int):
    report_context = _build_judge_acta_context(judge_id)
    if not report_context:
        abort(404)
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.evaluation_report_judge_preview", judge_id=judge_id))
    pdf_bytes = _render_judge_acta_pdf(report_context)
    safe_name = _normalize_code(report_context["judge"].full_name) or f"juez_{judge_id}"
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"acta_juez_{safe_name}.pdf",
    )


def _build_assignments_report_rows(context: dict) -> list[dict]:
    judges = [j for j in context.get("judges", []) if j.effective_role == Judge.ROLE_JUDGE]
    project_map = {p.id: p for p in context.get("projects", [])}
    category_map = context.get("category_map", {})
    rows = []
    for judge in judges:
        assignments = [
            a for a in context.get("assignments", [])
            if a.judge_id == judge.id
        ]
        for a in assignments:
            project = project_map.get(a.project_id)
            rows.append({
                "judge_name": judge.full_name,
                "judge_email": judge.email,
                "eval_scope": judge.evaluation_scope_label,
                "category_scope": judge.category_scope_label,
                "project_title": project.title if project else f"Proyecto #{a.project_id}",
                "project_category": category_map.get(project.category, project.category) if project else "—",
                "status": "Borrador" if a.is_draft else "Confirmado",
            })
    rows.sort(key=lambda r: (r["judge_name"], r["project_title"]))
    return rows


@admin_module_required("assignments")
def assignments_report_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        flash("Instala openpyxl en el servidor para exportar Excel.", "error")
        return redirect(url_for("admin.assignments_page"))

    context = _base_context("assignments")
    rows = _build_assignments_report_rows(context)

    wb = Workbook()

    # ── Hoja 1: Datos ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Asignaciones"

    headers = ["Juez", "Correo", "Alcance evaluación", "Alcance categoría", "Proyecto", "Categoría proyecto", "Estado"]
    col_widths = [30, 36, 22, 20, 55, 20, 14]

    header_fill = PatternFill("solid", fgColor="1A4A7A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    draft_fill  = PatternFill("solid", fgColor="FFF3CD")
    thin = Side(style="thin", color="C8DDF0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap   = Alignment(vertical="top", wrap_text=True)

    ws.append(headers)
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 22

    for row in rows:
        ws.append([
            row["judge_name"], row["judge_email"], row["eval_scope"],
            row["category_scope"], row["project_title"],
            row["project_category"], row["status"],
        ])
        r = ws.max_row
        is_draft = row["status"] == "Borrador"
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = wrap
            cell.border = border
            if is_draft:
                cell.fill = draft_fill

    # Excel table with autofilter
    last_row = ws.max_row
    last_col = get_column_letter(len(headers))
    tbl = Table(displayName="Asignaciones", ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.freeze_panes = "A2"

    # ── Hoja 2: Resumen por juez (base para tabla dinámica) ────────────────
    ws2 = wb.create_sheet("Resumen por juez")
    summary: dict[str, dict] = {}
    for row in rows:
        key = row["judge_name"]
        if key not in summary:
            summary[key] = {"Juez": key, "Correo": row["judge_email"],
                            "Alcance": row["eval_scope"], "Confirmadas": 0, "Borradores": 0, "Total": 0}
        if row["status"] == "Confirmado":
            summary[key]["Confirmadas"] += 1
        else:
            summary[key]["Borradores"] += 1
        summary[key]["Total"] += 1

    summary_headers = ["Juez", "Correo", "Alcance", "Confirmadas", "Borradores", "Total"]
    summary_widths  = [30, 36, 22, 14, 14, 10]
    ws2.append(summary_headers)
    for col_idx, (h, w) in enumerate(zip(summary_headers, summary_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    for s in sorted(summary.values(), key=lambda x: -x["Total"]):
        ws2.append([s["Juez"], s["Correo"], s["Alcance"], s["Confirmadas"], s["Borradores"], s["Total"]])
        r = ws2.max_row
        for col_idx in range(1, len(summary_headers) + 1):
            ws2.cell(row=r, column=col_idx).border = border

    last_row2 = ws2.max_row
    last_col2 = get_column_letter(len(summary_headers))
    tbl2 = Table(displayName="ResumenJuez", ref=f"A1:{last_col2}{last_row2}")
    tbl2.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws2.add_table(tbl2)
    ws2.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="asignaciones_por_juez.xlsx",
    )


@admin_module_required("assignments")
def assignments_report_pdf():
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.assignments_page"))
    context = _base_context("assignments")
    rows = _build_assignments_report_rows(context)
    institution = _institution_name()
    buffer = BytesIO()

    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=13, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#5d7897"), spaceAfter=10)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=10)

    elements = [
        Paragraph(_pdf_normalize_text(f"Reporte de Asignaciones por Juez"), title_style),
        Paragraph(_pdf_normalize_text(institution), sub_style),
    ]

    header = ["Juez", "Correo", "Alcance evaluación", "Categoría", "Proyecto", "Estado"]
    col_widths = [4*cm, 5.5*cm, 3.5*cm, 2.8*cm, 9*cm, 2.5*cm]
    table_data = [header]
    for r in rows:
        table_data.append([
            Paragraph(_pdf_normalize_text(r["judge_name"]), cell_style),
            Paragraph(r["judge_email"], cell_style),
            Paragraph(_pdf_normalize_text(r["eval_scope"]), cell_style),
            Paragraph(_pdf_normalize_text(r["project_category"]), cell_style),
            Paragraph(_pdf_normalize_text(r["project_title"]), cell_style),
            Paragraph(_pdf_normalize_text(r["status"]), cell_style),
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a4a7a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f8ff")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c8ddf0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="asignaciones_por_juez.pdf",
    )


@admin_module_required("documents")
def participation_certificates_preview():
    certificate_context = _build_participation_certificate_context()
    context = _base_context("documents")
    context.update(certificate_context)
    return render_template("admin/certificates_participation.html", **context)


@admin_module_required("documents")
def participation_certificates_pdf():
    certificate_context = _build_participation_certificate_context()
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.participation_certificates_preview"))
    pdf_bytes = _render_participation_certificates_pdf(certificate_context)
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=False,
        download_name="certificados_participacion_proyectos_activos.pdf",
    )


@admin_module_required("documents")
def participation_certificates_download():
    certificate_context = _build_participation_certificate_context()
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.participation_certificates_preview"))
    pdf_bytes = _render_participation_certificates_pdf(certificate_context)
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="certificados_participacion_proyectos_activos.pdf",
    )


@admin_module_required("documents")
def project_certificates_preview(project_id: int):
    certificate_context = _build_participation_certificate_context(project_id)
    if not certificate_context["projects"]:
        abort(404)
    context = _base_context("documents")
    context.update(certificate_context)
    return render_template("admin/certificates_participation.html", **context)


@admin_module_required("documents")
def project_certificates_pdf(project_id: int):
    certificate_context = _build_participation_certificate_context(project_id)
    if not certificate_context["projects"]:
        abort(404)
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.project_certificates_preview", project_id=project_id))
    pdf_bytes = _render_participation_certificates_pdf(certificate_context)
    safe_name = _normalize_code(certificate_context["single_project"].title) or f"proyecto_{project_id}"
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=f"certificados_{safe_name}.pdf",
    )


@admin_module_required("documents")
def project_certificates_download(project_id: int):
    certificate_context = _build_participation_certificate_context(project_id)
    if not certificate_context["projects"]:
        abort(404)
    if not REPORTLAB_AVAILABLE:
        flash("No se pudo generar PDF. Instala reportlab en el entorno.", "error")
        return redirect(url_for("admin.project_certificates_preview", project_id=project_id))
    pdf_bytes = _render_participation_certificates_pdf(certificate_context)
    safe_name = _normalize_code(certificate_context["single_project"].title) or f"proyecto_{project_id}"
    return send_file(
        pdf_bytes,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"certificados_{safe_name}.pdf",
    )


@admin_module_required("smtp")
def smtp_page():
    return _render("admin/smtp.html", "smtp")


@admin_module_required("logs")
def logs_page():
    q = (request.args.get("q", "") or "").strip()
    action = (request.args.get("action", "") or "").strip()
    entity = (request.args.get("entity", "") or "").strip()

    query = SystemAuditLog.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            (SystemAuditLog.actor_name.ilike(like))
            | (SystemAuditLog.actor_email.ilike(like))
            | (SystemAuditLog.detail.ilike(like))
        )
    if action:
        query = query.filter(SystemAuditLog.action == action)
    if entity:
        query = query.filter(SystemAuditLog.entity == entity)

    logs = query.order_by(SystemAuditLog.created_at.desc()).limit(500).all()
    actions = [row[0] for row in db.session.query(SystemAuditLog.action).distinct().order_by(SystemAuditLog.action.asc()).all()]
    entities = [row[0] for row in db.session.query(SystemAuditLog.entity).distinct().order_by(SystemAuditLog.entity.asc()).all()]

    context = _base_context("logs")
    context.update(
        {
            "logs": logs,
            "audit_actions": actions,
            "audit_entities": entities,
            "filter_q": q,
            "filter_action": action,
            "filter_entity": entity,
        }
    )
    return render_template("admin/logs.html", **context)


@admin_module_required("campaigns")
def campaigns_page():
    return _render("admin/campaigns.html", "campaigns")


@admin_module_required("institution")
def institution_page():
    return _render("admin/institution.html", "institution")


@admin_module_required("maintenance")
def maintenance_page():
    return _render("admin/maintenance.html", "maintenance")


@admin_module_required("database")
def database_page():
    return _render("admin/database.html", "database")


@admin_module_required("database")
def database_backup_download(filename: str):
    try:
        safe_filename = _safe_backup_filename(filename)
    except ValueError:
        abort(404)
    backup_path = _database_backup_dir() / safe_filename
    return send_file(
        backup_path,
        mimetype="application/sql",
        as_attachment=True,
        download_name=safe_filename,
    )


@admin_module_required("gitops")
def gitops_page():
    return _render("admin/gitops.html", "gitops")


def _build_students_stats(context: dict) -> dict:
    from collections import Counter, defaultdict
    active_projects = [p for p in context.get("projects", []) if p.is_active]
    active_ids = {p.id for p in active_projects}
    members = [m for p in active_projects for m in p.members]

    total = len(members)
    gender_counts = Counter((m.gender or "No indicado").lower() for m in members)
    male   = gender_counts.get("masculino", 0)
    female = gender_counts.get("femenino", 0)
    other  = total - male - female
    scholarship = sum(1 for m in members if m.has_dining_scholarship)
    english     = sum(1 for m in members if m.participates_in_english)

    # by category
    cat_counts: dict[str, int] = defaultdict(int)
    cat_english: dict[str, int] = defaultdict(int)
    for p in active_projects:
        cat = (p.category or "Sin categoría").capitalize()
        for m in p.members:
            cat_counts[cat] += 1
            if m.participates_in_english:
                cat_english[cat] += 1

    # by specialty (top 8)
    spec_counts = Counter(
        (m.specialty or "Sin especialidad").strip() for m in members
    )
    top_specs = spec_counts.most_common(8)

    # by section
    section_counts = Counter(
        (m.section_name or "Sin sección").strip() for m in members
    )
    top_sections = sorted(section_counts.items(), key=lambda x: x[0])

    # team size distribution
    team_sizes = Counter(len(p.members) for p in active_projects)
    avg_team = round(total / len(active_projects), 1) if active_projects else 0

    # projects by logistics status
    logistics_labels = {
        "pendiente_revision": "Pendiente revisión",
        "en_revision": "En revisión",
        "aprobado": "Aprobado",
        "rechazado": "Rechazado",
    }
    logistics_counts = Counter(p.logistics_status or "pendiente_revision" for p in active_projects)

    # projects by section (derived from first member's section_name)
    proj_section_counts = Counter(
        (p.members[0].section_name if p.members and p.members[0].section_name else "Sin sección").strip()
        for p in active_projects
    )
    proj_sections_sorted = sorted(proj_section_counts.items(), key=lambda x: x[0])

    # projects by specialty (from project.specialty_ref relationship or specialty text)
    proj_specialty_counts = Counter(
        (p.specialty_ref.name if p.specialty_ref else (p.specialty or "Sin especialidad")).strip()
        for p in active_projects
    )
    top_proj_specialties = proj_specialty_counts.most_common(8)

    # projects by advisor
    advisor_proj: dict[str, int] = defaultdict(int)
    for p in active_projects:
        raw = (p.advisor_name or "Sin tutor").strip()
        key = " ".join(w.capitalize() for w in raw.split())
        advisor_proj[key] += 1
    top_advisors = sorted(advisor_proj.items(), key=lambda x: -x[1])[:10]

    return {
        "total": total,
        "male": male,
        "female": female,
        "other_gender": other,
        "scholarship": scholarship,
        "scholarship_pct": round(scholarship / total * 100) if total else 0,
        "english": english,
        "english_pct": round(english / total * 100) if total else 0,
        "total_projects": len(active_projects),
        "avg_team": avg_team,
        "total_sections": len(proj_section_counts),
        "total_advisors": len(advisor_proj),
        "gender_chart": {
            "labels": ["Masculino", "Femenino", "No indicado"],
            "data": [male, female, other],
        },
        "category_chart": {
            "labels": list(cat_counts.keys()),
            "data": list(cat_counts.values()),
            "english": [cat_english.get(c, 0) for c in cat_counts],
        },
        "specialty_chart": {
            "labels": [s[0] for s in top_specs],
            "data": [s[1] for s in top_specs],
        },
        "section_chart": {
            "labels": [s[0] for s in top_sections],
            "data": [s[1] for s in top_sections],
        },
        "team_size_chart": {
            "labels": [f"{k} integrante{'s' if k != 1 else ''}" for k in sorted(team_sizes)],
            "data": [team_sizes[k] for k in sorted(team_sizes)],
        },
        "logistics_chart": {
            "labels": [logistics_labels.get(k, k) for k in logistics_counts],
            "data": list(logistics_counts.values()),
        },
        "proj_section_chart": {
            "labels": [s[0] for s in proj_sections_sorted],
            "data": [s[1] for s in proj_sections_sorted],
        },
        "proj_specialty_chart": {
            "labels": [s[0] for s in top_proj_specialties],
            "data": [s[1] for s in top_proj_specialties],
        },
        "advisor_chart": {
            "labels": [a[0] for a in top_advisors],
            "data": [a[1] for a in top_advisors],
        },
    }


@admin_module_required("students_stats")
def students_stats_page():
    context = _base_context("students_stats")
    context["stats"] = _build_students_stats(context)
    return render_template("admin/students_stats.html", **context)


def _superadmin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapped(*args, **kwargs):
        if not current_user.has_admin_access or not current_user.is_superadmin:
            flash("Solo el superadministrador puede acceder a esta sección.", "error")
            return redirect(url_for("admin.overview"))
        return view_func(*args, **kwargs)
    wrapped.__name__ = view_func.__name__
    return wrapped


def _pip_list_installed() -> list[dict]:
    import sys
    try:
        result = subprocess.run(
            [sys.executable, "-m", "pip", "list", "--format=freeze"],
            capture_output=True, text=True, timeout=30,
        )
        packages = []
        for line in result.stdout.strip().splitlines():
            if "==" in line:
                name, version = line.split("==", 1)
                packages.append({"name": name.strip(), "version": version.strip()})
        return sorted(packages, key=lambda p: p["name"].lower())
    except Exception:
        return []


def _requirements_packages() -> list[dict]:
    req_path = Path(current_app.root_path).parent / "requirements.txt"
    packages = []
    if req_path.exists():
        for line in req_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                if "==" in line:
                    name, version = line.split("==", 1)
                    packages.append({"name": name.strip(), "version": version.strip(), "spec": line})
                else:
                    packages.append({"name": line, "version": "", "spec": line})
    return packages


@_superadmin_required
def dependencies_page():
    installed = _pip_list_installed()
    installed_map = {p["name"].lower(): p["version"] for p in installed}
    requirements = _requirements_packages()
    for req in requirements:
        req["installed_version"] = installed_map.get(req["name"].lower(), "")
        req["up_to_date"] = req["installed_version"] == req["version"] if req["version"] else bool(req["installed_version"])
    last_output = session.pop("dep_last_output", None)
    last_status = session.pop("dep_last_status", None)
    context = _base_context("dependencies")
    context.update({
        "installed_packages": installed,
        "requirements": requirements,
        "last_output": last_output,
        "last_status": last_status,
    })
    return render_template("admin/dependencies.html", **context)


def judge_form_webhook():
    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({"ok": False, "error": "El cuerpo debe ser JSON."}), 400

    if SystemSetting.get_value("judge_form_enabled", "0") != "1":
        log_event("forms.judge_access.blocked", "judge", detail="Webhook deshabilitado")
        db.session.commit()
        return jsonify({"ok": False, "error": "Webhook deshabilitado."}), 403

    expected_secret = _get_judge_form_secret()
    received_secret = _request_judge_form_token(payload)
    if not expected_secret or not hmac.compare_digest(expected_secret, received_secret):
        log_event("forms.judge_access.invalid_token", "judge", detail="Solicitud con token invalido")
        db.session.commit()
        return jsonify({"ok": False, "error": "Token invalido."}), 401

    judge, temporary_password, error = _create_or_update_judge_from_form(payload)
    if error:
        log_event("forms.judge_access.invalid_payload", "judge", detail=error)
        db.session.commit()
        return jsonify({"ok": False, "error": error}), 400

    return jsonify(
        {
            "ok": True,
            "judge_id": judge.id,
            "email": judge.email,
            "created_password": bool(temporary_password),
            "must_change_password": bool(judge.must_change_password),
        }
    )


JUDGE_CAPTCHA_QUESTION_KEY = "judge_registration_captcha_question"
JUDGE_CAPTCHA_ANSWER_KEY = "judge_registration_captcha_answer"


def _new_judge_registration_captcha():
    left = secrets.randbelow(8) + 2
    right = secrets.randbelow(8) + 2
    session[JUDGE_CAPTCHA_QUESTION_KEY] = f"{left} + {right}"
    session[JUDGE_CAPTCHA_ANSWER_KEY] = str(left + right)
    session.modified = True
    return session[JUDGE_CAPTCHA_QUESTION_KEY]


def _judge_registration_captcha_question():
    return session.get(JUDGE_CAPTCHA_QUESTION_KEY) or _new_judge_registration_captcha()


def _validate_judge_registration_captcha(answer):
    expected = str(session.get(JUDGE_CAPTCHA_ANSWER_KEY, "")).strip()
    received = str(answer or "").strip()
    valid = bool(expected) and received == expected
    _new_judge_registration_captcha()
    return valid


def public_judge_registration():
    if request.method == "POST":
        if request.form.get("website", "").strip():
            return redirect(url_for("public.judge_registration"))

        payload = request.form.to_dict()
        if not _validate_judge_registration_captcha(payload.get("captcha_answer")):
            flash("La verificación anti-spam no coincide. Inténtalo de nuevo.", "error")
            return render_template(
                "public/judge_registration.html",
                form_data=payload,
                captcha_question=_judge_registration_captcha_question(),
            )

        judge, temporary_password, error = _create_or_update_judge_from_form(payload)
        if error:
            flash(error, "error")
            return render_template(
                "public/judge_registration.html",
                form_data=payload,
                captcha_question=_judge_registration_captcha_question(),
            )

        email_sent = bool(temporary_password) and bool(getattr(judge, "_credentials_email_sent", False))
        return render_template(
            "public/judge_registration_success.html",
            judge=judge,
            credentials_generated=bool(temporary_password),
            temporary_password=temporary_password if not email_sent else "",
            email_sent=email_sent,
        )

    return render_template(
        "public/judge_registration.html",
        form_data={},
        captcha_question=_judge_registration_captcha_question(),
    )
